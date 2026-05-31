from datetime import datetime, timedelta, date
from collections import defaultdict
import csv
import json
import re
from urllib.parse import urlsplit
from io import BytesIO, StringIO, TextIOWrapper
from flask import Blueprint, Response, jsonify, render_template, redirect, url_for, flash, request
from flask_login import login_required, current_user
from werkzeug.security import generate_password_hash
from sqlalchemy import func, or_, and_
from openpyxl import Workbook, load_workbook
from app.extensions import db
from app.decorators import role_required
from app.services.majlis_enrollment_service import ensure_majlis_participant_acceptance, list_active_majlis_participants
from app.services.rumah_quran_service import (
    apply_rumah_quran_student_filter,
    assign_student_rumah_quran_class,
    ensure_rumah_quran_program_group,
    get_student_rumah_quran_classroom,
    is_rumah_quran_classroom,
    list_rumah_quran_classes,
    list_rumah_quran_students_for_class,
)
from app.services.bahasa_service import (
    apply_bahasa_student_filter,
    assign_student_bahasa_class,
    ensure_bahasa_program_group,
    get_student_bahasa_classroom,
    is_bahasa_classroom,
    list_bahasa_classes,
    list_bahasa_students_for_class,
)
from app.services.formal_service import (
    ensure_formal_program_group,
    list_formal_students_for_class,
    sync_student_formal_class_membership,
)
from app.services.staff_assignment_service import (
    cleanup_rumah_quran_subject_data,
    display_assignment_role,
    ensure_assignment_label_configs,
    sync_class_homeroom_assignment,
)
from app.services.ppdb_fee_service import (
    build_candidate_fee_drafts,
)
from app.services.ppdb_config_service import (
    list_active_ppdb_document_requirements,
    list_active_ppdb_form_fields,
)
from app.routes.ppdb_config_views import ppdb_form_builder_view, ppdb_settings_view
from app.services.finance_posting_service import (
    create_cash_bank_transaction,
    post_invoice_payment,
    post_journal,
    post_savings_transaction,
    reverse_journal,
)
from app.services.grade_formula_service import (
    REPORT_ADJUSTMENT_STATUS_ACTIVE,
    REPORT_ADJUSTMENT_STATUS_VOID,
    calculate_report_final_detail,
)
from app.utils.timezone import local_day_bounds_utc_naive, local_now, local_today
from app.forms import StudentForm, FeeTypeForm  # Pastikan Anda punya form untuk Guru/Mapel nanti
from app.models import (
    # Base & Enums
    UserRole, Gender, AttendanceStatus, PaymentStatus, RegistrationStatus, ProgramType, EducationLevel, AssignmentRole, TenantStatus, BehaviorReportType, ParticipantType,
    # Users
    User, UserRoleAssignment, Student, Parent, Teacher, Staff, MajlisParticipant, BoardingGuardian, Tenant,
    # Academic
    AcademicYear, ClassRoom, Subject, Schedule, Program, ProgramGroup, StaffAssignment,
    # Finance
    FeeType, Invoice, Transaction,
    FinanceAccount, FinanceAccountCategory, FinanceNormalBalance,
    FinanceSetting, FinanceAccountingBasis,
    FinancePeriod, FinancePeriodStatus,
    FinanceCashBankAccount, FinanceCashBankAccountType,
    FinanceCashBankTransaction, FinanceCashBankTransactionType,
    FinanceJournal, FinanceJournalLine, FinanceJournalStatus, FinanceJournalSourceType, FinanceEntrySide,
    SavingsTransactionType, SavingsTransactionStatus, StudentSavingsTransaction,
    # Student Related
    StudentClassHistory, Attendance, BoardingAttendance, Grade, ReportCard, ReportScoreAdjustment, StudentAttitude,
    Violation, BehaviorReport, TahfidzRecord, TahfidzSummary, RecitationRecord, TahfidzEvaluation,
    student_extracurriculars, StudentSavingsAccount,
    # User/System Related
    Announcement, AnnouncementRead, NotificationQueue, AuditLog, BoardingDormitory, BoardingActivitySchedule,
    # Activities
    Extracurricular,
    # PPDB
    StudentCandidate,
    # Config
    AppConfig
)
from app.utils.nis import generate_nip, generate_nis
from app.utils.roles import validate_role_combination, role_label, ROLE_PRIORITY
from app.utils.money import to_rupiah_int
from app.utils.invoice import generate_invoice_number
from app.utils.tenant_modules import (
    PACKAGE_FULL,
    PACKAGE_RUMAH_QURAN,
    PACKAGE_SEKOLAH,
    PACKAGE_OPTIONS,
    TENANT_PACKAGE_KEY,
    get_tenant_package,
    normalize_tenant_package,
)
from app.utils.tenant import (
    classroom_in_tenant,
    resolve_tenant_id,
    scoped_classrooms_query,
)

admin_bp = Blueprint('admin', __name__)


def _current_tenant_id():
    return resolve_tenant_id(current_user)


def _loads_object(raw_value):
    if not raw_value:
        return {}
    try:
        parsed = json.loads(raw_value)
    except (TypeError, ValueError):
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _tenant_teachers_query(tenant_id):
    return Teacher.query.join(User, Teacher.user_id == User.id).filter(
        Teacher.is_deleted.is_(False),
        User.tenant_id == tenant_id,
    )


def _safe_students_list_return_url(next_url, fallback_endpoint='admin.list_students'):
    fallback_url = url_for(fallback_endpoint)
    if not next_url:
        return fallback_url

    parsed = urlsplit(next_url)
    if parsed.scheme or parsed.netloc:
        return fallback_url

    allowed_paths = {url_for('admin.list_students'), url_for('staff.list_students')}
    if parsed.path not in allowed_paths:
        return fallback_url

    return next_url


def _infer_user_display_name(user):
    if user.teacher_profile and user.teacher_profile.full_name:
        return user.teacher_profile.full_name
    if user.staff_profile and user.staff_profile.full_name:
        return user.staff_profile.full_name
    if user.parent_profile and user.parent_profile.full_name:
        return user.parent_profile.full_name
    if user.student_profile and user.student_profile.full_name:
        return user.student_profile.full_name
    if user.majlis_profile and user.majlis_profile.full_name:
        return user.majlis_profile.full_name
    if user.boarding_guardian_profile and user.boarding_guardian_profile.full_name:
        return user.boarding_guardian_profile.full_name
    return user.username


def _infer_user_phone(user):
    if user.teacher_profile and user.teacher_profile.phone:
        return user.teacher_profile.phone
    if user.parent_profile and user.parent_profile.phone:
        return user.parent_profile.phone
    if user.boarding_guardian_profile and user.boarding_guardian_profile.phone:
        return user.boarding_guardian_profile.phone
    if user.majlis_profile and user.majlis_profile.phone:
        return user.majlis_profile.phone
    return None


def _parse_tenant_status(raw_value):
    if not raw_value:
        return None
    normalized = str(raw_value).strip()
    if not normalized:
        return None
    try:
        return TenantStatus[normalized.upper()]
    except KeyError:
        pass
    for item in TenantStatus:
        if item.value.lower() == normalized.lower():
            return item
    return None


def _upsert_tenant_config(tenant_id, key, value, description):
    row = AppConfig.query.filter_by(tenant_id=tenant_id, key=key).first()
    clean_value = (value or '').strip()
    if row:
        row.value = clean_value
        row.description = description
    else:
        db.session.add(AppConfig(
            tenant_id=tenant_id,
            key=key,
            value=clean_value,
            description=description,
        ))


def _slugify_tenant(raw_value):
    slug = re.sub(r"[^a-z0-9]+", "-", (raw_value or "").strip().lower()).strip("-")
    return slug or "tenant"


def _normalize_tenant_code(raw_value):
    code = re.sub(r"[^A-Za-z0-9_-]+", "", (raw_value or "").strip().upper())
    return code


def _parse_iso_date(value):
    raw = (value or '').strip()
    if not raw:
        return None
    try:
        return datetime.strptime(raw, '%Y-%m-%d').date()
    except ValueError:
        return None


def _iter_upload_rows(file):
    def _normalize_cell(value):
        if value is None:
            return ''
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return str(value).strip()
        if isinstance(value, int):
            return str(value)
        return str(value).strip()

    filename = (file.filename or "").lower()
    if filename.endswith('.xlsx'):
        workbook = load_workbook(file, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell).strip() if cell is not None else '' for cell in rows[0]]
        parsed = []
        for idx, row in enumerate(rows[1:], start=2):
            row_data = {}
            for col_idx, header in enumerate(headers):
                value = row[col_idx] if col_idx < len(row) else None
                row_data[header] = _normalize_cell(value)
            parsed.append((idx, row_data))
        return parsed

    wrapper = TextIOWrapper(file.stream, encoding='utf-8-sig')
    reader = csv.DictReader(wrapper)
    return [(idx, {k: (v.strip() if isinstance(v, str) else '' if v is None else str(v).strip())
                   for k, v in row.items()})
            for idx, row in enumerate(reader, start=2)]


# =========================================================
# 1. DASHBOARD & KONFIGURASI SISTEM
# =========================================================

@admin_bp.route('/dashboard')
@login_required
@role_required(UserRole.ADMIN)
def dashboard():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    # 1. Hitung Total Siswa & Guru
    total_students = (
        Student.query.join(User, Student.user_id == User.id)
        .filter(
            Student.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .count()
    )
    total_teachers = (
        Teacher.query.join(User, Teacher.user_id == User.id)
        .filter(
            Teacher.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .count()
    )

    # 2. Hitung Pemasukan Hari Ini (PENTING!)
    start_utc, end_utc = local_day_bounds_utc_naive()
    income_today = (
        db.session.query(func.sum(Transaction.amount))
        .join(Invoice, Invoice.id == Transaction.invoice_id)
        .join(Student, Student.id == Invoice.student_id)
        .join(User, User.id == Student.user_id)
        .filter(
            Transaction.date >= start_utc,
            Transaction.date < end_utc,
            Invoice.is_deleted.is_(False),
            Student.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .scalar()
        or 0
    )  # <--- "or 0" penting agar tidak None

    # 3. Kirim variabel ke HTML (income_today wajib ada)
    return render_template('admin/dashboard.html',
                           total_students=total_students,
                           total_teachers=total_teachers,
                           income_today=income_today)  # <--- JANGAN LUPA INI


def _format_chart_date(day):
    return day.strftime('%d/%m')


def _shift_month(month_start, offset):
    month_index = month_start.month - 1 + offset
    year = month_start.year + month_index // 12
    month = month_index % 12 + 1
    return date(year, month, 1)


def _next_month(month_start):
    return _shift_month(month_start, 1)


def _leadership_period(raw_period):
    today = local_today()
    options = {
        '7d': ('7 Hari', today - timedelta(days=6), today),
        '14d': ('14 Hari', today - timedelta(days=13), today),
        '30d': ('30 Hari', today - timedelta(days=29), today),
        '90d': ('90 Hari', today - timedelta(days=89), today),
        'month': ('Bulan Ini', date(today.year, today.month, 1), today),
        'year': ('Tahun Ini', date(today.year, 1, 1), today),
    }
    key = raw_period if raw_period in options else '14d'
    label, start_date, end_date = options[key]
    return key, label, start_date, end_date, [{'key': item_key, 'label': item[0]} for item_key, item in options.items()]


def _date_series(start_date, end_date):
    days = (end_date - start_date).days
    return [start_date + timedelta(days=offset) for offset in range(days + 1)]


def _month_series(start_date, end_date):
    cursor = date(start_date.year, start_date.month, 1)
    end_month = date(end_date.year, end_date.month, 1)
    months = []
    while cursor <= end_month:
        months.append(cursor)
        cursor = _next_month(cursor)
    return months


def _finance_amount_for_categories(tenant_id, start_date, end_date, categories):
    rows = (
        db.session.query(
            FinanceAccount.category,
            FinanceJournalLine.entry_side,
            func.sum(FinanceJournalLine.amount),
        )
        .join(FinanceJournal, FinanceJournal.id == FinanceJournalLine.journal_id)
        .join(FinanceAccount, FinanceAccount.id == FinanceJournalLine.account_id)
        .filter(
            FinanceJournalLine.tenant_id == tenant_id,
            FinanceJournal.tenant_id == tenant_id,
            FinanceJournal.status == FinanceJournalStatus.POSTED,
            FinanceJournal.journal_date >= start_date,
            FinanceJournal.journal_date <= end_date,
            FinanceAccount.category.in_(categories),
        )
        .group_by(FinanceAccount.category, FinanceJournalLine.entry_side)
        .all()
    )
    totals = {category: {'debit': 0, 'credit': 0} for category in categories}
    for category, entry_side, amount in rows:
        key = 'debit' if entry_side == FinanceEntrySide.DEBIT else 'credit'
        totals[category][key] += int(amount or 0)
    return totals


def _leadership_analytics_payload(tenant_id, period_key='14d'):
    today = local_today()
    period_key, period_label, period_start, period_end, period_options = _leadership_period(period_key)
    start_utc, end_utc = local_day_bounds_utc_naive()

    total_students = (
        Student.query.join(User, Student.user_id == User.id)
        .filter(Student.is_deleted.is_(False), User.tenant_id == tenant_id)
        .count()
    )
    total_teachers = (
        Teacher.query.join(User, Teacher.user_id == User.id)
        .filter(Teacher.is_deleted.is_(False), User.tenant_id == tenant_id)
        .count()
    )
    total_staff = (
        Staff.query.join(User, Staff.user_id == User.id)
        .filter(Staff.is_deleted.is_(False), User.tenant_id == tenant_id)
        .count()
    )

    income_today = (
        db.session.query(func.sum(Transaction.amount))
        .join(Invoice, Invoice.id == Transaction.invoice_id)
        .join(Student, Student.id == Invoice.student_id)
        .join(User, User.id == Student.user_id)
        .filter(
            Transaction.date >= start_utc,
            Transaction.date < end_utc,
            Invoice.is_deleted.is_(False),
            Student.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .scalar()
        or 0
    )
    receivable_total = (
        db.session.query(func.sum(Invoice.total_amount - Invoice.paid_amount))
        .join(Student, Student.id == Invoice.student_id)
        .join(User, User.id == Student.user_id)
        .filter(
            Invoice.is_deleted.is_(False),
            Student.is_deleted.is_(False),
            User.tenant_id == tenant_id,
            Invoice.status != PaymentStatus.PAID,
        )
        .scalar()
        or 0
    )

    finance_totals = _finance_amount_for_categories(
        tenant_id,
        period_start,
        period_end,
        [FinanceAccountCategory.REVENUE, FinanceAccountCategory.EXPENSE],
    )
    monthly_revenue = (
        finance_totals[FinanceAccountCategory.REVENUE]['credit']
        - finance_totals[FinanceAccountCategory.REVENUE]['debit']
    )
    monthly_expense = (
        finance_totals[FinanceAccountCategory.EXPENSE]['debit']
        - finance_totals[FinanceAccountCategory.EXPENSE]['credit']
    )

    attendance_counts = {
        status.value: Attendance.query.join(Student, Student.id == Attendance.student_id)
        .join(User, User.id == Student.user_id)
        .filter(
            Attendance.date == today,
            Attendance.status == status,
            Attendance.student_id.isnot(None),
            Student.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .count()
        for status in AttendanceStatus
    }
    attendance_total = sum(attendance_counts.values())
    attendance_present = attendance_counts.get(AttendanceStatus.HADIR.value, 0)
    attendance_rate = round((attendance_present / attendance_total) * 100, 1) if attendance_total else 0

    attendance_trend_rows = (
        db.session.query(Attendance.date, Attendance.status, func.count(Attendance.id))
        .join(Student, Student.id == Attendance.student_id)
        .join(User, User.id == Student.user_id)
        .filter(
            Attendance.date >= period_start,
            Attendance.date <= period_end,
            Attendance.student_id.isnot(None),
            Student.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .group_by(Attendance.date, Attendance.status)
        .all()
    )
    attendance_by_day = {}
    for day, status, count in attendance_trend_rows:
        day_key = day.isoformat()
        attendance_by_day.setdefault(day_key, {'total': 0, 'present': 0})
        attendance_by_day[day_key]['total'] += int(count or 0)
        if status == AttendanceStatus.HADIR:
            attendance_by_day[day_key]['present'] += int(count or 0)
    attendance_trend = []
    for day in _date_series(period_start, period_end):
        day_data = attendance_by_day.get(day.isoformat(), {'total': 0, 'present': 0})
        rate = round((day_data['present'] / day_data['total']) * 100, 1) if day_data['total'] else 0
        attendance_trend.append({
            'label': _format_chart_date(day),
            'present': day_data['present'],
            'total': day_data['total'],
            'rate': rate,
        })

    payment_rows = (
        db.session.query(func.date(Transaction.date), func.sum(Transaction.amount))
        .join(Invoice, Invoice.id == Transaction.invoice_id)
        .join(Student, Student.id == Invoice.student_id)
        .join(User, User.id == Student.user_id)
        .filter(
            Transaction.date >= datetime.combine(period_start, datetime.min.time()),
            Transaction.date < datetime.combine(period_end + timedelta(days=1), datetime.min.time()),
            Invoice.is_deleted.is_(False),
            Student.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .group_by(func.date(Transaction.date))
        .all()
    )
    payments_by_day = {str(day): int(amount or 0) for day, amount in payment_rows}
    payment_trend = []
    for day in _date_series(period_start, period_end):
        payment_trend.append({
            'label': _format_chart_date(day),
            'amount': payments_by_day.get(day.isoformat(), 0),
        })

    finance_trend = []
    for trend_month_start in _month_series(period_start, period_end):
        trend_month_end = min(_next_month(trend_month_start) - timedelta(days=1), period_end)
        trend_totals = _finance_amount_for_categories(
            tenant_id,
            max(trend_month_start, period_start),
            min(trend_month_end, today),
            [FinanceAccountCategory.REVENUE, FinanceAccountCategory.EXPENSE],
        )
        revenue = (
            trend_totals[FinanceAccountCategory.REVENUE]['credit']
            - trend_totals[FinanceAccountCategory.REVENUE]['debit']
        )
        expense = (
            trend_totals[FinanceAccountCategory.EXPENSE]['debit']
            - trend_totals[FinanceAccountCategory.EXPENSE]['credit']
        )
        finance_trend.append({
            'label': trend_month_start.strftime('%b %Y'),
            'revenue': int(revenue),
            'expense': int(expense),
            'surplus': int(revenue - expense),
        })

    invoice_status_rows = (
        db.session.query(
            Invoice.status,
            func.count(Invoice.id),
            func.sum(Invoice.total_amount),
            func.sum(Invoice.paid_amount),
        )
        .join(Student, Student.id == Invoice.student_id)
        .join(User, User.id == Student.user_id)
        .filter(
            Invoice.is_deleted.is_(False),
            Student.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .group_by(Invoice.status)
        .all()
    )
    invoice_status = []
    for status, count, total_amount, paid_amount in invoice_status_rows:
        total_amount = int(total_amount or 0)
        paid_amount = int(paid_amount or 0)
        invoice_status.append({
            'label': status.value if status else '-',
            'count': int(count or 0),
            'total_amount': total_amount,
            'paid_amount': paid_amount,
            'outstanding_amount': max(total_amount - paid_amount, 0),
        })

    class_overview = []
    for class_room in scoped_classrooms_query(tenant_id).order_by(ClassRoom.name.asc()).limit(8).all():
        student_count = (
            Student.query.join(User, Student.user_id == User.id)
            .filter(
                Student.current_class_id == class_room.id,
                Student.is_deleted.is_(False),
                User.tenant_id == tenant_id,
            )
            .count()
        )
        today_attendance = (
            Attendance.query.join(Student, Student.id == Attendance.student_id)
            .join(User, User.id == Student.user_id)
            .filter(
                Attendance.class_id == class_room.id,
                Attendance.date == today,
                Attendance.student_id.isnot(None),
                Student.is_deleted.is_(False),
                User.tenant_id == tenant_id,
            )
            .count()
        )
        class_overview.append({
            'name': class_room.name,
            'student_count': student_count,
            'attendance_count': today_attendance,
        })

    ppdb_pending = StudentCandidate.query.filter_by(
        tenant_id=tenant_id,
        status=RegistrationStatus.PENDING,
    ).count()
    behavior_open = (
        BehaviorReport.query.join(Student, Student.id == BehaviorReport.student_id)
        .join(User, User.id == Student.user_id)
        .filter(
            BehaviorReport.is_resolved.is_(False),
            Student.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .count()
    )
    behavior_rows = (
        db.session.query(BehaviorReport.report_type, func.count(BehaviorReport.id))
        .join(Student, Student.id == BehaviorReport.student_id)
        .join(User, User.id == Student.user_id)
        .filter(
            BehaviorReport.report_date >= period_start,
            BehaviorReport.report_date <= period_end,
            Student.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .group_by(BehaviorReport.report_type)
        .all()
    )
    behavior_summary = [
        {
            'label': report_type.value if report_type else '-',
            'count': int(count or 0),
        }
        for report_type, count in behavior_rows
    ]
    ppdb_status_rows = (
        db.session.query(StudentCandidate.status, func.count(StudentCandidate.id))
        .filter(
            StudentCandidate.tenant_id == tenant_id,
            StudentCandidate.created_at >= datetime.combine(period_start, datetime.min.time()),
            StudentCandidate.created_at < datetime.combine(period_end + timedelta(days=1), datetime.min.time()),
        )
        .group_by(StudentCandidate.status)
        .all()
    )
    ppdb_status = [
        {
            'label': status.value if status else '-',
            'count': int(count or 0),
        }
        for status, count in ppdb_status_rows
    ]
    draft_journals = FinanceJournal.query.filter_by(
        tenant_id=tenant_id,
        status=FinanceJournalStatus.DRAFT,
    ).count()
    period_today = FinancePeriod.query.filter(
        FinancePeriod.tenant_id == tenant_id,
        FinancePeriod.start_date <= today,
        FinancePeriod.end_date >= today,
    ).first()

    recent_payments = (
        Transaction.query.join(Invoice, Invoice.id == Transaction.invoice_id)
        .join(Student, Student.id == Invoice.student_id)
        .join(User, User.id == Student.user_id)
        .filter(
            Invoice.is_deleted.is_(False),
            Student.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .order_by(Transaction.date.desc(), Transaction.id.desc())
        .limit(5)
        .all()
    )

    return {
        'generated_at': local_now().strftime('%d/%m/%Y %H:%M:%S'),
        'period': {
            'key': period_key,
            'label': period_label,
            'start_date': period_start.isoformat(),
            'end_date': period_end.isoformat(),
            'options': period_options,
        },
        'cards': {
            'students': total_students,
            'teachers_staff': total_teachers + total_staff,
            'attendance_rate': attendance_rate,
            'income_today': int(income_today or 0),
            'receivable_total': int(receivable_total or 0),
            'monthly_surplus': int(monthly_revenue - monthly_expense),
        },
        'finance': {
            'monthly_revenue': int(monthly_revenue),
            'monthly_expense': int(monthly_expense),
            'monthly_surplus': int(monthly_revenue - monthly_expense),
            'draft_journals': draft_journals,
            'period_status': period_today.status.value if period_today else 'BELUM ADA',
        },
        'attendance': {
            'counts': attendance_counts,
            'total': attendance_total,
            'rate': attendance_rate,
        },
        'payment_trend': payment_trend,
        'attendance_trend': attendance_trend,
        'finance_trend': finance_trend,
        'invoice_status': invoice_status,
        'class_overview': class_overview,
        'behavior_summary': behavior_summary,
        'ppdb_status': ppdb_status,
        'alerts': {
            'ppdb_pending': ppdb_pending,
            'behavior_open': behavior_open,
            'draft_journals': draft_journals,
            'period_status': period_today.status.value if period_today else 'BELUM ADA',
        },
        'recent_payments': [
            {
                'student': payment.invoice.student.full_name if payment.invoice and payment.invoice.student else '-',
                'amount': int(payment.amount or 0),
                'method': payment.method or '-',
                'date': payment.date.strftime('%d/%m/%Y %H:%M') if payment.date else '-',
            }
            for payment in recent_payments
        ],
    }


@admin_bp.route('/dashboard/pimpinan')
@login_required
@role_required(UserRole.PIMPINAN)
def leadership_dashboard():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))
    payload = _leadership_analytics_payload(tenant_id, request.args.get('period'))
    return render_template('admin/leadership_dashboard.html', analytics=payload)


@admin_bp.route('/dashboard/pimpinan/data')
@login_required
@role_required(UserRole.PIMPINAN)
def leadership_dashboard_data():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        return jsonify({'error': 'Tenant default tidak ditemukan.'}), 404
    return jsonify(_leadership_analytics_payload(tenant_id, request.args.get('period')))


def _leadership_detail_dates():
    start_date = _parse_iso_date(request.args.get('start_date'))
    end_date = _parse_iso_date(request.args.get('end_date'))
    if start_date and end_date:
        return start_date, end_date
    _, _, period_start, period_end, _ = _leadership_period(request.args.get('period'))
    return start_date or period_start, end_date or period_end


def _enum_from_request(enum_class, raw_value):
    raw_value = (raw_value or '').strip()
    if not raw_value:
        return None
    if raw_value in enum_class.__members__:
        return enum_class[raw_value]
    for item in enum_class:
        if raw_value == item.value:
            return item
    return None


def _leadership_attendance_detail(tenant_id, start_date, end_date):
    class_rows = (
        db.session.query(ClassRoom.id, ClassRoom.name, Attendance.status, func.count(Attendance.id))
        .join(Attendance, Attendance.class_id == ClassRoom.id)
        .join(Student, Student.id == Attendance.student_id)
        .join(User, User.id == Student.user_id)
        .filter(
            Attendance.date >= start_date,
            Attendance.date <= end_date,
            Attendance.student_id.isnot(None),
            Student.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .group_by(ClassRoom.id, ClassRoom.name, Attendance.status)
        .order_by(ClassRoom.name.asc(), Attendance.status.asc())
        .all()
    )
    class_map = {}
    for class_id, class_name, status, count in class_rows:
        row = class_map.setdefault(class_id, {
            'class_id': class_id,
            'name': class_name,
            'counts': {item.value: 0 for item in AttendanceStatus},
            'total': 0,
        })
        status_label = status.value if status else '-'
        row['counts'][status_label] = int(count or 0)
        row['total'] += int(count or 0)

    boarding_rows = (
        db.session.query(
            BoardingDormitory.id,
            BoardingDormitory.name,
            BoardingActivitySchedule.id,
            BoardingActivitySchedule.activity_name,
            BoardingAttendance.status,
            func.count(BoardingAttendance.id),
        )
        .join(BoardingDormitory, BoardingDormitory.id == BoardingAttendance.dormitory_id)
        .join(BoardingActivitySchedule, BoardingActivitySchedule.id == BoardingAttendance.schedule_id)
        .join(Student, Student.id == BoardingAttendance.student_id)
        .join(User, User.id == Student.user_id)
        .filter(
            BoardingAttendance.date >= start_date,
            BoardingAttendance.date <= end_date,
            Student.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .group_by(
            BoardingDormitory.id,
            BoardingDormitory.name,
            BoardingActivitySchedule.id,
            BoardingActivitySchedule.activity_name,
            BoardingAttendance.status,
        )
        .order_by(BoardingDormitory.name.asc(), BoardingActivitySchedule.activity_name.asc())
        .all()
    )
    boarding_map = {}
    for dormitory_id, dormitory_name, schedule_id, activity_name, status, count in boarding_rows:
        key = (dormitory_id, schedule_id)
        row = boarding_map.setdefault(key, {
            'dormitory_id': dormitory_id,
            'dormitory': dormitory_name,
            'schedule_id': schedule_id,
            'activity': activity_name,
            'counts': {item.value: 0 for item in AttendanceStatus},
            'total': 0,
        })
        status_label = status.value if status else '-'
        row['counts'][status_label] = int(count or 0)
        row['total'] += int(count or 0)

    recent_records = (
        Attendance.query.join(Student, Student.id == Attendance.student_id)
        .join(User, User.id == Student.user_id)
        .filter(
            Attendance.date >= start_date,
            Attendance.date <= end_date,
            Attendance.student_id.isnot(None),
            Student.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .order_by(Attendance.date.desc(), Attendance.id.desc())
        .limit(100)
        .all()
    )

    drilldown = None
    status_filter = _enum_from_request(AttendanceStatus, request.args.get('status'))
    source_filter = (request.args.get('source') or '').strip()
    if status_filter and source_filter == 'formal':
        class_id = request.args.get('class_id', type=int)
        selected_class = ClassRoom.query.filter_by(id=class_id).first() if class_id else None
        query = (
            db.session.query(
                Attendance.date,
                Student.full_name.label('student_name'),
                ClassRoom.name.label('group_name'),
                Attendance.status,
                Attendance.notes,
            )
            .join(Student, Student.id == Attendance.student_id)
            .join(User, User.id == Student.user_id)
            .join(ClassRoom, ClassRoom.id == Attendance.class_id)
            .filter(
                Attendance.date >= start_date,
                Attendance.date <= end_date,
                Attendance.status == status_filter,
                Attendance.student_id.isnot(None),
                Student.is_deleted.is_(False),
                User.tenant_id == tenant_id,
            )
        )
        if selected_class:
            query = query.filter(Attendance.class_id == selected_class.id)
        drilldown = {
            'type': 'attendance_formal',
            'title': f"Absensi Formal {status_filter.value}" + (f" - {selected_class.name}" if selected_class else ''),
            'rows': [
                {
                    'date': row.date,
                    'student_name': row.student_name,
                    'group_name': row.group_name,
                    'activity_name': '-',
                    'status': row.status.value if row.status else '-',
                    'notes': row.notes or '-',
                }
                for row in query.order_by(Attendance.date.desc(), Student.full_name.asc()).limit(300).all()
            ],
        }
    elif status_filter and source_filter == 'boarding':
        dormitory_id = request.args.get('dormitory_id', type=int)
        schedule_id = request.args.get('schedule_id', type=int)
        query = (
            db.session.query(
                BoardingAttendance.date,
                Student.full_name.label('student_name'),
                BoardingDormitory.name.label('group_name'),
                BoardingActivitySchedule.activity_name.label('activity_name'),
                BoardingAttendance.status,
                BoardingAttendance.notes,
            )
            .join(Student, Student.id == BoardingAttendance.student_id)
            .join(User, User.id == Student.user_id)
            .join(BoardingDormitory, BoardingDormitory.id == BoardingAttendance.dormitory_id)
            .join(BoardingActivitySchedule, BoardingActivitySchedule.id == BoardingAttendance.schedule_id)
            .filter(
                BoardingAttendance.date >= start_date,
                BoardingAttendance.date <= end_date,
                BoardingAttendance.status == status_filter,
                Student.is_deleted.is_(False),
                User.tenant_id == tenant_id,
            )
        )
        selected_dormitory = BoardingDormitory.query.filter_by(id=dormitory_id).first() if dormitory_id else None
        selected_schedule = BoardingActivitySchedule.query.filter_by(id=schedule_id).first() if schedule_id else None
        if selected_dormitory:
            query = query.filter(BoardingAttendance.dormitory_id == selected_dormitory.id)
        if selected_schedule:
            query = query.filter(BoardingAttendance.schedule_id == selected_schedule.id)
        suffix = []
        if selected_dormitory:
            suffix.append(selected_dormitory.name)
        if selected_schedule:
            suffix.append(selected_schedule.activity_name)
        drilldown = {
            'type': 'attendance_boarding',
            'title': f"Absensi Asrama {status_filter.value}" + (f" - {' / '.join(suffix)}" if suffix else ''),
            'rows': [
                {
                    'date': row.date,
                    'student_name': row.student_name,
                    'group_name': row.group_name,
                    'activity_name': row.activity_name,
                    'status': row.status.value if row.status else '-',
                    'notes': row.notes or '-',
                }
                for row in query.order_by(BoardingAttendance.date.desc(), Student.full_name.asc()).limit(300).all()
            ],
        }
    return {
        'class_rows': list(class_map.values()),
        'boarding_rows': list(boarding_map.values()),
        'recent_records': recent_records,
        'drilldown': drilldown,
    }


def _leadership_finance_detail(tenant_id, start_date, end_date):
    revenue_rows, expense_rows, total_revenue, total_expense, net_income = _income_statement_data(
        tenant_id, start_date, end_date
    )
    payment_start = datetime.combine(start_date, datetime.min.time())
    payment_end = datetime.combine(end_date + timedelta(days=1), datetime.min.time())
    transactions = (
        Transaction.query.join(Invoice, Invoice.id == Transaction.invoice_id)
        .join(Student, Student.id == Invoice.student_id)
        .join(User, User.id == Student.user_id)
        .filter(
            Transaction.date >= payment_start,
            Transaction.date < payment_end,
            Invoice.is_deleted.is_(False),
            Student.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .order_by(Transaction.date.desc(), Transaction.id.desc())
        .limit(100)
        .all()
    )
    invoice_rows = (
        db.session.query(
            Invoice.status,
            func.count(Invoice.id),
            func.sum(Invoice.total_amount),
            func.sum(Invoice.paid_amount),
        )
        .join(Student, Student.id == Invoice.student_id)
        .join(User, User.id == Student.user_id)
        .filter(
            Invoice.is_deleted.is_(False),
            Student.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .group_by(Invoice.status)
        .all()
    )
    journals = (
        FinanceJournal.query.filter(
            FinanceJournal.tenant_id == tenant_id,
            FinanceJournal.journal_date >= start_date,
            FinanceJournal.journal_date <= end_date,
        )
        .order_by(FinanceJournal.journal_date.desc(), FinanceJournal.id.desc())
        .limit(100)
        .all()
    )
    drilldown = None
    drill_type = (request.args.get('drill') or '').strip()
    if drill_type == 'invoice_status':
        status_filter = _enum_from_request(PaymentStatus, request.args.get('status'))
        query = (
            Invoice.query.join(Student, Student.id == Invoice.student_id)
            .join(User, User.id == Student.user_id)
            .filter(
                Invoice.is_deleted.is_(False),
                Student.is_deleted.is_(False),
                User.tenant_id == tenant_id,
            )
        )
        if status_filter:
            query = query.filter(Invoice.status == status_filter)
        drilldown = {
            'type': 'invoice_status',
            'title': f"Tagihan {status_filter.value if status_filter else 'Semua Status'}",
            'rows': query.order_by(Invoice.due_date.asc(), Student.full_name.asc()).limit(300).all(),
        }
    elif drill_type == 'finance_account':
        account_id = request.args.get('account_id', type=int)
        account = FinanceAccount.query.filter_by(id=account_id, tenant_id=tenant_id).first() if account_id else None
        if account:
            rows = (
                _posted_finance_lines_query(tenant_id, start_date, end_date)
                .filter(FinanceJournalLine.account_id == account.id)
                .order_by(FinanceJournal.journal_date.desc(), FinanceJournalLine.id.desc())
                .limit(300)
                .all()
            )
            drilldown = {
                'type': 'finance_account',
                'title': f"{account.code} - {account.name}",
                'rows': rows,
            }
    return {
        'revenue_rows': revenue_rows,
        'expense_rows': expense_rows,
        'total_revenue': total_revenue,
        'total_expense': total_expense,
        'net_income': net_income,
        'transactions': transactions,
        'invoice_rows': invoice_rows,
        'journals': journals,
        'drilldown': drilldown,
    }


def _leadership_operations_detail(tenant_id, start_date, end_date):
    teachers = (
        Teacher.query.join(User, Teacher.user_id == User.id)
        .filter(Teacher.is_deleted.is_(False), User.tenant_id == tenant_id)
        .order_by(Teacher.full_name.asc())
        .limit(100)
        .all()
    )
    staff_members = (
        Staff.query.join(User, Staff.user_id == User.id)
        .filter(Staff.is_deleted.is_(False), User.tenant_id == tenant_id)
        .order_by(Staff.full_name.asc())
        .limit(100)
        .all()
    )
    candidates = (
        StudentCandidate.query.filter(
            StudentCandidate.tenant_id == tenant_id,
            StudentCandidate.created_at >= datetime.combine(start_date, datetime.min.time()),
            StudentCandidate.created_at < datetime.combine(end_date + timedelta(days=1), datetime.min.time()),
        )
        .order_by(StudentCandidate.created_at.desc(), StudentCandidate.id.desc())
        .limit(100)
        .all()
    )
    candidate_status_rows = (
        db.session.query(StudentCandidate.status, func.count(StudentCandidate.id))
        .filter(
            StudentCandidate.tenant_id == tenant_id,
            StudentCandidate.created_at >= datetime.combine(start_date, datetime.min.time()),
            StudentCandidate.created_at < datetime.combine(end_date + timedelta(days=1), datetime.min.time()),
        )
        .group_by(StudentCandidate.status)
        .all()
    )
    behavior_reports = (
        BehaviorReport.query.join(Student, Student.id == BehaviorReport.student_id)
        .join(User, User.id == Student.user_id)
        .filter(
            BehaviorReport.report_date >= start_date,
            BehaviorReport.report_date <= end_date,
            Student.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .order_by(BehaviorReport.report_date.desc(), BehaviorReport.id.desc())
        .limit(100)
        .all()
    )
    behavior_type_rows = (
        db.session.query(BehaviorReport.report_type, func.count(BehaviorReport.id))
        .join(Student, Student.id == BehaviorReport.student_id)
        .join(User, User.id == Student.user_id)
        .filter(
            BehaviorReport.report_date >= start_date,
            BehaviorReport.report_date <= end_date,
            Student.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .group_by(BehaviorReport.report_type)
        .all()
    )
    draft_journals = (
        FinanceJournal.query.filter_by(tenant_id=tenant_id, status=FinanceJournalStatus.DRAFT)
        .order_by(FinanceJournal.journal_date.desc(), FinanceJournal.id.desc())
        .limit(100)
        .all()
    )
    drilldown = None
    drill_type = (request.args.get('drill') or '').strip()
    if drill_type == 'ppdb_status':
        status_filter = _enum_from_request(RegistrationStatus, request.args.get('status'))
        query = StudentCandidate.query.filter(
            StudentCandidate.tenant_id == tenant_id,
            StudentCandidate.created_at >= datetime.combine(start_date, datetime.min.time()),
            StudentCandidate.created_at < datetime.combine(end_date + timedelta(days=1), datetime.min.time()),
        )
        if status_filter:
            query = query.filter(StudentCandidate.status == status_filter)
        drilldown = {
            'type': 'ppdb_status',
            'title': f"PPDB {status_filter.value if status_filter else 'Semua Status'}",
            'rows': query.order_by(StudentCandidate.created_at.desc(), StudentCandidate.full_name.asc()).limit(300).all(),
        }
    elif drill_type == 'behavior_type':
        type_filter = _enum_from_request(BehaviorReportType, request.args.get('report_type'))
        query = (
            BehaviorReport.query.join(Student, Student.id == BehaviorReport.student_id)
            .join(User, User.id == Student.user_id)
            .filter(
                BehaviorReport.report_date >= start_date,
                BehaviorReport.report_date <= end_date,
                Student.is_deleted.is_(False),
                User.tenant_id == tenant_id,
            )
        )
        if type_filter:
            query = query.filter(BehaviorReport.report_type == type_filter)
        drilldown = {
            'type': 'behavior_type',
            'title': f"Laporan Pembinaan {type_filter.value if type_filter else 'Semua Tipe'}",
            'rows': query.order_by(BehaviorReport.report_date.desc(), Student.full_name.asc()).limit(300).all(),
        }
    return {
        'teachers': teachers,
        'staff_members': staff_members,
        'candidates': candidates,
        'candidate_status_rows': candidate_status_rows,
        'behavior_reports': behavior_reports,
        'behavior_type_rows': behavior_type_rows,
        'draft_journals': draft_journals,
        'drilldown': drilldown,
    }


def _leadership_students_detail(tenant_id):
    class_rows = (
        db.session.query(ClassRoom.id, ClassRoom.name, func.count(Student.id))
        .join(Student, Student.current_class_id == ClassRoom.id)
        .join(User, User.id == Student.user_id)
        .filter(Student.is_deleted.is_(False), User.tenant_id == tenant_id)
        .group_by(ClassRoom.id, ClassRoom.name)
        .order_by(ClassRoom.name.asc())
        .all()
    )
    dormitory_rows = (
        db.session.query(BoardingDormitory.id, BoardingDormitory.name, func.count(Student.id))
        .join(Student, Student.boarding_dormitory_id == BoardingDormitory.id)
        .join(User, User.id == Student.user_id)
        .filter(Student.is_deleted.is_(False), User.tenant_id == tenant_id)
        .group_by(BoardingDormitory.id, BoardingDormitory.name)
        .order_by(BoardingDormitory.name.asc())
        .all()
    )
    students_query = Student.query.join(User, Student.user_id == User.id).filter(
        Student.is_deleted.is_(False),
        User.tenant_id == tenant_id,
    )
    drill_title = None
    class_id = request.args.get('class_id', type=int)
    dormitory_id = request.args.get('dormitory_id', type=int)
    if class_id:
        selected_class = ClassRoom.query.filter_by(id=class_id).first()
        if selected_class:
            students_query = students_query.filter(Student.current_class_id == selected_class.id)
            drill_title = f"Santri Kelas {selected_class.name}"
    if dormitory_id:
        selected_dormitory = BoardingDormitory.query.filter_by(id=dormitory_id).first()
        if selected_dormitory:
            students_query = students_query.filter(Student.boarding_dormitory_id == selected_dormitory.id)
            drill_title = f"Santri Asrama {selected_dormitory.name}"
    students = students_query.order_by(Student.full_name.asc()).limit(300).all()
    return {
        'class_rows': class_rows,
        'dormitory_rows': dormitory_rows,
        'students': students,
        'drill_title': drill_title,
    }


@admin_bp.route('/dashboard/pimpinan/detail/<section>')
@login_required
@role_required(UserRole.PIMPINAN)
def leadership_detail(section):
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    allowed_sections = {
        'students': 'Detail Santri',
        'attendance': 'Detail Absensi',
        'finance': 'Detail Keuangan',
        'operations': 'Detail Operasional',
    }
    if section not in allowed_sections:
        flash('Detail laporan tidak ditemukan.', 'warning')
        return redirect(url_for('admin.leadership_dashboard'))

    start_date, end_date = _leadership_detail_dates()
    detail_data = {}
    if section == 'students':
        detail_data = _leadership_students_detail(tenant_id)
    elif section == 'attendance':
        detail_data = _leadership_attendance_detail(tenant_id, start_date, end_date)
    elif section == 'finance':
        detail_data = _leadership_finance_detail(tenant_id, start_date, end_date)
    elif section == 'operations':
        detail_data = _leadership_operations_detail(tenant_id, start_date, end_date)

    return render_template(
        'admin/leadership_detail.html',
        section=section,
        section_title=allowed_sections[section],
        start_date=start_date,
        end_date=end_date,
        detail=detail_data,
    )


@admin_bp.route('/pengaturan/sistem', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def manage_app_config():
    """Mengelola Variable Global (Misal: Biaya Denda, Pesan Pengumuman, dll)"""
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        key = request.form.get('key')
        value = request.form.get('value')
        description = request.form.get('description')

        config = AppConfig.query.filter_by(tenant_id=tenant_id, key=key).first()
        if config:
            config.value = value
            config.description = description
        else:
            new_config = AppConfig(tenant_id=tenant_id, key=key, value=value, description=description)
            db.session.add(new_config)

        db.session.commit()
        flash('Konfigurasi tersimpan.', 'success')
        return redirect(url_for('admin.manage_app_config'))

    ensure_assignment_label_configs(tenant_id=tenant_id)
    query = (request.args.get('q') or '').strip()
    configs_query = AppConfig.query.filter(AppConfig.tenant_id == tenant_id)
    if query:
        configs_query = configs_query.filter(
            or_(
                AppConfig.key.ilike(f'%{query}%'),
                AppConfig.value.ilike(f'%{query}%'),
                AppConfig.description.ilike(f'%{query}%')
            )
        )

    configs = configs_query.order_by(AppConfig.key.asc()).all()
    return render_template('admin/system/configs.html', configs=configs, query=query)


def _student_in_tenant_query(tenant_id):
    class_ids = [row.id for row in scoped_classrooms_query(tenant_id).all()]
    if not class_ids:
        return Student.query.filter(False)
    return Student.query.filter(
        Student.is_deleted.is_(False),
        Student.current_class_id.in_(class_ids),
    )


def _calculated_final_for_adjustment(student_id, academic_year_id, subject_id, tenant_id, class_id=None):
    rows = Grade.query.filter(
        Grade.is_deleted.is_(False),
        Grade.participant_type == ParticipantType.STUDENT,
        Grade.student_id == student_id,
        Grade.academic_year_id == academic_year_id,
        Grade.subject_id == subject_id,
    ).all()
    type_scores = defaultdict(list)
    for row in rows:
        if row.type:
            type_scores[row.type.name].append(float(row.score or 0))
    type_averages = {
        type_name: round(sum(scores) / len(scores), 2)
        for type_name, scores in type_scores.items()
        if scores
    }
    detail = calculate_report_final_detail(
        type_averages,
        tenant_id=tenant_id,
        academic_year_id=academic_year_id,
        subject_id=subject_id,
        student_id=student_id,
        class_id=class_id,
    )
    return detail['final_score']


def _row_value(row, *keys):
    for key in keys:
        value = row.get(key)
        if value not in (None, ''):
            return str(value).strip()
    lower_map = {str(k).strip().lower(): v for k, v in row.items()}
    for key in keys:
        value = lower_map.get(str(key).strip().lower())
        if value not in (None, ''):
            return str(value).strip()
    return ''


def _resolve_adjustment_student(row, tenant_id):
    student_id_raw = _row_value(row, 'student_id', 'id_siswa')
    nis = _row_value(row, 'nis', 'NIS')
    full_name = _row_value(row, 'nama', 'nama_siswa', 'full_name')

    query = _student_in_tenant_query(tenant_id)
    if student_id_raw:
        try:
            student_id = int(student_id_raw)
        except ValueError:
            return None, f'ID siswa tidak valid: {student_id_raw}'
        student = query.filter(Student.id == student_id).first()
        return (student, None) if student else (None, f'Siswa ID {student_id} tidak ditemukan di tenant ini.')

    if nis:
        student = query.filter(Student.nis == nis).first()
        return (student, None) if student else (None, f'NIS {nis} tidak ditemukan di tenant ini.')

    if full_name:
        matches = query.filter(func.lower(Student.full_name) == full_name.lower()).all()
        if len(matches) == 1:
            return matches[0], None
        if len(matches) > 1:
            return None, f'Nama siswa "{full_name}" duplikat, gunakan NIS atau student_id.'
        return None, f'Nama siswa "{full_name}" tidak ditemukan.'

    return None, 'Isi salah satu kolom student_id, nis, atau nama_siswa.'


def _resolve_adjustment_subject(row):
    subject_id_raw = _row_value(row, 'subject_id', 'id_mapel')
    subject_name = _row_value(row, 'mapel', 'mata_pelajaran', 'subject')

    if subject_id_raw:
        try:
            subject_id = int(subject_id_raw)
        except ValueError:
            return None, f'ID mapel tidak valid: {subject_id_raw}'
        subject = Subject.query.filter_by(id=subject_id, is_deleted=False).first()
        return (subject, None) if subject else (None, f'Mapel ID {subject_id} tidak ditemukan.')

    if subject_name:
        matches = Subject.query.filter(
            Subject.is_deleted.is_(False),
            func.lower(Subject.name) == subject_name.lower(),
        ).all()
        if len(matches) == 1:
            return matches[0], None
        if len(matches) > 1:
            return None, f'Mapel "{subject_name}" duplikat, gunakan subject_id.'
        return None, f'Mapel "{subject_name}" tidak ditemukan.'

    return None, 'Isi salah satu kolom subject_id atau mapel.'


def _resolve_adjustment_academic_year(row):
    academic_year_id_raw = _row_value(row, 'academic_year_id', 'id_tahun_ajaran')
    year_name = _row_value(row, 'tahun_ajaran', 'academic_year', 'tahun')
    semester = _row_value(row, 'semester')

    if academic_year_id_raw:
        try:
            academic_year_id = int(academic_year_id_raw)
        except ValueError:
            return None, f'ID tahun ajaran tidak valid: {academic_year_id_raw}'
        academic_year = AcademicYear.query.filter_by(id=academic_year_id, is_deleted=False).first()
        return (academic_year, None) if academic_year else (None, f'Tahun ajaran ID {academic_year_id} tidak ditemukan.')

    if year_name:
        query = AcademicYear.query.filter(
            AcademicYear.is_deleted.is_(False),
            func.lower(AcademicYear.name) == year_name.lower(),
        )
        if semester:
            query = query.filter(func.lower(AcademicYear.semester) == semester.lower())
        matches = query.all()
        if len(matches) == 1:
            return matches[0], None
        if len(matches) > 1:
            return None, f'Tahun ajaran "{year_name}" memiliki beberapa semester, isi kolom semester.'
        return None, f'Tahun ajaran "{year_name}" tidak ditemukan.'

    return None, 'Isi salah satu kolom academic_year_id atau tahun_ajaran.'


def _resolve_adjustment_class(row, tenant_id, student):
    class_id_raw = _row_value(row, 'class_id', 'id_kelas')
    class_name = _row_value(row, 'kelas', 'class_name')

    if class_id_raw:
        try:
            class_id = int(class_id_raw)
        except ValueError:
            return None, f'ID kelas tidak valid: {class_id_raw}'
        class_room = scoped_classrooms_query(tenant_id).filter(ClassRoom.id == class_id).first()
        return (class_room, None) if class_room else (None, f'Kelas ID {class_id} tidak ditemukan di tenant ini.')

    if class_name:
        matches = scoped_classrooms_query(tenant_id).filter(func.lower(ClassRoom.name) == class_name.lower()).all()
        if len(matches) == 1:
            return matches[0], None
        if len(matches) > 1:
            return None, f'Kelas "{class_name}" duplikat, gunakan class_id.'
        return None, f'Kelas "{class_name}" tidak ditemukan di tenant ini.'

    return student.current_class, None


def _create_report_score_adjustment(student, class_room, academic_year, subject, adjusted_score, approval_reference, reason, tenant_id):
    class_id = class_room.id if class_room else student.current_class_id
    original_score = _calculated_final_for_adjustment(
        student_id=student.id,
        academic_year_id=academic_year.id,
        subject_id=subject.id,
        tenant_id=tenant_id,
        class_id=class_id,
    )

    existing_active = ReportScoreAdjustment.query.filter(
        ReportScoreAdjustment.tenant_id == tenant_id,
        ReportScoreAdjustment.student_id == student.id,
        ReportScoreAdjustment.academic_year_id == academic_year.id,
        ReportScoreAdjustment.subject_id == subject.id,
        ReportScoreAdjustment.status == REPORT_ADJUSTMENT_STATUS_ACTIVE,
        ReportScoreAdjustment.is_deleted.is_(False),
    ).all()
    for existing in existing_active:
        existing.status = REPORT_ADJUSTMENT_STATUS_VOID
        existing.void_reason = 'Digantikan oleh adjustment resmi baru.'
        existing.voided_by_user_id = current_user.id
        existing.voided_at = local_now()

    db.session.add(ReportScoreAdjustment(
        tenant_id=tenant_id,
        student_id=student.id,
        class_id=class_id,
        academic_year_id=academic_year.id,
        subject_id=subject.id,
        original_score=original_score,
        adjusted_score=adjusted_score,
        reason=reason,
        approval_reference=approval_reference,
        approved_by_user_id=current_user.id,
        approved_at=local_now(),
        status=REPORT_ADJUSTMENT_STATUS_ACTIVE,
    ))


@admin_bp.route('/akademik/adjustment-raport/template')
@login_required
@role_required(UserRole.ADMIN)
def report_score_adjustment_template():
    return _xlsx_response(
        'template_adjustment_nilai_raport.xlsx',
        'Adjustment Nilai',
        [
            'nis',
            'nama_siswa',
            'kelas',
            'tahun_ajaran',
            'semester',
            'mapel',
            'nilai_adjustment',
            'nomor_dokumen',
            'alasan',
        ],
        [
            [
                '12345',
                'Nama Siswa',
                'Kelas 7A',
                '2025/2026',
                'Ganjil',
                'Matematika',
                88.5,
                'BA-NILAI/2026/001',
                'Adjustment resmi sesuai berita acara.',
            ],
        ],
    )


@admin_bp.route('/akademik/adjustment-raport', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def manage_report_score_adjustments():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    class_options = scoped_classrooms_query(tenant_id).order_by(ClassRoom.name.asc()).all()
    student_options = _student_in_tenant_query(tenant_id).order_by(Student.full_name.asc()).all()
    academic_year_options = AcademicYear.query.filter(AcademicYear.is_deleted.is_(False)).order_by(
        AcademicYear.name.desc(),
        AcademicYear.semester.asc(),
    ).all()
    subject_options = Subject.query.filter(Subject.is_deleted.is_(False)).order_by(Subject.name.asc()).all()

    if request.method == 'POST':
        action = (request.form.get('action') or 'create').strip()
        if action == 'void':
            adjustment_id = request.form.get('adjustment_id', type=int)
            void_reason = (request.form.get('void_reason') or '').strip()
            adjustment = ReportScoreAdjustment.query.filter_by(
                id=adjustment_id,
                tenant_id=tenant_id,
                is_deleted=False,
            ).first()
            if not adjustment:
                flash('Data adjustment tidak ditemukan.', 'danger')
                return redirect(url_for('admin.manage_report_score_adjustments'))
            if not void_reason:
                flash('Alasan pembatalan wajib diisi.', 'warning')
                return redirect(url_for('admin.manage_report_score_adjustments'))
            adjustment.status = REPORT_ADJUSTMENT_STATUS_VOID
            adjustment.void_reason = void_reason
            adjustment.voided_by_user_id = current_user.id
            adjustment.voided_at = local_now()
            db.session.commit()
            flash('Adjustment raport resmi dibatalkan.', 'success')
            return redirect(url_for('admin.manage_report_score_adjustments'))

        if action == 'upload_excel':
            file = request.files.get('file')
            if not file or not file.filename:
                flash('File Excel belum dipilih.', 'warning')
                return redirect(url_for('admin.manage_report_score_adjustments'))
            if not file.filename.lower().endswith('.xlsx'):
                flash('Format file harus XLSX.', 'warning')
                return redirect(url_for('admin.manage_report_score_adjustments'))

            created = 0
            skipped = 0
            errors = []
            for idx, row in _iter_upload_rows(file):
                if not any((value or '').strip() for value in row.values()):
                    continue

                student, error = _resolve_adjustment_student(row, tenant_id)
                if error:
                    skipped += 1
                    errors.append(f'Baris {idx}: {error}')
                    continue

                subject, error = _resolve_adjustment_subject(row)
                if error:
                    skipped += 1
                    errors.append(f'Baris {idx}: {error}')
                    continue

                academic_year, error = _resolve_adjustment_academic_year(row)
                if error:
                    skipped += 1
                    errors.append(f'Baris {idx}: {error}')
                    continue

                class_room, error = _resolve_adjustment_class(row, tenant_id, student)
                if error:
                    skipped += 1
                    errors.append(f'Baris {idx}: {error}')
                    continue

                adjusted_score_raw = _row_value(row, 'nilai_adjustment', 'adjusted_score', 'nilai', 'nilai_akhir')
                approval_reference = _row_value(row, 'nomor_dokumen', 'approval_reference', 'dokumen', 'no_dokumen')
                reason = _row_value(row, 'alasan', 'reason', 'keterangan')

                if not approval_reference or not reason:
                    skipped += 1
                    errors.append(f'Baris {idx}: Nomor dokumen dan alasan wajib diisi.')
                    continue

                try:
                    adjusted_score = round(float(adjusted_score_raw.replace(',', '.')), 2)
                except (AttributeError, ValueError):
                    skipped += 1
                    errors.append(f'Baris {idx}: Nilai adjustment harus berupa angka.')
                    continue
                if adjusted_score < 0 or adjusted_score > 100:
                    skipped += 1
                    errors.append(f'Baris {idx}: Nilai adjustment harus berada pada rentang 0 sampai 100.')
                    continue

                try:
                    with db.session.begin_nested():
                        _create_report_score_adjustment(
                            student=student,
                            class_room=class_room,
                            academic_year=academic_year,
                            subject=subject,
                            adjusted_score=adjusted_score,
                            approval_reference=approval_reference,
                            reason=reason,
                            tenant_id=tenant_id,
                        )
                        created += 1
                except Exception as exc:
                    skipped += 1
                    errors.append(f'Baris {idx}: {exc}')

            db.session.commit()
            flash(f'Upload adjustment selesai. Berhasil: {created}, Dilewati: {skipped}.', 'success')
            if errors:
                flash('Contoh error: ' + '; '.join(errors[:5]), 'warning')
            return redirect(url_for('admin.manage_report_score_adjustments'))

        student_id = request.form.get('student_id', type=int)
        class_id = request.form.get('class_id', type=int) or None
        academic_year_id = request.form.get('academic_year_id', type=int)
        subject_id = request.form.get('subject_id', type=int)
        adjusted_score_raw = request.form.get('adjusted_score')
        approval_reference = (request.form.get('approval_reference') or '').strip()
        reason = (request.form.get('reason') or '').strip()

        student = _student_in_tenant_query(tenant_id).filter(Student.id == student_id).first()
        subject = Subject.query.filter_by(id=subject_id, is_deleted=False).first()
        academic_year = AcademicYear.query.filter_by(id=academic_year_id, is_deleted=False).first()
        class_room = None
        if class_id:
            class_room = ClassRoom.query.filter_by(id=class_id, is_deleted=False).first()
            if not class_room or not classroom_in_tenant(class_room, tenant_id):
                class_room = None

        if not student or not subject or not academic_year:
            flash('Siswa, tahun ajaran, atau mata pelajaran tidak valid untuk tenant ini.', 'danger')
            return redirect(url_for('admin.manage_report_score_adjustments'))
        if class_id and not class_room:
            flash('Kelas tidak valid untuk tenant ini.', 'danger')
            return redirect(url_for('admin.manage_report_score_adjustments'))
        if not approval_reference or not reason:
            flash('Nomor dokumen persetujuan dan alasan adjustment wajib diisi.', 'warning')
            return redirect(url_for('admin.manage_report_score_adjustments'))
        try:
            adjusted_score = round(float(adjusted_score_raw), 2)
        except (TypeError, ValueError):
            flash('Nilai adjustment harus berupa angka.', 'warning')
            return redirect(url_for('admin.manage_report_score_adjustments'))
        if adjusted_score < 0 or adjusted_score > 100:
            flash('Nilai adjustment harus berada pada rentang 0 sampai 100.', 'warning')
            return redirect(url_for('admin.manage_report_score_adjustments'))

        _create_report_score_adjustment(
            student=student,
            class_room=class_room,
            academic_year=academic_year,
            subject=subject,
            adjusted_score=adjusted_score,
            approval_reference=approval_reference,
            reason=reason,
            tenant_id=tenant_id,
        )
        db.session.commit()
        flash('Adjustment nilai raport resmi tersimpan.', 'success')
        return redirect(url_for('admin.manage_report_score_adjustments'))

    adjustments = (
        ReportScoreAdjustment.query.filter(
            ReportScoreAdjustment.tenant_id == tenant_id,
            ReportScoreAdjustment.is_deleted.is_(False),
        )
        .order_by(ReportScoreAdjustment.created_at.desc(), ReportScoreAdjustment.id.desc())
        .limit(300)
        .all()
    )
    return render_template(
        'admin/academic/report_score_adjustments.html',
        adjustments=adjustments,
        class_options=class_options,
        student_options=student_options,
        academic_year_options=academic_year_options,
        subject_options=subject_options,
        active_status=REPORT_ADJUSTMENT_STATUS_ACTIVE,
    )


@admin_bp.route('/platform/tenants', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.SUPER_ADMIN)
def manage_tenants():
    if request.method == 'POST':
        action = (request.form.get('action') or 'update_tenant').strip()

        if action == 'create_tenant':
            name = (request.form.get('name') or '').strip()
            code = _normalize_tenant_code(request.form.get('code'))
            slug = _slugify_tenant(request.form.get('slug') or name)
            timezone = (request.form.get('timezone') or 'Asia/Jakarta').strip()
            package = normalize_tenant_package(request.form.get('module_package'))

            if not name:
                flash('Nama tenant wajib diisi.', 'danger')
                return redirect(url_for('admin.manage_tenants'))
            if not code:
                flash('Kode tenant wajib diisi (huruf/angka).', 'danger')
                return redirect(url_for('admin.manage_tenants'))
            if Tenant.query.filter_by(code=code, is_deleted=False).first():
                flash(f'Kode tenant "{code}" sudah dipakai.', 'danger')
                return redirect(url_for('admin.manage_tenants'))
            if Tenant.query.filter_by(slug=slug, is_deleted=False).first():
                flash(f'Slug tenant "{slug}" sudah dipakai.', 'danger')
                return redirect(url_for('admin.manage_tenants'))

            tenant = Tenant(
                name=name,
                slug=slug,
                code=code,
                timezone=timezone or 'Asia/Jakarta',
                status=TenantStatus.ACTIVE,
                is_default=False,
            )
            db.session.add(tenant)
            db.session.flush()
            _upsert_tenant_config(tenant.id, TENANT_PACKAGE_KEY, package, 'Paket modul tenant.')
            _upsert_tenant_config(tenant.id, 'institution_name', name, 'Nama lembaga penerbit dokumen tenant.')
            _upsert_tenant_config(
                tenant.id,
                'institution_address',
                request.form.get('institution_address') or '',
                'Alamat lembaga untuk dokumen resmi tenant.',
            )
            _upsert_tenant_config(
                tenant.id,
                'institution_phone',
                request.form.get('institution_phone') or '',
                'Nomor telepon lembaga untuk dokumen resmi tenant.',
            )
            db.session.commit()
            flash(f'Tenant baru "{code}" berhasil dibuat.', 'success')
            return redirect(url_for('admin.manage_tenants'))

        if action == 'create_tenant_admin':
            tenant_id = request.form.get('tenant_id', type=int)
            tenant = Tenant.query.filter_by(id=tenant_id, is_deleted=False).first_or_404()
            username = (request.form.get('username') or '').strip()
            email = (request.form.get('email') or '').strip()
            password = request.form.get('password') or ''
            full_name = (request.form.get('full_name') or '').strip()

            if not username or not email or not password:
                flash('Username, email, dan password admin tenant wajib diisi.', 'danger')
                return redirect(url_for('admin.manage_tenants'))
            if len(password) < 8:
                flash('Password admin tenant minimal 8 karakter.', 'danger')
                return redirect(url_for('admin.manage_tenants'))
            if User.query.filter(User.username == username, User.is_deleted.is_(False)).first():
                flash(f'Username "{username}" sudah dipakai.', 'danger')
                return redirect(url_for('admin.manage_tenants'))
            if User.query.filter(User.email == email, User.is_deleted.is_(False)).first():
                flash(f'Email "{email}" sudah dipakai.', 'danger')
                return redirect(url_for('admin.manage_tenants'))

            new_user = User(
                tenant_id=tenant.id,
                username=username,
                email=email,
                role=UserRole.ADMIN,
                must_change_password=True,
            )
            new_user.set_password(password)
            db.session.add(new_user)
            db.session.flush()

            if full_name:
                db.session.add(Staff(
                    user_id=new_user.id,
                    full_name=full_name,
                    position='Admin Tenant',
                ))

            db.session.commit()
            flash(f'Admin tenant baru untuk "{tenant.code}" berhasil dibuat.', 'success')
            return redirect(url_for('admin.manage_tenants'))

        tenant_id = request.form.get('tenant_id', type=int)
        tenant = Tenant.query.filter_by(id=tenant_id, is_deleted=False).first_or_404()

        name = (request.form.get('name') or '').strip()
        timezone = (request.form.get('timezone') or '').strip()
        status = _parse_tenant_status(request.form.get('status'))
        address = request.form.get('institution_address') or ''
        phone = request.form.get('institution_phone') or ''
        package = normalize_tenant_package(request.form.get('module_package'))

        if not name:
            flash('Nama tenant wajib diisi.', 'danger')
            return redirect(url_for('admin.manage_tenants'))
        if not status:
            flash('Status tenant tidak valid.', 'danger')
            return redirect(url_for('admin.manage_tenants'))

        tenant.name = name
        tenant.status = status
        if timezone:
            tenant.timezone = timezone

        _upsert_tenant_config(tenant.id, TENANT_PACKAGE_KEY, package, 'Paket modul tenant.')
        _upsert_tenant_config(
            tenant.id,
            'institution_address',
            address,
            'Alamat lembaga untuk dokumen resmi tenant.',
        )
        _upsert_tenant_config(
            tenant.id,
            'institution_phone',
            phone,
            'Nomor telepon lembaga untuk dokumen resmi tenant.',
        )
        _upsert_tenant_config(
            tenant.id,
            'institution_name',
            name,
            'Nama lembaga penerbit dokumen tenant.',
        )

        db.session.commit()
        flash(f'Konfigurasi tenant "{tenant.code}" berhasil diperbarui.', 'success')
        return redirect(url_for('admin.manage_tenants'))

    query = (request.args.get('q') or '').strip()
    status_filter = _parse_tenant_status(request.args.get('status'))

    tenants_query = Tenant.query.filter(Tenant.is_deleted.is_(False))
    if status_filter:
        tenants_query = tenants_query.filter(Tenant.status == status_filter)
    if query:
        tenants_query = tenants_query.filter(
            or_(
                Tenant.name.ilike(f'%{query}%'),
                Tenant.code.ilike(f'%{query}%'),
                Tenant.slug.ilike(f'%{query}%'),
            )
        )

    tenants = tenants_query.order_by(Tenant.name.asc()).all()
    tenant_ids = [tenant.id for tenant in tenants]
    config_map = {}
    if tenant_ids:
        configs = (
            AppConfig.query
            .filter(
                AppConfig.tenant_id.in_(tenant_ids),
                AppConfig.key.in_(('institution_address', 'institution_phone', TENANT_PACKAGE_KEY)),
                AppConfig.is_deleted.is_(False),
            )
            .all()
        )
        for row in configs:
            config_map[(row.tenant_id, row.key)] = (row.value or '').strip()

    tenant_admins = (
        User.query
        .filter(
            User.tenant_id.in_(tenant_ids if tenant_ids else [-1]),
            User.is_deleted.is_(False),
            or_(
                User.role == UserRole.ADMIN,
                User.role_assignments.any(role=UserRole.ADMIN),
            ),
        )
        .order_by(User.tenant_id.asc(), User.username.asc())
        .all()
    )
    admin_map = {}
    for row in tenant_admins:
        admin_map.setdefault(row.tenant_id, []).append(row)

    return render_template(
        'admin/system/tenants.html',
        tenants=tenants,
        config_map=config_map,
        admin_map=admin_map,
        query=query,
        status_filter=(status_filter.name if status_filter else ''),
        status_options=list(TenantStatus),
        package_options=PACKAGE_OPTIONS,
        package_full=PACKAGE_FULL,
    )


# =========================================================
# 2. MASTER AKADEMIK (TAHUN AJARAN & MATA PELAJARAN)
# =========================================================

@admin_bp.route('/akademik/tahun-ajaran', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def manage_academic_years():
    if request.method == 'POST':
        name = request.form.get('name')  # 2025/2026
        semester = request.form.get('semester')  # Ganjil/Genap
        is_active = request.form.get('is_active') == 'on'

        if is_active:
            # Nonaktifkan tahun lain jika ini di-set aktif
            AcademicYear.query.update({AcademicYear.is_active: False})

        new_year = AcademicYear(name=name, semester=semester, is_active=is_active)
        db.session.add(new_year)
        db.session.commit()
        flash('Tahun ajaran berhasil dibuat.', 'success')
        return redirect(url_for('admin.manage_academic_years'))

    query = (request.args.get('q') or '').strip()
    years_query = AcademicYear.query
    if query:
        years_query = years_query.filter(
            or_(
                AcademicYear.name.ilike(f'%{query}%'),
                AcademicYear.semester.ilike(f'%{query}%')
            )
        )

    years = years_query.order_by(AcademicYear.id.desc()).all()
    return render_template('admin/academic/years.html', years=years, query=query)


@admin_bp.route('/akademik/aktifkan-tahun/<int:id>')
@login_required
@role_required(UserRole.ADMIN)
def activate_academic_year(id):
    # Nonaktifkan semua
    AcademicYear.query.update({AcademicYear.is_active: False})
    # Aktifkan yang dipilih
    year = AcademicYear.query.get_or_404(id)
    year.is_active = True
    db.session.commit()
    flash(f'Tahun Ajaran {year.name} - {year.semester} sekarang AKTIF.', 'success')
    return redirect(url_for('admin.manage_academic_years'))


@admin_bp.route('/akademik/mapel', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def manage_subjects():
    if request.method == 'POST':
        code = request.form.get('code')
        name = request.form.get('name')
        kkm = request.form.get('kkm')

        new_subject = Subject(code=code, name=name, kkm=float(kkm))
        db.session.add(new_subject)
        db.session.commit()
        flash('Mata Pelajaran ditambahkan.', 'success')
        return redirect(url_for('admin.manage_subjects'))

    query = (request.args.get('q') or '').strip()
    subjects_query = Subject.query.filter_by(is_deleted=False)
    if query:
        subjects_query = subjects_query.filter(
            or_(
                Subject.code.ilike(f'%{query}%'),
                Subject.name.ilike(f'%{query}%')
            )
        )

    subjects = subjects_query.order_by(Subject.name.asc()).all()
    return render_template('admin/academic/subjects.html', subjects=subjects, query=query)


@admin_bp.route('/akademik/mapel/edit/<int:subject_id>', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def edit_subject(subject_id):
    subject = Subject.query.get_or_404(subject_id)

    if request.method == 'POST':
        subject.code = request.form.get('code')
        subject.name = request.form.get('name')
        subject.kkm = float(request.form.get('kkm')) # nilai kkm dengan type float

        try:
            db.session.commit()
            flash(f'Mapel {subject.name} berhasil diperbarui.', 'success')
            return redirect(url_for('admin.manage_subjects'))
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal update: {e}', 'danger')

    return render_template('admin/academic/edit_subject.html', subject=subject)


# =========================================================
# 3. MASTER SDM (GURU & STAFF)
# =========================================================

@admin_bp.route('/sdm/guru', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def manage_teachers():
    if request.method == 'POST':
        tenant_id = _current_tenant_id()
        if tenant_id is None:
            flash('Tenant default tidak ditemukan.', 'danger')
            return redirect(url_for('admin.manage_teachers'))

        # 1. Buat User Login
        username = generate_nip()
        password = request.form.get('password') or "guru123"
        full_name = request.form.get('full_name')
        phone = request.form.get('phone')
        specialty = request.form.get('specialty')

        try:
            user = User(
                tenant_id=tenant_id,
                username=username,
                email=f"{username}@sekolah.id",
                password_hash=generate_password_hash(password),
                role=UserRole.GURU,
                must_change_password=True
            )
            db.session.add(user)
            db.session.flush()  # Dapat ID

            # 2. Buat Profile Guru
            teacher = Teacher(
                user_id=user.id,
                nip=username,
                full_name=full_name,
                phone=phone,
                specialty=specialty
            )
            db.session.add(teacher)
            db.session.commit()
            flash(f'Data Guru berhasil ditambahkan. NIP/Login: {username}', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {e}', 'danger')

        return redirect(url_for('admin.manage_teachers'))

    query = (request.args.get('q') or '').strip()
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return render_template('admin/hr/teachers.html', teachers=[], query=query)

    teachers_query = Teacher.query.join(User, Teacher.user_id == User.id).filter(
        Teacher.is_deleted == False,
        User.tenant_id == tenant_id,
    )
    if query:
        teachers_query = teachers_query.filter(
            or_(
                Teacher.full_name.ilike(f'%{query}%'),
                Teacher.nip.ilike(f'%{query}%'),
                Teacher.phone.ilike(f'%{query}%'),
                Teacher.specialty.ilike(f'%{query}%')
            )
        )

    teachers = teachers_query.order_by(Teacher.full_name.asc()).all()
    return render_template('admin/hr/teachers.html', teachers=teachers, query=query)


def _display_assignment_note(note):
    if not note:
        return None
    normalized = note.strip()
    internal_prefixes = (
        'Legacy ',
        'Class homeroom sync',
        'Admin assignment',
    )
    if normalized.startswith(internal_prefixes):
        return None
    return normalized


@admin_bp.route('/sdm/guru/<int:id>/assignments')
@login_required
@role_required(UserRole.ADMIN)
def teacher_assignments(id):
    tenant_id = _current_tenant_id()

    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.manage_teachers'))

    teacher = (
        Teacher.query.join(User, Teacher.user_id == User.id)
        .filter(
            Teacher.id == id,
            Teacher.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .first_or_404()
    )

    assignment_rows = (
        StaffAssignment.query
        .outerjoin(Program, StaffAssignment.program_id == Program.id)
        .outerjoin(ProgramGroup, StaffAssignment.group_id == ProgramGroup.id)
        .filter(
            StaffAssignment.tenant_id == tenant_id,
            StaffAssignment.person_id == teacher.person_id,
            StaffAssignment.is_deleted == False,
        )
        .order_by(
            StaffAssignment.end_date.isnot(None),
            Program.code.asc(),
            StaffAssignment.assignment_role.asc(),
            ProgramGroup.name.asc(),
            StaffAssignment.id.asc(),
        )
        .all()
    )

    grouped_assignments = []
    grouped_map = {}
    for assignment in assignment_rows:
        program = assignment.program
        if program is None:
            continue
        program_key = program.code
        if program_key not in grouped_map:
            grouped_map[program_key] = {
                'program_code': program.code,
                'program_name': program.name,
                'assignments': [],
                'active_count': 0,
            }
            grouped_assignments.append(grouped_map[program_key])

        row = {
            'id': assignment.id,
            'role': display_assignment_role(
                assignment.assignment_role,
                program.code,
                tenant_id=tenant_id,
            ),
            'group_name': assignment.group.name if assignment.group else '-',
            'academic_year': assignment.academic_year.name if assignment.academic_year else '-',
            'start_date': assignment.start_date,
            'end_date': assignment.end_date,
            'notes': _display_assignment_note(assignment.notes),
            'is_active': assignment.end_date is None,
        }
        grouped_map[program_key]['assignments'].append(row)
        if row['is_active']:
            grouped_map[program_key]['active_count'] += 1

    role_summary = {}
    for assignment in assignment_rows:
        if assignment.assignment_role:
            label = display_assignment_role(
                assignment.assignment_role,
                assignment.program.code if assignment.program else None,
                tenant_id=tenant_id,
            )
            role_summary[label] = role_summary.get(label, 0) + 1

    return render_template(
        'admin/hr/teacher_assignments.html',
        teacher=teacher,
        grouped_assignments=grouped_assignments,
        total_assignments=len(assignment_rows),
        role_summary=role_summary,
    )


@admin_bp.route('/sdm/guru/hapus/<int:id>', methods=['POST'])
@login_required
@role_required(UserRole.ADMIN)
def delete_teacher(id):
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.manage_teachers'))

    teacher = (
        Teacher.query.join(User, Teacher.user_id == User.id)
        .filter(
            Teacher.id == id,
            Teacher.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .first_or_404()
    )
    teacher.delete()
    if teacher.user:
        teacher.user.delete()
    db.session.commit()
    flash('Data guru berhasil dihapus.', 'success')
    return redirect(url_for('admin.manage_teachers'))


@admin_bp.route('/sdm/guru/upload', methods=['POST'])
@login_required
@role_required(UserRole.ADMIN)
def upload_teachers():
    file = request.files.get('file')
    if not file or file.filename == '':
        flash('File belum dipilih.', 'warning')
        return redirect(url_for('admin.manage_teachers'))

    if not file.filename.lower().endswith(('.csv', '.xlsx')):
        flash('Format file harus CSV atau XLSX.', 'warning')
        return redirect(url_for('admin.manage_teachers'))

    created = 0
    skipped = 0
    errors = []
    tenant_id = _current_tenant_id()

    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.manage_teachers'))

    for idx, row in _iter_upload_rows(file):
        nip = (row.get('nip') or row.get('NIP') or '').strip()
        full_name = (row.get('full_name') or row.get('nama') or row.get('nama_lengkap') or '').strip()
        specialty = (row.get('specialty') or row.get('mapel') or '').strip()
        phone = (row.get('phone') or row.get('no_hp') or row.get('whatsapp') or '').strip()
        password = (row.get('password') or '').strip() or "guru123"

        if not nip or not full_name:
            skipped += 1
            errors.append(f'Baris {idx}: NIP dan Nama wajib diisi.')
            continue

        existing_user = User.query.filter_by(username=nip).first()
        if existing_user:
            skipped += 1
            if existing_user.tenant_id != tenant_id:
                errors.append(f'Baris {idx}: NIP {nip} sudah dipakai tenant lain.')
            else:
                errors.append(f'Baris {idx}: NIP {nip} sudah terdaftar.')
            continue

        try:
            with db.session.begin_nested():
                user = User(
                    tenant_id=tenant_id,
                    username=nip,
                    email=f"{nip}@sekolah.id",
                    password_hash=generate_password_hash(password),
                    role=UserRole.GURU,
                    must_change_password=True
                )
                db.session.add(user)
                db.session.flush()

                teacher = Teacher(
                    user_id=user.id,
                    nip=nip,
                    full_name=full_name,
                    phone=phone,
                    specialty=specialty
                )
                db.session.add(teacher)
                created += 1
        except Exception as exc:
            skipped += 1
            errors.append(f'Baris {idx}: {exc}')

    db.session.commit()
    flash(f'Upload guru selesai. Berhasil: {created}, Dilewati: {skipped}.', 'success')
    if errors:
        flash('Contoh error: ' + '; '.join(errors[:3]), 'warning')
    return redirect(url_for('admin.manage_teachers'))


@admin_bp.route('/sdm/guru/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def edit_teacher(id):
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.manage_teachers'))

    # Ambil data guru, return 404 jika tidak ada
    teacher = (
        Teacher.query.join(User, Teacher.user_id == User.id)
        .filter(
            Teacher.id == id,
            Teacher.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .first_or_404()
    )

    if request.method == 'POST':
        # Ambil data dari form
        new_nip = request.form.get('nip')
        full_name = request.form.get('full_name')
        specialty = request.form.get('specialty')
        phone = request.form.get('phone')
        new_password = request.form.get('password')

        try:
            # 1. Update Data Profil Guru
            teacher.full_name = full_name
            teacher.specialty = specialty
            teacher.phone = phone

            # 2. Cek apakah NIP berubah? (NIP berpengaruh ke Username Login)
            if new_nip and new_nip != teacher.nip:
                # Cek apakah NIP baru sudah dipakai orang lain?
                existing_user = User.query.filter_by(username=new_nip).first()
                if existing_user:
                    flash(f'NIP {new_nip} sudah digunakan oleh user lain.', 'danger')
                    return redirect(url_for('admin.edit_teacher', id=id))

                # Jika aman, update NIP di tabel Teacher dan Username di tabel User
                teacher.nip = new_nip
                teacher.user.username = new_nip
                teacher.user.email = f"{new_nip}@sekolah.id"

            # 3. Update Password (hanya jika diisi)
            if new_password:
                teacher.user.password_hash = generate_password_hash(new_password)
                teacher.user.must_change_password = True  # Opsional: paksa ganti pas login

            db.session.commit()
            flash('Data Guru berhasil diperbarui.', 'success')
            return redirect(url_for('admin.manage_teachers'))

        except Exception as e:
            db.session.rollback()
            flash(f'Gagal update data: {e}', 'danger')

    # Render template edit
    return render_template('admin/hr/edit_teacher.html', teacher=teacher)

@admin_bp.route('/sdm/staff', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def manage_staff():
    if request.method == 'POST':
        tenant_id = _current_tenant_id()
        if tenant_id is None:
            flash('Tenant default tidak ditemukan.', 'danger')
            return redirect(url_for('admin.manage_staff'))

        username = (request.form.get('username') or '').strip()
        password = request.form.get('password') or "staff123"
        full_name = request.form.get('full_name')
        position = request.form.get('position') # Misal: Kepala TU, Staff Keuangan

        # Cek Username Kembar
        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            if existing_user.tenant_id != tenant_id:
                flash('Username sudah digunakan tenant lain.', 'danger')
            else:
                flash('Username sudah digunakan.', 'danger')
        else:
            try:
                # 1. Buat User Login
                user = User(
                    tenant_id=tenant_id,
                    username=username,
                    email=f"{username}@sekolah.id",
                    role=UserRole.TU,
                )
                user.set_password(password)
                db.session.add(user)
                db.session.flush()

                # 2. Buat Profil Staff
                staff = Staff(user_id=user.id, full_name=full_name, position=position)
                db.session.add(staff)
                db.session.commit()
                flash('Staff Tata Usaha berhasil ditambahkan.', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Error: {e}', 'danger')
        return redirect(url_for('admin.manage_staff'))

    query = (request.args.get('q') or '').strip()
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return render_template('admin/hr/staff.html', staff_list=[], query=query)

    staff_query = Staff.query.filter_by(is_deleted=False).join(User, Staff.user_id == User.id).filter(
        User.tenant_id == tenant_id
    )
    if query:
        staff_query = staff_query.filter(
            or_(
                Staff.full_name.ilike(f'%{query}%'),
                Staff.position.ilike(f'%{query}%'),
                User.username.ilike(f'%{query}%')
            )
        )

    staff_list = staff_query.order_by(Staff.full_name.asc()).all()
    return render_template('admin/hr/staff.html', staff_list=staff_list, query=query)


@admin_bp.route('/sdm/staff/hapus/<int:id>', methods=['POST'])
@login_required
@role_required(UserRole.ADMIN)
def delete_staff(id):
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.manage_staff'))

    staff = (
        Staff.query.join(User, Staff.user_id == User.id)
        .filter(
            Staff.id == id,
            Staff.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .first_or_404()
    )
    staff.delete()
    if staff.user:
        staff.user.delete()
    db.session.commit()
    flash('Data staff berhasil dihapus.', 'success')
    return redirect(url_for('admin.manage_staff'))


@admin_bp.route('/sdm/staff/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def edit_staff(id):
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.manage_staff'))

    # Ambil data staff, jika tidak ada return 404
    staff = (
        Staff.query.join(User, Staff.user_id == User.id)
        .filter(
            Staff.id == id,
            Staff.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .first_or_404()
    )

    if request.method == 'POST':
        # Ambil data input
        staff.full_name = request.form.get('full_name')
        staff.position = request.form.get('position')

        # Opsional: Jika ingin admin bisa reset password staff dari sini
        new_password = request.form.get('password')
        if new_password:  # Hanya update jika kolom password diisi
            staff.user.set_password(new_password)

        try:
            db.session.commit()
            flash('Data Staff berhasil diperbarui.', 'success')
            # Redirect kembali ke fungsi manage_staff (nama blueprint 'admin')
            return redirect(url_for('admin.manage_staff'))
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal memperbarui data: {e}', 'danger')

    # Render template edit
    return render_template('admin/hr/edit_staff.html', staff=staff)
# =========================================================
# 4. MASTER KELAS & WALI KELAS
# =========================================================

@admin_bp.route('/sekolah/kelas', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def manage_classes():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return render_template(
            'admin/academic/classes.html',
            classes=[],
            class_student_counts={},
            teachers=[],
            query=(request.args.get('q') or '').strip(),
            ProgramType=ProgramType,
            EducationLevel=EducationLevel,
        )

    if request.method == 'POST':
        name = request.form.get('name')
        grade_level = request.form.get('grade_level')
        homeroom_id = request.form.get('homeroom_teacher_id', type=int)  # ID Guru
        program_type_raw = request.form.get('program_type')
        education_level_raw = request.form.get('education_level')

        if homeroom_id:
            homeroom_teacher = _tenant_teachers_query(tenant_id).filter(Teacher.id == homeroom_id).first()
            if homeroom_teacher is None:
                flash('Wali kelas tidak valid untuk tenant aktif.', 'warning')
                return redirect(url_for('admin.manage_classes'))

        program_type = ProgramType[program_type_raw] if program_type_raw else None
        education_level = EducationLevel[education_level_raw] if education_level_raw else None

        new_class = ClassRoom(
            name=name,
            grade_level=grade_level,
            homeroom_teacher_id=homeroom_id,
            program_type=program_type,
            education_level=education_level
        )
        db.session.add(new_class)
        db.session.flush()
        ensure_formal_program_group(new_class, tenant_id=tenant_id)
        ensure_rumah_quran_program_group(new_class, tenant_id=tenant_id)
        ensure_bahasa_program_group(new_class, tenant_id=tenant_id)
        sync_class_homeroom_assignment(new_class)
        db.session.commit()
        flash('Kelas berhasil dibuat.', 'success')
        return redirect(url_for('admin.manage_classes'))

    query = (request.args.get('q') or '').strip()
    classes_query = scoped_classrooms_query(tenant_id).outerjoin(Teacher, ClassRoom.homeroom_teacher_id == Teacher.id)
    if query:
        classes_query = classes_query.filter(
            or_(
                ClassRoom.name.ilike(f'%{query}%'),
                Teacher.full_name.ilike(f'%{query}%')
            )
        )

    classes = classes_query.order_by(ClassRoom.name.asc()).all()
    class_student_counts = {}
    for class_room in classes:
        if is_rumah_quran_classroom(class_room):
            class_student_counts[class_room.id] = len(list_rumah_quran_students_for_class(class_room.id))
        elif is_bahasa_classroom(class_room):
            class_student_counts[class_room.id] = len(list_bahasa_students_for_class(class_room.id))
        elif class_room.program_group_id:
            class_student_counts[class_room.id] = len(list_formal_students_for_class(class_room.id))
        else:
            class_student_counts[class_room.id] = len(class_room.students)
    teachers = _tenant_teachers_query(tenant_id).order_by(Teacher.full_name.asc()).all()  # Untuk dropdown
    return render_template(
        'admin/academic/classes.html',
        classes=classes,
        class_student_counts=class_student_counts,
        teachers=teachers,
        query=query,
        ProgramType=ProgramType,
        EducationLevel=EducationLevel
    )


@admin_bp.route('/sekolah/kelas/edit/<int:class_id>', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def edit_class(class_id):
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.manage_classes'))

    # Ambil data kelas atau 404 jika tidak ada
    class_room = scoped_classrooms_query(tenant_id).filter(ClassRoom.id == class_id).first_or_404()
    teachers = _tenant_teachers_query(tenant_id).order_by(Teacher.full_name.asc()).all()

    if request.method == 'POST':
        class_room.name = request.form.get('name')
        class_room.grade_level = request.form.get('grade_level')
        program_type_raw = request.form.get('program_type')
        education_level_raw = request.form.get('education_level')
        class_room.program_type = ProgramType[program_type_raw] if program_type_raw else None
        class_room.education_level = EducationLevel[education_level_raw] if education_level_raw else None

        # Handle Wali Kelas (Bisa Kosong/None)
        homeroom_id = request.form.get('homeroom_teacher_id', type=int)
        if homeroom_id:
            homeroom_teacher = _tenant_teachers_query(tenant_id).filter(Teacher.id == homeroom_id).first()
            if homeroom_teacher is None:
                flash('Wali kelas tidak valid untuk tenant aktif.', 'warning')
                return redirect(url_for('admin.edit_class', class_id=class_id))
        class_room.homeroom_teacher_id = homeroom_id if homeroom_id else None

        try:
            ensure_formal_program_group(class_room, tenant_id=tenant_id)
            ensure_rumah_quran_program_group(class_room, tenant_id=tenant_id)
            ensure_bahasa_program_group(class_room, tenant_id=tenant_id)
            sync_class_homeroom_assignment(class_room)
            db.session.commit()
            flash(f'Kelas {class_room.name} berhasil diperbarui.', 'success')
            return redirect(url_for('admin.manage_classes'))
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal update: {e}', 'danger')

    return render_template(
        'admin/academic/edit_class.html',
        class_room=class_room,
        teachers=teachers,
        ProgramType=ProgramType,
        EducationLevel=EducationLevel
    )


@admin_bp.route('/sekolah/kelas/hapus/<int:class_id>', methods=['POST'])
@login_required
@role_required(UserRole.ADMIN)
def delete_class(class_id):
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.manage_classes'))

    class_room = scoped_classrooms_query(tenant_id).filter(ClassRoom.id == class_id).first_or_404()

    if is_rumah_quran_classroom(class_room):
        student_count = len(list_rumah_quran_students_for_class(class_room.id))
    elif is_bahasa_classroom(class_room):
        student_count = len(list_bahasa_students_for_class(class_room.id))
    elif class_room.program_group_id:
        student_count = len(list_formal_students_for_class(class_room.id))
    else:
        student_count = Student.query.filter_by(
            current_class_id=class_room.id,
            is_deleted=False,
        ).count()

    majlis_parent_count = Parent.query.filter_by(
        majlis_class_id=class_room.id,
        is_deleted=False,
    ).count()
    majlis_participant_count = MajlisParticipant.query.filter_by(
        majlis_class_id=class_room.id,
        is_deleted=False,
    ).count()

    if student_count > 0 or majlis_parent_count > 0 or majlis_participant_count > 0:
        flash(
            (
                f'Kelas "{class_room.name}" tidak bisa dihapus karena masih memiliki peserta aktif '
                f'(siswa: {student_count}, peserta majlis: {majlis_participant_count}, wali majlis: {majlis_parent_count}).'
            ),
            'danger'
        )
        return redirect(url_for('admin.manage_classes'))

    try:
        schedules = Schedule.query.filter_by(class_id=class_room.id, is_deleted=False).all()
        for schedule in schedules:
            schedule.is_deleted = True

        class_room.homeroom_teacher_id = None
        sync_class_homeroom_assignment(class_room)
        class_room.is_deleted = True

        if class_room.program_group_id:
            program_group = ProgramGroup.query.filter_by(
                id=class_room.program_group_id,
                tenant_id=tenant_id,
                is_deleted=False,
            ).first()
            if program_group:
                program_group.is_active = False

        db.session.commit()
        flash(f'Kelas "{class_room.name}" berhasil dihapus.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal menghapus kelas: {e}', 'danger')

    return redirect(url_for('admin.manage_classes'))

# =========================================================
# 5. MASTER KESISWAAN (EKSKUL)
# =========================================================

@admin_bp.route('/kesiswaan/ekskul', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def manage_extracurriculars():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        name = request.form.get('name')
        supervisor_id = request.form.get('supervisor_id', type=int)
        if supervisor_id and _tenant_teachers_query(tenant_id).filter(Teacher.id == supervisor_id).first() is None:
            flash('Pembina ekstrakurikuler tidak valid untuk tenant aktif.', 'warning')
            return redirect(url_for('admin.manage_extracurriculars'))

        ekskul = Extracurricular(name=name, supervisor_teacher_id=supervisor_id)
        db.session.add(ekskul)
        db.session.commit()
        flash('Ekstrakurikuler ditambahkan.', 'success')
        return redirect(url_for('admin.manage_extracurriculars'))

    ekskuls = Extracurricular.query.filter_by(is_deleted=False).all()
    teachers = _tenant_teachers_query(tenant_id).order_by(Teacher.full_name.asc()).all()
    return render_template('admin/student_affairs/extracurriculars.html', ekskuls=ekskuls, teachers=teachers)


# =========================================================
# 6. MANAJEMEN SISWA
# =========================================================

from app.forms import StudentForm
from app.models import User, Student, Parent, ClassRoom, UserRole, Gender


@admin_bp.route('/student/tambah', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def add_student():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.list_students'))

    form = StudentForm()

    # 1. Isi Pilihan Kelas (Wajib diisi dinamis setiap loading halaman)
    # Kita ambil ID dan Nama Kelas dari database
    form.class_id.choices = [(c.id, c.name) for c in scoped_classrooms_query(tenant_id).all()]

    # Jika belum ada kelas sama sekali, kasih opsi dummy biar gak error
    if not form.class_id.choices:
        form.class_id.choices = [(0, 'Belum ada kelas')]

    if form.validate_on_submit():
        try:
            if form.class_id.data and form.class_id.data != 0:
                selected_class = scoped_classrooms_query(tenant_id).filter(ClassRoom.id == form.class_id.data).first()
                if selected_class is None:
                    raise ValueError('Kelas tidak valid untuk tenant aktif.')

            nis = (form.nis.data or '').strip() or generate_nis()

            # A. CEK DUPLIKASI (Penting!)
            existing_student_user = User.query.filter_by(username=nis).first()
            if existing_student_user:
                if existing_student_user.tenant_id != tenant_id:
                    flash('NIS sudah dipakai tenant lain.', 'warning')
                else:
                    flash('NIS sudah terdaftar sebagai User.', 'warning')
                return render_template('admin/add_student.html', form=form)

            # B. BUAT USER SISWA
            student_user = User(
                tenant_id=tenant_id,
                username=nis,  # Login pakai NIS
                email=form.email.data,  # Pakai email dari inputan form
                role=UserRole.SISWA
            )
            student_user.set_password("123456")  # Default Pass
            db.session.add(student_user)
            db.session.flush()

            # C. BUAT PROFIL SISWA
            new_student = Student(
                user_id=student_user.id,
                nis=nis,
                full_name=form.full_name.data,
                gender=Gender[form.gender.data],  # Konversi string 'L'/'P' ke Enum
                place_of_birth=form.place_of_birth.data,
                date_of_birth=form.date_of_birth.data,
                current_class_id=form.class_id.data if form.class_id.data != 0 else None,
                address=form.address.data
            )
            db.session.add(new_student)

            # D. BUAT USER & PROFIL WALI
            # Cek dulu takutnya ortu sudah punya akun (kakak kelas)
            parent_user = User.query.filter_by(username=form.parent_phone.data).first()

            if not parent_user:
                # Buat Akun Wali Baru
                parent_user = User(
                    tenant_id=tenant_id,
                    username=form.parent_phone.data,  # Login pakai No WA
                    email=f"{form.parent_phone.data}@wali.sekolah.id",  # Email dummy
                    role=UserRole.WALI_MURID
                )
                parent_user.set_password(form.parent_phone.data)  # Default Pass = No WA
                db.session.add(parent_user)
                db.session.flush()
            elif parent_user.tenant_id != tenant_id:
                raise ValueError('Nomor HP wali sudah dipakai tenant lain.')

            if parent_user.tenant_id is None:
                parent_user.tenant_id = tenant_id

            if parent_user.role == UserRole.ADMIN:
                raise ValueError('Akun admin tidak boleh dipakai sebagai wali murid.')

            if not parent_user.has_role(UserRole.WALI_MURID):
                db.session.add(UserRoleAssignment(user_id=parent_user.id, role=UserRole.WALI_MURID))

            if parent_user.role != UserRole.WALI_MURID:
                parent_user.role = UserRole.WALI_MURID
            db.session.flush()

            # Buat/ambil profil wali
            parent_profile = parent_user.parent_profile
            if not parent_profile:
                parent_profile = Parent(
                    user_id=parent_user.id,
                    full_name=form.parent_name.data,
                    phone=form.parent_phone.data,
                    job=form.parent_job.data,
                    address=form.address.data
                )
                db.session.add(parent_profile)
                db.session.flush()

            # Sambungkan Siswa ke Wali
            new_student.parent_id = parent_profile.id

            sync_student_formal_class_membership(new_student, new_student.current_class_id)
            db.session.commit()
            flash(f'Siswa {form.full_name.data} berhasil ditambahkan. NIS/Login: {nis}', 'success')
            return redirect(url_for('admin.list_students'))

        except Exception as e:
            db.session.rollback()
            flash(f"Gagal menyimpan: {str(e)}", 'danger')

    return render_template('admin/add_student.html', form=form)


@admin_bp.route('/student/edit/<int:student_id>', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def edit_student(student_id):
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.list_students'))

    student = (
        Student.query.join(User, Student.user_id == User.id)
        .filter(
            Student.id == student_id,
            Student.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .first_or_404()
    )
    return_url = _safe_students_list_return_url(
        request.args.get('next') or request.form.get('next'),
        fallback_endpoint='admin.list_students'
    )
    classes = scoped_classrooms_query(tenant_id).all()
    rumah_quran_classes = [class_room for class_room in list_rumah_quran_classes() if classroom_in_tenant(class_room, tenant_id)]
    rumah_quran_class = get_student_rumah_quran_classroom(student)
    bahasa_classes = [class_room for class_room in list_bahasa_classes() if classroom_in_tenant(class_room, tenant_id)]
    bahasa_class = get_student_bahasa_classroom(student)

    if request.method == 'POST':
        # Update Data Dasar
        student.full_name = request.form.get('full_name')
        student.nis = request.form.get('nis')
        student.nisn = (request.form.get('nisn') or '').strip() or None

        # Update Kelas
        cid = request.form.get('class_id')
        selected_class_id = int(cid) if cid else None
        student.current_class_id = selected_class_id
        rumah_quran_class_id = request.form.get('rumah_quran_class_id')
        rumah_quran_class_id = int(rumah_quran_class_id) if rumah_quran_class_id else None
        bahasa_class_id = request.form.get('bahasa_class_id')
        bahasa_class_id = int(bahasa_class_id) if bahasa_class_id else None

        selected_class = (
            scoped_classrooms_query(tenant_id).filter(ClassRoom.id == selected_class_id).first()
            if selected_class_id else None
        )
        if selected_class_id and selected_class is None:
            flash('Kelas tidak valid untuk tenant aktif.', 'warning')
            return redirect(url_for('admin.edit_student', student_id=student_id, next=return_url))
        if selected_class and selected_class.program_type in (ProgramType.RQDF_SORE, ProgramType.TAKHOSUS_TAHFIDZ):
            rumah_quran_class_id = selected_class.id
        if selected_class and selected_class.program_type == ProgramType.BAHASA:
            bahasa_class_id = selected_class.id

        # Update SPP Khusus
        spp = request.form.get('custom_spp')
        if spp:
            student.custom_spp_fee = int(''.join(filter(str.isdigit, spp)))
        else:
            student.custom_spp_fee = None

        try:
            sync_student_formal_class_membership(student, selected_class_id)
            rumah_quran_ok = assign_student_rumah_quran_class(student, rumah_quran_class_id)
            if not rumah_quran_ok:
                raise ValueError("Gagal memperbarui penempatan halaqoh Rumah Qur'an untuk siswa ini.")
            assign_student_bahasa_class(student, bahasa_class_id)
            student.save()  # Menggunakan method save() dari BaseModel
            flash('Data siswa diupdate.', 'success')
            return redirect(return_url)
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal update data siswa: {e}', 'danger')

    return render_template('staff/edit_student.html',
                           student=student,
                           classes=classes,
                           rumah_quran_classes=rumah_quran_classes,
                           rumah_quran_class=rumah_quran_class,
                           bahasa_classes=bahasa_classes,
                           bahasa_class=bahasa_class,
                           return_url=return_url)


@admin_bp.route('/daftar-student')
@login_required
@role_required(UserRole.ADMIN)
def list_students():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    query = (request.args.get('q') or '').strip()
    query_majlis = (request.args.get('q_majlis') or '').strip()
    active_category = (request.args.get('category') or 'all').strip().lower()
    package = get_tenant_package(tenant_id)
    if package == PACKAGE_SEKOLAH:
        allowed_categories = {'all', 'sbq_sd', 'sbq_smp', 'sbq_sma', 'bahasa'}
    elif package == PACKAGE_RUMAH_QURAN:
        allowed_categories = {'all', 'reguler', 'takhosus', 'bahasa'}
    else:
        allowed_categories = {'all', 'sbq_sd', 'sbq_smp', 'sbq_sma', 'bahasa', 'reguler', 'takhosus'}
    if active_category not in allowed_categories:
        active_category = 'all'
    if package == PACKAGE_SEKOLAH:
        query_majlis = ''

    students_query = (
        Student.query.join(User, Student.user_id == User.id)
        .filter(
            Student.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .outerjoin(ClassRoom, Student.current_class_id == ClassRoom.id)
    )

    if query:
        students_query = students_query.outerjoin(Parent, Student.parent_id == Parent.id).filter(
            db.or_(
                Student.full_name.ilike(f'%{query}%'),
                Student.nis.ilike(f'%{query}%'),
                Parent.full_name.ilike(f'%{query}%'),
                Parent.phone.ilike(f'%{query}%'),
                ClassRoom.name.ilike(f'%{query}%')
            )
        )

    if active_category == 'sbq_sd':
        students_query = students_query.filter(
            or_(
                and_(
                    ClassRoom.program_type == ProgramType.SEKOLAH_FULLDAY,
                    ClassRoom.education_level == EducationLevel.SD
                ),
                and_(
                    ClassRoom.program_type.is_(None),
                    or_(
                        ClassRoom.name.ilike('%sd%'),
                        ClassRoom.grade_level.in_([1, 2, 3, 4, 5, 6])
                    )
                )
            )
        )
    elif active_category == 'sbq_smp':
        students_query = students_query.filter(
            or_(
                and_(
                    ClassRoom.program_type == ProgramType.SEKOLAH_FULLDAY,
                    ClassRoom.education_level == EducationLevel.SMP
                ),
                and_(
                    ClassRoom.program_type.is_(None),
                    or_(
                        ClassRoom.name.ilike('%smp%'),
                        ClassRoom.grade_level.in_([7, 8, 9])
                    )
                )
            )
        )
    elif active_category == 'sbq_sma':
        students_query = students_query.filter(
            or_(
                and_(
                    ClassRoom.program_type == ProgramType.SEKOLAH_FULLDAY,
                    ClassRoom.education_level == EducationLevel.SMA
                ),
                and_(
                    ClassRoom.program_type.is_(None),
                    or_(
                        ClassRoom.name.ilike('%sma%'),
                        ClassRoom.grade_level.in_([10, 11, 12])
                    )
                )
            )
        )
    elif active_category == 'reguler':
        students_query = apply_rumah_quran_student_filter(students_query, track='reguler')
    elif active_category == 'takhosus':
        students_query = apply_rumah_quran_student_filter(students_query, track='takhosus')
    elif active_category == 'bahasa':
        students_query = apply_bahasa_student_filter(students_query)

    students = students_query.order_by(Student.id.desc()).all()
    bahasa_class_map = {}
    if active_category == 'bahasa':
        bahasa_class_map = {
            student.id: get_student_bahasa_classroom(student)
            for student in students
        }
    majlis_participants = (
        []
        if package == PACKAGE_SEKOLAH
        else list_active_majlis_participants(search=query_majlis, tenant_id=tenant_id)
    )

    return render_template(
        'student/list_students.html',
        students=students,
        bahasa_class_map=bahasa_class_map,
        majlis_participants=majlis_participants,
        query=query,
        query_majlis=query_majlis,
        active_category=active_category
    )


@admin_bp.route('/student/upload', methods=['POST'])
@login_required
@role_required(UserRole.ADMIN)
def upload_students():
    file = request.files.get('file')
    if not file or file.filename == '':
        flash('File belum dipilih.', 'warning')
        return redirect(url_for('admin.list_students'))

    if not file.filename.lower().endswith(('.csv', '.xlsx')):
        flash('Format file harus CSV atau XLSX.', 'warning')
        return redirect(url_for('admin.list_students'))

    created = 0
    skipped = 0
    errors = []
    tenant_id = _current_tenant_id()

    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.list_students'))

    for idx, row in _iter_upload_rows(file):
        nis = (row.get('nis') or row.get('NIS') or '').strip()
        full_name = (row.get('full_name') or row.get('nama') or row.get('nama_lengkap') or '').strip()
        gender_raw = (row.get('gender') or row.get('jk') or row.get('jenis_kelamin') or '').strip().upper()
        class_name = (row.get('class') or row.get('kelas') or row.get('class_name') or '').strip()
        place_of_birth = (row.get('place_of_birth') or row.get('tempat_lahir') or '').strip()
        date_of_birth = (row.get('date_of_birth') or row.get('tanggal_lahir') or '').strip()
        address = (row.get('address') or row.get('alamat') or '').strip()
        email = (row.get('email') or '').strip()
        parent_name = (row.get('parent_name') or row.get('nama_wali') or '').strip()
        parent_phone = (row.get('parent_phone') or row.get('no_wa') or row.get('no_hp_wali') or '').strip()
        parent_job = (row.get('parent_job') or row.get('pekerjaan_wali') or '').strip()

        if not full_name:
            skipped += 1
            errors.append(f'Baris {idx}: Nama wajib diisi.')
            continue

        if not parent_phone:
            skipped += 1
            errors.append(f'Baris {idx}: Nomor HP wali wajib diisi.')
            continue

        if not nis:
            nis = generate_nis()

        existing_student_user = User.query.filter_by(username=nis).first()
        if existing_student_user:
            skipped += 1
            if existing_student_user.tenant_id != tenant_id:
                errors.append(f'Baris {idx}: NIS {nis} sudah dipakai tenant lain.')
            else:
                errors.append(f'Baris {idx}: NIS {nis} sudah terdaftar.')
            continue

        if gender_raw not in {'L', 'P'}:
            skipped += 1
            errors.append(f'Baris {idx}: Gender harus L atau P.')
            continue

        try:
            dob = datetime.strptime(date_of_birth, '%Y-%m-%d').date()
        except ValueError:
            skipped += 1
            errors.append(f'Baris {idx}: Tanggal lahir harus YYYY-MM-DD.')
            continue

        class_id = None
        if class_name:
            class_room = scoped_classrooms_query(tenant_id).filter(ClassRoom.name == class_name).first()
            if class_room:
                class_id = class_room.id

        try:
            with db.session.begin_nested():
                student_user = User(
                    tenant_id=tenant_id,
                    username=nis,
                    email=email or f"{nis}@sekolah.id",
                    role=UserRole.SISWA
                )
                student_user.set_password("123456")
                db.session.add(student_user)
                db.session.flush()

                new_student = Student(
                    user_id=student_user.id,
                    nis=nis,
                    full_name=full_name,
                    gender=Gender[gender_raw],
                    place_of_birth=place_of_birth,
                    date_of_birth=dob,
                    current_class_id=class_id,
                    address=address
                )
                db.session.add(new_student)

                parent_user = User.query.filter_by(username=parent_phone).first()
                if not parent_user:
                    parent_user = User(
                        tenant_id=tenant_id,
                        username=parent_phone,
                        email=f"{parent_phone}@wali.sekolah.id",
                        role=UserRole.WALI_MURID
                    )
                    parent_user.set_password(parent_phone)
                    db.session.add(parent_user)
                    db.session.flush()
                elif parent_user.tenant_id != tenant_id:
                    raise ValueError('Nomor HP wali sudah terdaftar pada tenant lain.')

                if parent_user.tenant_id is None:
                    parent_user.tenant_id = tenant_id

                if not parent_user.has_role(UserRole.WALI_MURID):
                    db.session.add(UserRoleAssignment(user_id=parent_user.id, role=UserRole.WALI_MURID))

                if parent_user.role == UserRole.ADMIN:
                    raise ValueError('Akun admin tidak boleh dipakai sebagai wali murid.')

                if parent_user.role != UserRole.WALI_MURID:
                    parent_user.role = UserRole.WALI_MURID

                db.session.flush()

                parent_profile = parent_user.parent_profile
                if not parent_profile:
                    parent_profile = Parent(
                        user_id=parent_user.id,
                        full_name=parent_name or "Wali Murid",
                        phone=parent_phone,
                        job=parent_job,
                        address=address
                    )
                    db.session.add(parent_profile)
                    db.session.flush()

                new_student.parent_id = parent_profile.id

                created += 1
        except Exception as exc:
            skipped += 1
            errors.append(f'Baris {idx}: {exc}')

    db.session.commit()
    flash(f'Upload siswa selesai. Berhasil: {created}, Dilewati: {skipped}.', 'success')
    if errors:
        flash('Contoh error: ' + '; '.join(errors[:3]), 'warning')
    return redirect(url_for('admin.list_students'))


@admin_bp.route('/student/hapus/<int:id>')
@login_required
@role_required(UserRole.ADMIN)
def delete_student(id):
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.list_students'))

    student = (
        Student.query.join(User, Student.user_id == User.id)
        .filter(
            Student.id == id,
            Student.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .first_or_404()
    )
    student.delete()  # Menggunakan method Soft Delete dari BaseModel
    flash('Data siswa berhasil dihapus (Soft Delete).', 'warning')
    return redirect(url_for('admin.list_students'))


# =========================================================
# 7. MANAJEMEN KEUANGAN
# =========================================================

@admin_bp.route('/keuangan/akun', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.TU)
def manage_finance_accounts():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        action = (request.form.get('action') or 'create').strip()

        if action == 'create':
            code = (request.form.get('code') or '').strip()
            name = (request.form.get('name') or '').strip()
            category_raw = (request.form.get('category') or '').strip().upper()
            normal_balance_raw = (request.form.get('normal_balance') or '').strip().upper()

            if not code or not name:
                flash('Kode akun dan nama akun wajib diisi.', 'warning')
                return redirect(url_for('admin.manage_finance_accounts'))

            try:
                category = FinanceAccountCategory[category_raw]
                normal_balance = FinanceNormalBalance[normal_balance_raw]
            except KeyError:
                flash('Kategori akun atau normal balance tidak valid.', 'danger')
                return redirect(url_for('admin.manage_finance_accounts'))

            existing = FinanceAccount.query.filter_by(tenant_id=tenant_id, code=code).first()
            if existing:
                flash(f'Kode akun "{code}" sudah digunakan.', 'warning')
                return redirect(url_for('admin.manage_finance_accounts'))

            db.session.add(FinanceAccount(
                tenant_id=tenant_id,
                code=code,
                name=name,
                category=category,
                normal_balance=normal_balance,
                is_active=True,
            ))
            db.session.commit()
            flash(f'Akun "{code} - {name}" berhasil ditambahkan.', 'success')
            return redirect(url_for('admin.manage_finance_accounts'))

        if action == 'toggle':
            account_id = request.form.get('account_id', type=int)
            account = FinanceAccount.query.filter_by(id=account_id, tenant_id=tenant_id).first_or_404()
            account.is_active = not account.is_active
            db.session.commit()
            flash(
                f'Akun "{account.code} - {account.name}" {"diaktifkan" if account.is_active else "dinonaktifkan"}.',
                'success'
            )
            return redirect(url_for('admin.manage_finance_accounts'))

    query = (request.args.get('q') or '').strip()
    category_filter = (request.args.get('category') or '').strip().upper()
    accounts_query = FinanceAccount.query.filter(FinanceAccount.tenant_id == tenant_id)
    if query:
        accounts_query = accounts_query.filter(
            or_(
                FinanceAccount.code.ilike(f'%{query}%'),
                FinanceAccount.name.ilike(f'%{query}%'),
            )
        )
    if category_filter and category_filter in FinanceAccountCategory.__members__:
        accounts_query = accounts_query.filter(FinanceAccount.category == FinanceAccountCategory[category_filter])

    accounts = accounts_query.order_by(FinanceAccount.code.asc()).all()
    return render_template(
        'admin/finance/accounts.html',
        accounts=accounts,
        query=query,
        category_filter=category_filter,
        categories=list(FinanceAccountCategory),
        normal_balances=list(FinanceNormalBalance),
    )


@admin_bp.route('/keuangan/settings', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.TU)
def manage_finance_settings():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    settings = FinanceSetting.query.filter_by(tenant_id=tenant_id).first()

    if request.method == 'POST':
        action = (request.form.get('action') or 'save_defaults').strip()

        if action == 'save_defaults':
            if not settings:
                settings = FinanceSetting(tenant_id=tenant_id)
                db.session.add(settings)

            basis_raw = (request.form.get('accounting_basis') or 'CASH').strip().upper()
            if basis_raw not in FinanceAccountingBasis.__members__:
                flash('Accounting basis tidak valid.', 'warning')
                return redirect(url_for('admin.manage_finance_settings'))

            settings.accounting_basis = FinanceAccountingBasis[basis_raw]
            settings.default_cash_bank_account_id = request.form.get('default_cash_bank_account_id', type=int) or None
            settings.default_spp_revenue_account_id = request.form.get('default_spp_revenue_account_id', type=int) or None
            settings.default_registration_revenue_account_id = request.form.get('default_registration_revenue_account_id', type=int) or None
            settings.default_savings_liability_account_id = request.form.get('default_savings_liability_account_id', type=int) or None
            settings.default_donation_revenue_account_id = request.form.get('default_donation_revenue_account_id', type=int) or None

            if settings.default_cash_bank_account_id:
                cash_bank = FinanceCashBankAccount.query.filter_by(
                    id=settings.default_cash_bank_account_id,
                    tenant_id=tenant_id
                ).first()
                if not cash_bank:
                    flash('Default kas/bank tidak valid untuk tenant ini.', 'danger')
                    db.session.rollback()
                    return redirect(url_for('admin.manage_finance_settings'))

            account_ids = [
                settings.default_spp_revenue_account_id,
                settings.default_registration_revenue_account_id,
                settings.default_savings_liability_account_id,
                settings.default_donation_revenue_account_id,
            ]
            valid_account_ids = [account_id for account_id in account_ids if account_id]
            if valid_account_ids:
                count_valid = FinanceAccount.query.filter(
                    FinanceAccount.tenant_id == tenant_id,
                    FinanceAccount.id.in_(valid_account_ids)
                ).count()
                if count_valid != len(set(valid_account_ids)):
                    flash('Salah satu default akun tidak valid untuk tenant ini.', 'danger')
                    db.session.rollback()
                    return redirect(url_for('admin.manage_finance_settings'))

            db.session.commit()
            flash('Finance settings berhasil disimpan.', 'success')
            return redirect(url_for('admin.manage_finance_settings'))

        if action == 'create_period':
            name = (request.form.get('name') or '').strip()
            start_date = _parse_iso_date(request.form.get('start_date'))
            end_date = _parse_iso_date(request.form.get('end_date'))
            status_raw = (request.form.get('status') or 'OPEN').strip().upper()

            if not start_date or not end_date:
                flash('Tanggal mulai dan akhir periode wajib valid.', 'warning')
                return redirect(url_for('admin.manage_finance_settings'))
            if start_date > end_date:
                flash('Tanggal mulai periode tidak boleh lebih besar dari tanggal akhir.', 'warning')
                return redirect(url_for('admin.manage_finance_settings'))
            if status_raw not in FinancePeriodStatus.__members__:
                flash('Status periode tidak valid.', 'warning')
                return redirect(url_for('admin.manage_finance_settings'))
            if not name:
                name = f'{start_date.year}-{start_date.month:02d}'

            overlap = FinancePeriod.query.filter(
                FinancePeriod.tenant_id == tenant_id,
                FinancePeriod.start_date <= end_date,
                FinancePeriod.end_date >= start_date,
            ).first()
            if overlap:
                flash(
                    f'Periode bentrok dengan periode "{overlap.name}" ({overlap.start_date} s.d {overlap.end_date}).',
                    'warning'
                )
                return redirect(url_for('admin.manage_finance_settings'))

            db.session.add(FinancePeriod(
                tenant_id=tenant_id,
                name=name,
                start_date=start_date,
                end_date=end_date,
                status=FinancePeriodStatus[status_raw],
            ))
            db.session.commit()
            flash(f'Periode "{name}" berhasil dibuat.', 'success')
            return redirect(url_for('admin.manage_finance_settings'))

        if action == 'set_period_status':
            period_id = request.form.get('period_id', type=int)
            status_raw = (request.form.get('status') or '').strip().upper()
            if status_raw not in FinancePeriodStatus.__members__:
                flash('Status periode tidak valid.', 'warning')
                return redirect(url_for('admin.manage_finance_settings'))
            period = FinancePeriod.query.filter_by(id=period_id, tenant_id=tenant_id).first_or_404()
            period.status = FinancePeriodStatus[status_raw]
            if period.status in (FinancePeriodStatus.CLOSED, FinancePeriodStatus.LOCKED):
                period.closed_at = local_now()
                period.closed_by_user_id = current_user.id
            else:
                period.closed_at = None
                period.closed_by_user_id = None
            db.session.commit()
            flash(f'Status periode "{period.name}" diperbarui menjadi {period.status.value}.', 'success')
            return redirect(url_for('admin.manage_finance_settings'))

        if action == 'ensure_current_period':
            today = local_today()
            period_name = f'{today.year:04d}-{today.month:02d}'
            start_date = date(today.year, today.month, 1)
            if today.month == 12:
                end_date = date(today.year, 12, 31)
            else:
                next_month_start = date(today.year, today.month + 1, 1)
                end_date = next_month_start - timedelta(days=1)

            period = FinancePeriod.query.filter_by(tenant_id=tenant_id, name=period_name).first()
            if period:
                if period.status != FinancePeriodStatus.OPEN:
                    period.status = FinancePeriodStatus.OPEN
                    period.closed_at = None
                    period.closed_by_user_id = None
                    db.session.commit()
                    flash(f'Periode {period_name} sudah ada dan diubah ke status OPEN.', 'success')
                else:
                    flash(f'Periode {period_name} sudah tersedia dan OPEN.', 'info')
                return redirect(url_for('admin.manage_finance_settings'))

            db.session.add(FinancePeriod(
                tenant_id=tenant_id,
                name=period_name,
                start_date=start_date,
                end_date=end_date,
                status=FinancePeriodStatus.OPEN,
            ))
            db.session.commit()
            flash(f'Periode {period_name} berhasil dibuat dengan status OPEN.', 'success')
            return redirect(url_for('admin.manage_finance_settings'))

        if action == 'lock_old_periods':
            today = local_today()
            current_month_start = date(today.year, today.month, 1)
            periods = FinancePeriod.query.filter(
                FinancePeriod.tenant_id == tenant_id,
                FinancePeriod.end_date < current_month_start,
                FinancePeriod.status.in_([FinancePeriodStatus.OPEN, FinancePeriodStatus.CLOSED]),
            ).all()
            locked = 0
            for period in periods:
                period.status = FinancePeriodStatus.LOCKED
                period.closed_at = local_now()
                period.closed_by_user_id = current_user.id
                locked += 1
            db.session.commit()
            flash(f'Periode lama berhasil dikunci: {locked}.', 'success')
            return redirect(url_for('admin.manage_finance_settings'))

    settings = FinanceSetting.query.filter_by(tenant_id=tenant_id).first()
    accounts = FinanceAccount.query.filter_by(tenant_id=tenant_id, is_active=True).order_by(FinanceAccount.code.asc()).all()
    cash_bank_accounts = FinanceCashBankAccount.query.filter_by(tenant_id=tenant_id, is_active=True).order_by(FinanceCashBankAccount.account_name.asc()).all()
    periods = FinancePeriod.query.filter_by(tenant_id=tenant_id).order_by(FinancePeriod.start_date.desc(), FinancePeriod.id.desc()).all()
    finance_draft_count = FinanceJournal.query.filter_by(
        tenant_id=tenant_id,
        status=FinanceJournalStatus.DRAFT,
    ).count()
    return render_template(
        'admin/finance/settings.html',
        settings=settings,
        accounts=accounts,
        cash_bank_accounts=cash_bank_accounts,
        periods=periods,
        finance_draft_count=finance_draft_count,
        accounting_basis_options=list(FinanceAccountingBasis),
        period_status_options=list(FinancePeriodStatus),
    )


@admin_bp.route('/keuangan/kas-bank', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.TU)
def manage_finance_cash_bank():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        action = (request.form.get('action') or 'create_transaction').strip()

        if action == 'create_account':
            account_name = (request.form.get('account_name') or '').strip()
            account_type_raw = (request.form.get('account_type') or '').strip().upper()
            gl_account_id = request.form.get('gl_account_id', type=int)

            if not account_name or not gl_account_id:
                flash('Nama kas/bank dan GL account wajib diisi.', 'warning')
                return redirect(url_for('admin.manage_finance_cash_bank'))
            if account_type_raw not in FinanceCashBankAccountType.__members__:
                flash('Tipe kas/bank tidak valid.', 'warning')
                return redirect(url_for('admin.manage_finance_cash_bank'))

            gl_account = FinanceAccount.query.filter_by(
                id=gl_account_id,
                tenant_id=tenant_id,
                is_active=True,
            ).first()
            if not gl_account or gl_account.category != FinanceAccountCategory.ASSET:
                flash('GL account kas/bank harus akun ASSET aktif pada tenant ini.', 'danger')
                return redirect(url_for('admin.manage_finance_cash_bank'))

            existing = FinanceCashBankAccount.query.filter_by(
                tenant_id=tenant_id,
                account_name=account_name,
            ).first()
            if existing:
                flash(f'Kas/bank "{account_name}" sudah ada.', 'warning')
                return redirect(url_for('admin.manage_finance_cash_bank'))

            db.session.add(FinanceCashBankAccount(
                tenant_id=tenant_id,
                account_name=account_name,
                account_type=FinanceCashBankAccountType[account_type_raw],
                gl_account_id=gl_account.id,
                is_active=True,
            ))
            db.session.commit()
            flash(f'Kas/bank "{account_name}" berhasil ditambahkan.', 'success')
            return redirect(url_for('admin.manage_finance_cash_bank'))

        if action == 'toggle_account':
            account_id = request.form.get('account_id', type=int)
            cash_bank = FinanceCashBankAccount.query.filter_by(id=account_id, tenant_id=tenant_id).first_or_404()
            cash_bank.is_active = not cash_bank.is_active
            db.session.commit()
            flash(
                f'Kas/bank "{cash_bank.account_name}" {"diaktifkan" if cash_bank.is_active else "dinonaktifkan"}.',
                'success'
            )
            return redirect(url_for('admin.manage_finance_cash_bank'))

        trx_date = _parse_iso_date(request.form.get('trx_date'))
        cash_bank_account_id = request.form.get('cash_bank_account_id', type=int)
        trx_type = (request.form.get('trx_type') or '').strip().upper()
        amount = to_rupiah_int(request.form.get('amount'), default=0)
        counterpart_account_id = request.form.get('counterpart_account_id', type=int)
        description = (request.form.get('description') or '').strip()

        if not trx_date:
            flash('Tanggal transaksi wajib valid.', 'warning')
            return redirect(url_for('admin.manage_finance_cash_bank'))

        try:
            cash_bank_trx_id = create_cash_bank_transaction(
                tenant_id=tenant_id,
                trx_date=trx_date,
                cash_bank_account_id=cash_bank_account_id,
                trx_type=trx_type,
                amount=amount,
                counterpart_account_id=counterpart_account_id,
                description=description,
                actor_user_id=current_user.id,
            )
            flash(f'Transaksi kas/bank #{cash_bank_trx_id} berhasil dicatat.', 'success')
        except Exception as exc:
            db.session.rollback()
            flash(f'Gagal mencatat transaksi kas/bank: {exc}', 'danger')
        return redirect(url_for('admin.manage_finance_cash_bank'))

    asset_accounts = FinanceAccount.query.filter_by(
        tenant_id=tenant_id,
        category=FinanceAccountCategory.ASSET,
        is_active=True,
    ).order_by(FinanceAccount.code.asc()).all()
    accounts = FinanceAccount.query.filter_by(
        tenant_id=tenant_id,
        is_active=True,
    ).order_by(FinanceAccount.code.asc()).all()
    cash_bank_accounts = FinanceCashBankAccount.query.filter_by(
        tenant_id=tenant_id,
    ).order_by(FinanceCashBankAccount.is_active.desc(), FinanceCashBankAccount.account_name.asc()).all()
    transactions = FinanceCashBankTransaction.query.filter_by(
        tenant_id=tenant_id,
    ).order_by(FinanceCashBankTransaction.trx_date.desc(), FinanceCashBankTransaction.id.desc()).limit(100).all()

    return render_template(
        'admin/finance/cash_bank.html',
        today=local_today(),
        accounts=accounts,
        asset_accounts=asset_accounts,
        cash_bank_accounts=cash_bank_accounts,
        transactions=transactions,
        account_type_options=list(FinanceCashBankAccountType),
        trx_type_options=list(FinanceCashBankTransactionType),
    )


@admin_bp.route('/keuangan/jurnal')
@login_required
@role_required(UserRole.TU)
def finance_journals():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    status_filter = (request.args.get('status') or '').strip().upper()
    source_filter = (request.args.get('source_type') or '').strip().upper()
    start_date = _parse_iso_date(request.args.get('start_date'))
    end_date = _parse_iso_date(request.args.get('end_date'))
    query_text = (request.args.get('q') or '').strip()

    journals_query = FinanceJournal.query.filter_by(tenant_id=tenant_id)
    if status_filter and status_filter in FinanceJournalStatus.__members__:
        journals_query = journals_query.filter(FinanceJournal.status == FinanceJournalStatus[status_filter])
    if source_filter and source_filter in FinanceJournalSourceType.__members__:
        journals_query = journals_query.filter(FinanceJournal.source_type == FinanceJournalSourceType[source_filter])
    if start_date:
        journals_query = journals_query.filter(FinanceJournal.journal_date >= start_date)
    if end_date:
        journals_query = journals_query.filter(FinanceJournal.journal_date <= end_date)
    if query_text:
        journals_query = journals_query.filter(
            or_(
                FinanceJournal.journal_no.ilike(f'%{query_text}%'),
                FinanceJournal.description.ilike(f'%{query_text}%'),
            )
        )

    journals = journals_query.order_by(
        FinanceJournal.journal_date.desc(),
        FinanceJournal.id.desc(),
    ).limit(300).all()

    return render_template(
        'admin/finance/journals.html',
        journals=journals,
        status_filter=status_filter,
        source_filter=source_filter,
        start_date=start_date,
        end_date=end_date,
        query_text=query_text,
        status_options=list(FinanceJournalStatus),
        source_type_options=list(FinanceJournalSourceType),
    )


@admin_bp.route('/keuangan/jurnal/export')
@login_required
@role_required(UserRole.TU)
def finance_journals_export():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    status_filter = (request.args.get('status') or '').strip().upper()
    source_filter = (request.args.get('source_type') or '').strip().upper()
    start_date = _parse_iso_date(request.args.get('start_date'))
    end_date = _parse_iso_date(request.args.get('end_date'))
    query_text = (request.args.get('q') or '').strip()

    journals_query = FinanceJournal.query.filter_by(tenant_id=tenant_id)
    if status_filter and status_filter in FinanceJournalStatus.__members__:
        journals_query = journals_query.filter(FinanceJournal.status == FinanceJournalStatus[status_filter])
    if source_filter and source_filter in FinanceJournalSourceType.__members__:
        journals_query = journals_query.filter(FinanceJournal.source_type == FinanceJournalSourceType[source_filter])
    if start_date:
        journals_query = journals_query.filter(FinanceJournal.journal_date >= start_date)
    if end_date:
        journals_query = journals_query.filter(FinanceJournal.journal_date <= end_date)
    if query_text:
        journals_query = journals_query.filter(
            or_(
                FinanceJournal.journal_no.ilike(f'%{query_text}%'),
                FinanceJournal.description.ilike(f'%{query_text}%'),
            )
        )

    journals = journals_query.order_by(FinanceJournal.journal_date.asc(), FinanceJournal.id.asc()).all()
    rows = [
        [
            journal.journal_date,
            journal.journal_no,
            journal.status.value,
            journal.source_type.value if journal.source_type else '',
            journal.source_id or '',
            journal.description or '',
            journal.created_by.full_name if journal.created_by else '',
            journal.approved_by.full_name if journal.approved_by else '',
            journal.posted_at.strftime('%Y-%m-%d %H:%M:%S') if journal.posted_at else '',
        ]
        for journal in journals
    ]
    return _csv_response(
        'finance_jurnal_umum.csv',
        ['Tanggal', 'No Jurnal', 'Status', 'Source Type', 'Source ID', 'Deskripsi', 'Dibuat Oleh', 'Disetujui Oleh', 'Posted At'],
        rows,
    )


@admin_bp.route('/keuangan/jurnal/export-xlsx')
@login_required
@role_required(UserRole.TU)
def finance_journals_export_xlsx():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    status_filter = (request.args.get('status') or '').strip().upper()
    source_filter = (request.args.get('source_type') or '').strip().upper()
    start_date = _parse_iso_date(request.args.get('start_date'))
    end_date = _parse_iso_date(request.args.get('end_date'))
    query_text = (request.args.get('q') or '').strip()

    journals_query = FinanceJournal.query.filter_by(tenant_id=tenant_id)
    if status_filter and status_filter in FinanceJournalStatus.__members__:
        journals_query = journals_query.filter(FinanceJournal.status == FinanceJournalStatus[status_filter])
    if source_filter and source_filter in FinanceJournalSourceType.__members__:
        journals_query = journals_query.filter(FinanceJournal.source_type == FinanceJournalSourceType[source_filter])
    if start_date:
        journals_query = journals_query.filter(FinanceJournal.journal_date >= start_date)
    if end_date:
        journals_query = journals_query.filter(FinanceJournal.journal_date <= end_date)
    if query_text:
        journals_query = journals_query.filter(
            or_(
                FinanceJournal.journal_no.ilike(f'%{query_text}%'),
                FinanceJournal.description.ilike(f'%{query_text}%'),
            )
        )

    rows = [
        [
            journal.journal_date,
            journal.journal_no,
            journal.status.value,
            journal.source_type.value if journal.source_type else '',
            journal.source_id or '',
            journal.description or '',
            journal.created_by.full_name if journal.created_by else '',
            journal.approved_by.full_name if journal.approved_by else '',
            journal.posted_at.strftime('%Y-%m-%d %H:%M:%S') if journal.posted_at else '',
        ]
        for journal in journals_query.order_by(FinanceJournal.journal_date.asc(), FinanceJournal.id.asc()).all()
    ]
    return _xlsx_response(
        'finance_jurnal_umum.xlsx',
        'Jurnal Umum',
        ['Tanggal', 'No Jurnal', 'Status', 'Source Type', 'Source ID', 'Deskripsi', 'Dibuat Oleh', 'Disetujui Oleh', 'Posted At'],
        rows,
    )


@admin_bp.route('/keuangan/jurnal/<int:journal_id>')
@login_required
@role_required(UserRole.TU)
def finance_journal_detail(journal_id):
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    journal = FinanceJournal.query.filter_by(id=journal_id, tenant_id=tenant_id).first_or_404()
    lines = FinanceJournalLine.query.filter_by(
        tenant_id=tenant_id,
        journal_id=journal.id,
    ).order_by(FinanceJournalLine.id.asc()).all()
    debit_total = sum(int(line.amount or 0) for line in lines if line.entry_side == FinanceEntrySide.DEBIT)
    credit_total = sum(int(line.amount or 0) for line in lines if line.entry_side == FinanceEntrySide.CREDIT)

    return render_template(
        'admin/finance/journal_detail.html',
        journal=journal,
        lines=lines,
        debit_total=debit_total,
        credit_total=credit_total,
    )


@admin_bp.route('/keuangan/jurnal/<int:journal_id>/print')
@login_required
@role_required(UserRole.TU)
def finance_journal_detail_print(journal_id):
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    journal = FinanceJournal.query.filter_by(id=journal_id, tenant_id=tenant_id).first_or_404()
    lines = FinanceJournalLine.query.filter_by(
        tenant_id=tenant_id,
        journal_id=journal.id,
    ).order_by(FinanceJournalLine.id.asc()).all()
    debit_total = sum(int(line.amount or 0) for line in lines if line.entry_side == FinanceEntrySide.DEBIT)
    credit_total = sum(int(line.amount or 0) for line in lines if line.entry_side == FinanceEntrySide.CREDIT)
    return render_template(
        'admin/finance/print_journal_detail.html',
        journal=journal,
        lines=lines,
        debit_total=debit_total,
        credit_total=credit_total,
        signers=_report_signers_from_request(),
    )


@admin_bp.route('/keuangan/jurnal/<int:journal_id>/reverse', methods=['POST'])
@login_required
@role_required(UserRole.TU)
def finance_journal_reverse(journal_id):
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    reason = (request.form.get('reason') or '').strip()
    if not reason:
        flash('Alasan reversal wajib diisi.', 'warning')
        return redirect(url_for('admin.finance_journal_detail', journal_id=journal_id))

    try:
        reversal_id = reverse_journal(
            tenant_id=tenant_id,
            journal_id=journal_id,
            reason=reason,
            actor_user_id=current_user.id,
        )
        flash(f'Jurnal berhasil direversal. Jurnal reversal: #{reversal_id}.', 'success')
        return redirect(url_for('admin.finance_journal_detail', journal_id=reversal_id))
    except Exception as exc:
        db.session.rollback()
        flash(f'Gagal reversal jurnal: {exc}', 'danger')
        return redirect(url_for('admin.finance_journal_detail', journal_id=journal_id))


def _default_report_dates():
    today = local_today()
    return date(today.year, today.month, 1), today


def _posted_finance_lines_query(tenant_id, start_date, end_date):
    query = (
        FinanceJournalLine.query
        .join(FinanceJournal, FinanceJournal.id == FinanceJournalLine.journal_id)
        .filter(
            FinanceJournalLine.tenant_id == tenant_id,
            FinanceJournal.tenant_id == tenant_id,
            FinanceJournal.status == FinanceJournalStatus.POSTED,
        )
    )
    if start_date:
        query = query.filter(FinanceJournal.journal_date >= start_date)
    if end_date:
        query = query.filter(FinanceJournal.journal_date <= end_date)
    return query


def _csv_response(filename, header, rows):
    output = StringIO()
    output.write('\ufeff')
    writer = csv.writer(output)
    writer.writerow(header)
    writer.writerows(rows)
    return Response(
        output.getvalue(),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


def _xlsx_response(filename, sheet_title, header, rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_title[:31]
    worksheet.append(header)
    for row in rows:
        worksheet.append(row)
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or '')) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 40)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


def _report_signers_from_request():
    return [
        {
            'title': (request.args.get('signer1_title') or 'Mengetahui').strip(),
            'name': (request.args.get('signer1_name') or '').strip(),
        },
        {
            'title': (request.args.get('signer2_title') or 'Dibuat oleh').strip(),
            'name': (request.args.get('signer2_name') or '').strip(),
        },
    ]


def _trial_balance_data(tenant_id, start_date, end_date):
    rows_by_account_id = {
        account.id: {
            'account': account,
            'debit_total': 0,
            'credit_total': 0,
            'ending_debit': 0,
            'ending_credit': 0,
        }
        for account in FinanceAccount.query.filter_by(tenant_id=tenant_id).order_by(FinanceAccount.code.asc()).all()
    }
    for line in _posted_finance_lines_query(tenant_id, start_date, end_date).all():
        row = rows_by_account_id.get(line.account_id)
        if not row:
            continue
        amount = int(line.amount or 0)
        if line.entry_side == FinanceEntrySide.DEBIT:
            row['debit_total'] += amount
        elif line.entry_side == FinanceEntrySide.CREDIT:
            row['credit_total'] += amount

    report_rows = []
    total_debit = 0
    total_credit = 0
    for row in rows_by_account_id.values():
        balance = row['debit_total'] - row['credit_total']
        if balance > 0:
            row['ending_debit'] = balance
            total_debit += balance
        elif balance < 0:
            row['ending_credit'] = abs(balance)
            total_credit += abs(balance)
        if row['debit_total'] or row['credit_total']:
            report_rows.append(row)
    return report_rows, total_debit, total_credit


def _income_statement_data(tenant_id, start_date, end_date):
    rows_by_account_id = {
        account.id: {
            'account': account,
            'debit_total': 0,
            'credit_total': 0,
            'amount': 0,
        }
        for account in FinanceAccount.query.filter(
            FinanceAccount.tenant_id == tenant_id,
            FinanceAccount.category.in_([FinanceAccountCategory.REVENUE, FinanceAccountCategory.EXPENSE]),
        ).order_by(FinanceAccount.code.asc()).all()
    }
    lines = (
        _posted_finance_lines_query(tenant_id, start_date, end_date)
        .join(FinanceAccount, FinanceAccount.id == FinanceJournalLine.account_id)
        .filter(
            FinanceAccount.tenant_id == tenant_id,
            FinanceAccount.category.in_([FinanceAccountCategory.REVENUE, FinanceAccountCategory.EXPENSE]),
        )
        .all()
    )
    for line in lines:
        row = rows_by_account_id.get(line.account_id)
        if not row:
            continue
        amount = int(line.amount or 0)
        if line.entry_side == FinanceEntrySide.DEBIT:
            row['debit_total'] += amount
        elif line.entry_side == FinanceEntrySide.CREDIT:
            row['credit_total'] += amount

    revenue_rows = []
    expense_rows = []
    total_revenue = 0
    total_expense = 0
    for row in rows_by_account_id.values():
        account = row['account']
        if account.category == FinanceAccountCategory.REVENUE:
            row['amount'] = row['credit_total'] - row['debit_total']
            if row['amount']:
                total_revenue += row['amount']
                revenue_rows.append(row)
        elif account.category == FinanceAccountCategory.EXPENSE:
            row['amount'] = row['debit_total'] - row['credit_total']
            if row['amount']:
                total_expense += row['amount']
                expense_rows.append(row)
    return revenue_rows, expense_rows, total_revenue, total_expense, total_revenue - total_expense


def _financial_position_data(tenant_id, as_of_date):
    balance_categories = [
        FinanceAccountCategory.ASSET,
        FinanceAccountCategory.LIABILITY,
        FinanceAccountCategory.EQUITY,
    ]
    rows_by_account_id = {
        account.id: {
            'account': account,
            'debit_total': 0,
            'credit_total': 0,
            'amount': 0,
        }
        for account in FinanceAccount.query.filter(
            FinanceAccount.tenant_id == tenant_id,
            FinanceAccount.category.in_(balance_categories),
        ).order_by(FinanceAccount.code.asc()).all()
    }

    lines = (
        _posted_finance_lines_query(tenant_id, None, as_of_date)
        .join(FinanceAccount, FinanceAccount.id == FinanceJournalLine.account_id)
        .filter(
            FinanceAccount.tenant_id == tenant_id,
            FinanceAccount.category.in_(balance_categories),
        )
        .all()
    )
    for line in lines:
        row = rows_by_account_id.get(line.account_id)
        if not row:
            continue
        amount = int(line.amount or 0)
        if line.entry_side == FinanceEntrySide.DEBIT:
            row['debit_total'] += amount
        elif line.entry_side == FinanceEntrySide.CREDIT:
            row['credit_total'] += amount

    asset_rows = []
    liability_rows = []
    equity_rows = []
    total_assets = 0
    total_liabilities = 0
    total_equity = 0
    for row in rows_by_account_id.values():
        account = row['account']
        if account.category == FinanceAccountCategory.ASSET:
            row['amount'] = row['debit_total'] - row['credit_total']
            if row['amount']:
                total_assets += row['amount']
                asset_rows.append(row)
        elif account.category == FinanceAccountCategory.LIABILITY:
            row['amount'] = row['credit_total'] - row['debit_total']
            if row['amount']:
                total_liabilities += row['amount']
                liability_rows.append(row)
        elif account.category == FinanceAccountCategory.EQUITY:
            row['amount'] = row['credit_total'] - row['debit_total']
            if row['amount']:
                total_equity += row['amount']
                equity_rows.append(row)

    _, _, _, _, net_income = _income_statement_data(tenant_id, None, as_of_date)
    if net_income:
        total_equity += net_income
        equity_rows.append({
            'account': None,
            'code': 'LR-BERJALAN',
            'name': 'Laba/Rugi Berjalan',
            'amount': net_income,
        })

    return {
        'asset_rows': asset_rows,
        'liability_rows': liability_rows,
        'equity_rows': equity_rows,
        'total_assets': total_assets,
        'total_liabilities': total_liabilities,
        'total_equity': total_equity,
        'net_income': net_income,
        'total_liabilities_equity': total_liabilities + total_equity,
    }


def _ledger_data(tenant_id, selected_account, start_date, end_date):
    opening_debit = 0
    opening_credit = 0
    for line in _posted_finance_lines_query(tenant_id, None, start_date - timedelta(days=1)).filter(
        FinanceJournalLine.account_id == selected_account.id,
    ).all():
        amount = int(line.amount or 0)
        if line.entry_side == FinanceEntrySide.DEBIT:
            opening_debit += amount
        elif line.entry_side == FinanceEntrySide.CREDIT:
            opening_credit += amount

    running_balance = opening_debit - opening_credit
    ledger_rows = []
    period_lines = (
        _posted_finance_lines_query(tenant_id, start_date, end_date)
        .filter(FinanceJournalLine.account_id == selected_account.id)
        .order_by(FinanceJournal.journal_date.asc(), FinanceJournal.id.asc(), FinanceJournalLine.id.asc())
        .all()
    )
    for line in period_lines:
        debit = int(line.amount or 0) if line.entry_side == FinanceEntrySide.DEBIT else 0
        credit = int(line.amount or 0) if line.entry_side == FinanceEntrySide.CREDIT else 0
        running_balance += debit - credit
        ledger_rows.append({'line': line, 'debit': debit, 'credit': credit, 'balance': running_balance})

    closing_debit = running_balance if running_balance >= 0 else 0
    closing_credit = abs(running_balance) if running_balance < 0 else 0
    return opening_debit, opening_credit, ledger_rows, closing_debit, closing_credit


@admin_bp.route('/keuangan/laporan/neraca-saldo')
@login_required
@role_required(UserRole.TU)
def finance_trial_balance():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    default_start, default_end = _default_report_dates()
    start_date = _parse_iso_date(request.args.get('start_date')) or default_start
    end_date = _parse_iso_date(request.args.get('end_date')) or default_end

    accounts = FinanceAccount.query.filter_by(
        tenant_id=tenant_id,
    ).order_by(FinanceAccount.code.asc()).all()
    rows_by_account_id = {
        account.id: {
            'account': account,
            'debit_total': 0,
            'credit_total': 0,
            'ending_debit': 0,
            'ending_credit': 0,
        }
        for account in accounts
    }

    lines = _posted_finance_lines_query(tenant_id, start_date, end_date).all()
    for line in lines:
        row = rows_by_account_id.get(line.account_id)
        if not row:
            continue
        amount = int(line.amount or 0)
        if line.entry_side == FinanceEntrySide.DEBIT:
            row['debit_total'] += amount
        elif line.entry_side == FinanceEntrySide.CREDIT:
            row['credit_total'] += amount

    report_rows = []
    total_debit = 0
    total_credit = 0
    for row in rows_by_account_id.values():
        balance = row['debit_total'] - row['credit_total']
        if balance > 0:
            row['ending_debit'] = balance
            total_debit += balance
        elif balance < 0:
            row['ending_credit'] = abs(balance)
            total_credit += abs(balance)
        if row['debit_total'] or row['credit_total']:
            report_rows.append(row)

    return render_template(
        'admin/finance/trial_balance.html',
        rows=report_rows,
        start_date=start_date,
        end_date=end_date,
        total_debit=total_debit,
        total_credit=total_credit,
    )


@admin_bp.route('/keuangan/laporan/neraca-saldo/export')
@login_required
@role_required(UserRole.TU)
def finance_trial_balance_export():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    default_start, default_end = _default_report_dates()
    start_date = _parse_iso_date(request.args.get('start_date')) or default_start
    end_date = _parse_iso_date(request.args.get('end_date')) or default_end

    rows_by_account_id = {
        account.id: {
            'account': account,
            'debit_total': 0,
            'credit_total': 0,
            'ending_debit': 0,
            'ending_credit': 0,
        }
        for account in FinanceAccount.query.filter_by(tenant_id=tenant_id).order_by(FinanceAccount.code.asc()).all()
    }

    for line in _posted_finance_lines_query(tenant_id, start_date, end_date).all():
        row = rows_by_account_id.get(line.account_id)
        if not row:
            continue
        amount = int(line.amount or 0)
        if line.entry_side == FinanceEntrySide.DEBIT:
            row['debit_total'] += amount
        elif line.entry_side == FinanceEntrySide.CREDIT:
            row['credit_total'] += amount

    rows = []
    for row in rows_by_account_id.values():
        balance = row['debit_total'] - row['credit_total']
        if balance > 0:
            row['ending_debit'] = balance
        elif balance < 0:
            row['ending_credit'] = abs(balance)
        if row['debit_total'] or row['credit_total']:
            account = row['account']
            rows.append([
                start_date,
                end_date,
                account.code,
                account.name,
                account.category.value,
                row['ending_debit'],
                row['ending_credit'],
            ])

    return _csv_response(
        'finance_neraca_saldo.csv',
        ['Mulai', 'Selesai', 'Kode Akun', 'Nama Akun', 'Kategori', 'Debit', 'Credit'],
        rows,
    )


@admin_bp.route('/keuangan/laporan/neraca-saldo/export-xlsx')
@login_required
@role_required(UserRole.TU)
def finance_trial_balance_export_xlsx():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))
    default_start, default_end = _default_report_dates()
    start_date = _parse_iso_date(request.args.get('start_date')) or default_start
    end_date = _parse_iso_date(request.args.get('end_date')) or default_end
    report_rows, _, _ = _trial_balance_data(tenant_id, start_date, end_date)
    rows = [
        [start_date, end_date, row['account'].code, row['account'].name, row['account'].category.value, row['ending_debit'], row['ending_credit']]
        for row in report_rows
    ]
    return _xlsx_response(
        'finance_neraca_saldo.xlsx',
        'Neraca Saldo',
        ['Mulai', 'Selesai', 'Kode Akun', 'Nama Akun', 'Kategori', 'Debit', 'Credit'],
        rows,
    )


@admin_bp.route('/keuangan/laporan/neraca-saldo/print')
@login_required
@role_required(UserRole.TU)
def finance_trial_balance_print():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))
    default_start, default_end = _default_report_dates()
    start_date = _parse_iso_date(request.args.get('start_date')) or default_start
    end_date = _parse_iso_date(request.args.get('end_date')) or default_end
    rows, total_debit, total_credit = _trial_balance_data(tenant_id, start_date, end_date)
    return render_template(
        'admin/finance/print_trial_balance.html',
        rows=rows,
        start_date=start_date,
        end_date=end_date,
        total_debit=total_debit,
        total_credit=total_credit,
        signers=_report_signers_from_request(),
    )


@admin_bp.route('/keuangan/laporan/buku-besar')
@login_required
@role_required(UserRole.TU)
def finance_general_ledger():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    default_start, default_end = _default_report_dates()
    start_date = _parse_iso_date(request.args.get('start_date')) or default_start
    end_date = _parse_iso_date(request.args.get('end_date')) or default_end
    account_id = request.args.get('account_id', type=int)

    accounts = FinanceAccount.query.filter_by(
        tenant_id=tenant_id,
    ).order_by(FinanceAccount.code.asc()).all()

    selected_account = None
    if account_id:
        selected_account = FinanceAccount.query.filter_by(
            id=account_id,
            tenant_id=tenant_id,
        ).first()
        if not selected_account:
            flash('Akun buku besar tidak ditemukan untuk tenant ini.', 'warning')
            return redirect(url_for('admin.finance_general_ledger'))

    opening_debit = 0
    opening_credit = 0
    ledger_rows = []
    closing_debit = 0
    closing_credit = 0

    if selected_account:
        opening_lines = _posted_finance_lines_query(tenant_id, None, start_date - timedelta(days=1)).filter(
            FinanceJournalLine.account_id == selected_account.id,
        ).all()
        for line in opening_lines:
            amount = int(line.amount or 0)
            if line.entry_side == FinanceEntrySide.DEBIT:
                opening_debit += amount
            elif line.entry_side == FinanceEntrySide.CREDIT:
                opening_credit += amount

        running_balance = opening_debit - opening_credit
        period_lines = (
            _posted_finance_lines_query(tenant_id, start_date, end_date)
            .filter(FinanceJournalLine.account_id == selected_account.id)
            .order_by(FinanceJournal.journal_date.asc(), FinanceJournal.id.asc(), FinanceJournalLine.id.asc())
            .all()
        )
        for line in period_lines:
            debit = int(line.amount or 0) if line.entry_side == FinanceEntrySide.DEBIT else 0
            credit = int(line.amount or 0) if line.entry_side == FinanceEntrySide.CREDIT else 0
            running_balance += debit - credit
            ledger_rows.append({
                'line': line,
                'debit': debit,
                'credit': credit,
                'balance': running_balance,
            })

        if running_balance >= 0:
            closing_debit = running_balance
        else:
            closing_credit = abs(running_balance)

    return render_template(
        'admin/finance/general_ledger.html',
        accounts=accounts,
        selected_account=selected_account,
        account_id=account_id,
        start_date=start_date,
        end_date=end_date,
        opening_debit=opening_debit,
        opening_credit=opening_credit,
        ledger_rows=ledger_rows,
        closing_debit=closing_debit,
        closing_credit=closing_credit,
    )


@admin_bp.route('/keuangan/laporan/buku-besar/export')
@login_required
@role_required(UserRole.TU)
def finance_general_ledger_export():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    default_start, default_end = _default_report_dates()
    start_date = _parse_iso_date(request.args.get('start_date')) or default_start
    end_date = _parse_iso_date(request.args.get('end_date')) or default_end
    account_id = request.args.get('account_id', type=int)

    selected_account = FinanceAccount.query.filter_by(
        id=account_id,
        tenant_id=tenant_id,
    ).first() if account_id else None
    if not selected_account:
        flash('Pilih akun buku besar sebelum export.', 'warning')
        return redirect(url_for('admin.finance_general_ledger', start_date=start_date, end_date=end_date))

    opening_debit = 0
    opening_credit = 0
    for line in _posted_finance_lines_query(tenant_id, None, start_date - timedelta(days=1)).filter(
        FinanceJournalLine.account_id == selected_account.id,
    ).all():
        amount = int(line.amount or 0)
        if line.entry_side == FinanceEntrySide.DEBIT:
            opening_debit += amount
        elif line.entry_side == FinanceEntrySide.CREDIT:
            opening_credit += amount

    rows = [[start_date, end_date, selected_account.code, selected_account.name, 'SALDO AWAL', '', opening_debit, opening_credit, opening_debit - opening_credit, '']]
    running_balance = opening_debit - opening_credit
    period_lines = (
        _posted_finance_lines_query(tenant_id, start_date, end_date)
        .filter(FinanceJournalLine.account_id == selected_account.id)
        .order_by(FinanceJournal.journal_date.asc(), FinanceJournal.id.asc(), FinanceJournalLine.id.asc())
        .all()
    )
    for line in period_lines:
        debit = int(line.amount or 0) if line.entry_side == FinanceEntrySide.DEBIT else 0
        credit = int(line.amount or 0) if line.entry_side == FinanceEntrySide.CREDIT else 0
        running_balance += debit - credit
        journal = line.journal
        rows.append([
            start_date,
            end_date,
            selected_account.code,
            selected_account.name,
            journal.journal_date if journal else '',
            journal.journal_no if journal else '',
            debit,
            credit,
            running_balance,
            journal.description if journal else line.memo or '',
        ])

    return _csv_response(
        f'finance_buku_besar_{selected_account.code}.csv',
        ['Mulai', 'Selesai', 'Kode Akun', 'Nama Akun', 'Tanggal', 'No Jurnal', 'Debit', 'Credit', 'Saldo', 'Deskripsi'],
        rows,
    )


@admin_bp.route('/keuangan/laporan/buku-besar/export-xlsx')
@login_required
@role_required(UserRole.TU)
def finance_general_ledger_export_xlsx():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))
    default_start, default_end = _default_report_dates()
    start_date = _parse_iso_date(request.args.get('start_date')) or default_start
    end_date = _parse_iso_date(request.args.get('end_date')) or default_end
    account_id = request.args.get('account_id', type=int)
    selected_account = FinanceAccount.query.filter_by(id=account_id, tenant_id=tenant_id).first() if account_id else None
    if not selected_account:
        flash('Pilih akun buku besar sebelum export.', 'warning')
        return redirect(url_for('admin.finance_general_ledger', start_date=start_date, end_date=end_date))

    opening_debit, opening_credit, ledger_rows, _, _ = _ledger_data(tenant_id, selected_account, start_date, end_date)
    rows = [[start_date, end_date, selected_account.code, selected_account.name, 'SALDO AWAL', '', opening_debit, opening_credit, opening_debit - opening_credit, '']]
    for row in ledger_rows:
        journal = row['line'].journal
        rows.append([
            start_date,
            end_date,
            selected_account.code,
            selected_account.name,
            journal.journal_date if journal else '',
            journal.journal_no if journal else '',
            row['debit'],
            row['credit'],
            row['balance'],
            journal.description if journal else row['line'].memo or '',
        ])
    return _xlsx_response(
        f'finance_buku_besar_{selected_account.code}.xlsx',
        'Buku Besar',
        ['Mulai', 'Selesai', 'Kode Akun', 'Nama Akun', 'Tanggal', 'No Jurnal', 'Debit', 'Credit', 'Saldo', 'Deskripsi'],
        rows,
    )


@admin_bp.route('/keuangan/laporan/buku-besar/print')
@login_required
@role_required(UserRole.TU)
def finance_general_ledger_print():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))
    default_start, default_end = _default_report_dates()
    start_date = _parse_iso_date(request.args.get('start_date')) or default_start
    end_date = _parse_iso_date(request.args.get('end_date')) or default_end
    account_id = request.args.get('account_id', type=int)
    selected_account = FinanceAccount.query.filter_by(id=account_id, tenant_id=tenant_id).first() if account_id else None
    if not selected_account:
        flash('Pilih akun buku besar sebelum print.', 'warning')
        return redirect(url_for('admin.finance_general_ledger', start_date=start_date, end_date=end_date))
    opening_debit, opening_credit, ledger_rows, closing_debit, closing_credit = _ledger_data(
        tenant_id, selected_account, start_date, end_date
    )
    return render_template(
        'admin/finance/print_general_ledger.html',
        selected_account=selected_account,
        start_date=start_date,
        end_date=end_date,
        opening_debit=opening_debit,
        opening_credit=opening_credit,
        ledger_rows=ledger_rows,
        closing_debit=closing_debit,
        closing_credit=closing_credit,
        signers=_report_signers_from_request(),
    )


@admin_bp.route('/keuangan/laporan/laba-rugi')
@login_required
@role_required(UserRole.TU)
def finance_income_statement():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    default_start, default_end = _default_report_dates()
    start_date = _parse_iso_date(request.args.get('start_date')) or default_start
    end_date = _parse_iso_date(request.args.get('end_date')) or default_end

    accounts = FinanceAccount.query.filter(
        FinanceAccount.tenant_id == tenant_id,
        FinanceAccount.category.in_([FinanceAccountCategory.REVENUE, FinanceAccountCategory.EXPENSE]),
    ).order_by(FinanceAccount.code.asc()).all()
    rows_by_account_id = {
        account.id: {
            'account': account,
            'debit_total': 0,
            'credit_total': 0,
            'amount': 0,
        }
        for account in accounts
    }

    lines = (
        _posted_finance_lines_query(tenant_id, start_date, end_date)
        .join(FinanceAccount, FinanceAccount.id == FinanceJournalLine.account_id)
        .filter(
            FinanceAccount.tenant_id == tenant_id,
            FinanceAccount.category.in_([FinanceAccountCategory.REVENUE, FinanceAccountCategory.EXPENSE]),
        )
        .all()
    )
    for line in lines:
        row = rows_by_account_id.get(line.account_id)
        if not row:
            continue
        amount = int(line.amount or 0)
        if line.entry_side == FinanceEntrySide.DEBIT:
            row['debit_total'] += amount
        elif line.entry_side == FinanceEntrySide.CREDIT:
            row['credit_total'] += amount

    revenue_rows = []
    expense_rows = []
    total_revenue = 0
    total_expense = 0
    for row in rows_by_account_id.values():
        account = row['account']
        if account.category == FinanceAccountCategory.REVENUE:
            row['amount'] = row['credit_total'] - row['debit_total']
            if row['amount']:
                total_revenue += row['amount']
                revenue_rows.append(row)
        elif account.category == FinanceAccountCategory.EXPENSE:
            row['amount'] = row['debit_total'] - row['credit_total']
            if row['amount']:
                total_expense += row['amount']
                expense_rows.append(row)

    return render_template(
        'admin/finance/income_statement.html',
        start_date=start_date,
        end_date=end_date,
        revenue_rows=revenue_rows,
        expense_rows=expense_rows,
        total_revenue=total_revenue,
        total_expense=total_expense,
        net_income=total_revenue - total_expense,
    )


@admin_bp.route('/keuangan/laporan/laba-rugi/export')
@login_required
@role_required(UserRole.TU)
def finance_income_statement_export():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    default_start, default_end = _default_report_dates()
    start_date = _parse_iso_date(request.args.get('start_date')) or default_start
    end_date = _parse_iso_date(request.args.get('end_date')) or default_end

    rows_by_account_id = {
        account.id: {
            'account': account,
            'debit_total': 0,
            'credit_total': 0,
        }
        for account in FinanceAccount.query.filter(
            FinanceAccount.tenant_id == tenant_id,
            FinanceAccount.category.in_([FinanceAccountCategory.REVENUE, FinanceAccountCategory.EXPENSE]),
        ).order_by(FinanceAccount.code.asc()).all()
    }

    lines = (
        _posted_finance_lines_query(tenant_id, start_date, end_date)
        .join(FinanceAccount, FinanceAccount.id == FinanceJournalLine.account_id)
        .filter(
            FinanceAccount.tenant_id == tenant_id,
            FinanceAccount.category.in_([FinanceAccountCategory.REVENUE, FinanceAccountCategory.EXPENSE]),
        )
        .all()
    )
    for line in lines:
        row = rows_by_account_id.get(line.account_id)
        if not row:
            continue
        amount = int(line.amount or 0)
        if line.entry_side == FinanceEntrySide.DEBIT:
            row['debit_total'] += amount
        elif line.entry_side == FinanceEntrySide.CREDIT:
            row['credit_total'] += amount

    rows = []
    total_revenue = 0
    total_expense = 0
    for row in rows_by_account_id.values():
        account = row['account']
        if account.category == FinanceAccountCategory.REVENUE:
            amount = row['credit_total'] - row['debit_total']
            section = 'Pendapatan'
            total_revenue += amount
        else:
            amount = row['debit_total'] - row['credit_total']
            section = 'Beban'
            total_expense += amount
        if amount:
            rows.append([start_date, end_date, section, account.code, account.name, amount])

    rows.append([start_date, end_date, 'Ringkasan', '', 'Total Pendapatan', total_revenue])
    rows.append([start_date, end_date, 'Ringkasan', '', 'Total Beban', total_expense])
    rows.append([start_date, end_date, 'Ringkasan', '', 'Laba/Rugi Bersih', total_revenue - total_expense])

    return _csv_response(
        'finance_laba_rugi.csv',
        ['Mulai', 'Selesai', 'Bagian', 'Kode Akun', 'Nama Akun', 'Nominal'],
        rows,
    )


@admin_bp.route('/keuangan/laporan/laba-rugi/export-xlsx')
@login_required
@role_required(UserRole.TU)
def finance_income_statement_export_xlsx():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))
    default_start, default_end = _default_report_dates()
    start_date = _parse_iso_date(request.args.get('start_date')) or default_start
    end_date = _parse_iso_date(request.args.get('end_date')) or default_end
    revenue_rows, expense_rows, total_revenue, total_expense, net_income = _income_statement_data(
        tenant_id, start_date, end_date
    )
    rows = []
    for row in revenue_rows:
        rows.append([start_date, end_date, 'Pendapatan', row['account'].code, row['account'].name, row['amount']])
    for row in expense_rows:
        rows.append([start_date, end_date, 'Beban', row['account'].code, row['account'].name, row['amount']])
    rows.append([start_date, end_date, 'Ringkasan', '', 'Total Pendapatan', total_revenue])
    rows.append([start_date, end_date, 'Ringkasan', '', 'Total Beban', total_expense])
    rows.append([start_date, end_date, 'Ringkasan', '', 'Laba/Rugi Bersih', net_income])
    return _xlsx_response(
        'finance_laba_rugi.xlsx',
        'Laba Rugi',
        ['Mulai', 'Selesai', 'Bagian', 'Kode Akun', 'Nama Akun', 'Nominal'],
        rows,
    )


@admin_bp.route('/keuangan/laporan/laba-rugi/print')
@login_required
@role_required(UserRole.TU)
def finance_income_statement_print():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))
    default_start, default_end = _default_report_dates()
    start_date = _parse_iso_date(request.args.get('start_date')) or default_start
    end_date = _parse_iso_date(request.args.get('end_date')) or default_end
    revenue_rows, expense_rows, total_revenue, total_expense, net_income = _income_statement_data(
        tenant_id, start_date, end_date
    )
    return render_template(
        'admin/finance/print_income_statement.html',
        start_date=start_date,
        end_date=end_date,
        revenue_rows=revenue_rows,
        expense_rows=expense_rows,
        total_revenue=total_revenue,
        total_expense=total_expense,
        net_income=net_income,
        signers=_report_signers_from_request(),
    )


@admin_bp.route('/keuangan/laporan/posisi-keuangan')
@login_required
@role_required(UserRole.TU)
def finance_financial_position():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))
    as_of_date = _parse_iso_date(request.args.get('as_of_date')) or local_today()
    report = _financial_position_data(tenant_id, as_of_date)
    return render_template(
        'admin/finance/financial_position.html',
        as_of_date=as_of_date,
        **report,
    )


def _financial_position_export_rows(report, as_of_date):
    rows = []
    for row in report['asset_rows']:
        rows.append([as_of_date, 'Aset', row['account'].code, row['account'].name, row['amount']])
    rows.append([as_of_date, 'Ringkasan', '', 'Total Aset', report['total_assets']])
    for row in report['liability_rows']:
        rows.append([as_of_date, 'Kewajiban', row['account'].code, row['account'].name, row['amount']])
    rows.append([as_of_date, 'Ringkasan', '', 'Total Kewajiban', report['total_liabilities']])
    for row in report['equity_rows']:
        account = row.get('account')
        rows.append([
            as_of_date,
            'Ekuitas',
            account.code if account else row.get('code', ''),
            account.name if account else row.get('name', ''),
            row['amount'],
        ])
    rows.append([as_of_date, 'Ringkasan', '', 'Total Ekuitas', report['total_equity']])
    rows.append([as_of_date, 'Ringkasan', '', 'Total Kewajiban + Ekuitas', report['total_liabilities_equity']])
    return rows


@admin_bp.route('/keuangan/laporan/posisi-keuangan/export')
@login_required
@role_required(UserRole.TU)
def finance_financial_position_export():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))
    as_of_date = _parse_iso_date(request.args.get('as_of_date')) or local_today()
    report = _financial_position_data(tenant_id, as_of_date)
    return _csv_response(
        'finance_posisi_keuangan.csv',
        ['Tanggal Posisi', 'Bagian', 'Kode Akun', 'Nama Akun', 'Nominal'],
        _financial_position_export_rows(report, as_of_date),
    )


@admin_bp.route('/keuangan/laporan/posisi-keuangan/export-xlsx')
@login_required
@role_required(UserRole.TU)
def finance_financial_position_export_xlsx():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))
    as_of_date = _parse_iso_date(request.args.get('as_of_date')) or local_today()
    report = _financial_position_data(tenant_id, as_of_date)
    return _xlsx_response(
        'finance_posisi_keuangan.xlsx',
        'Posisi Keuangan',
        ['Tanggal Posisi', 'Bagian', 'Kode Akun', 'Nama Akun', 'Nominal'],
        _financial_position_export_rows(report, as_of_date),
    )


@admin_bp.route('/keuangan/laporan/posisi-keuangan/print')
@login_required
@role_required(UserRole.TU)
def finance_financial_position_print():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))
    as_of_date = _parse_iso_date(request.args.get('as_of_date')) or local_today()
    report = _financial_position_data(tenant_id, as_of_date)
    return render_template(
        'admin/finance/print_financial_position.html',
        as_of_date=as_of_date,
        signers=_report_signers_from_request(),
        **report,
    )


def _unposted_invoice_payment_transactions(tenant_id):
    posted_payment_ids = db.session.query(FinanceJournal.source_id).filter(
        FinanceJournal.tenant_id == tenant_id,
        FinanceJournal.source_type == FinanceJournalSourceType.INVOICE_PAYMENT,
        FinanceJournal.status == FinanceJournalStatus.POSTED,
    )
    return (
        Transaction.query
        .join(Invoice, Invoice.id == Transaction.invoice_id)
        .join(Student, Student.id == Invoice.student_id)
        .join(User, User.id == Student.user_id)
        .filter(
            User.tenant_id == tenant_id,
            Invoice.is_deleted.is_(False),
            Student.is_deleted.is_(False),
            ~Transaction.id.in_(posted_payment_ids),
        )
        .order_by(Transaction.date.desc(), Transaction.id.desc())
    )


def _unposted_savings_transactions(tenant_id):
    posted_deposit_ids = db.session.query(FinanceJournal.source_id).filter(
        FinanceJournal.tenant_id == tenant_id,
        FinanceJournal.source_type == FinanceJournalSourceType.SAVINGS_DEPOSIT,
        FinanceJournal.status == FinanceJournalStatus.POSTED,
    )
    posted_withdrawal_ids = db.session.query(FinanceJournal.source_id).filter(
        FinanceJournal.tenant_id == tenant_id,
        FinanceJournal.source_type == FinanceJournalSourceType.SAVINGS_WITHDRAWAL,
        FinanceJournal.status == FinanceJournalStatus.POSTED,
    )
    return (
        StudentSavingsTransaction.query
        .filter(
            StudentSavingsTransaction.tenant_id == tenant_id,
            StudentSavingsTransaction.status == SavingsTransactionStatus.APPROVED,
            or_(
                and_(
                    StudentSavingsTransaction.transaction_type == SavingsTransactionType.DEPOSIT,
                    ~StudentSavingsTransaction.id.in_(posted_deposit_ids),
                ),
                and_(
                    StudentSavingsTransaction.transaction_type == SavingsTransactionType.WITHDRAWAL,
                    ~StudentSavingsTransaction.id.in_(posted_withdrawal_ids),
                ),
            ),
        )
        .order_by(StudentSavingsTransaction.approved_at.desc(), StudentSavingsTransaction.id.desc())
    )


@admin_bp.route('/keuangan/rekonsiliasi')
@login_required
@role_required(UserRole.TU)
def finance_reconciliation():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    current_day = local_today()
    period_today = FinancePeriod.query.filter(
        FinancePeriod.tenant_id == tenant_id,
        FinancePeriod.start_date <= current_day,
        FinancePeriod.end_date >= current_day,
    ).first()

    invoice_unposted = _unposted_invoice_payment_transactions(tenant_id).limit(50).all()
    savings_unposted = _unposted_savings_transactions(tenant_id).limit(50).all()
    draft_journals = FinanceJournal.query.filter_by(
        tenant_id=tenant_id,
        status=FinanceJournalStatus.DRAFT,
    ).order_by(FinanceJournal.id.desc()).limit(100).all()

    return render_template(
        'admin/finance/reconciliation.html',
        invoice_unposted=invoice_unposted,
        savings_unposted=savings_unposted,
        draft_journals=draft_journals,
        period_today=period_today,
    )


@admin_bp.route('/keuangan/rekonsiliasi/retry-journal/<int:journal_id>', methods=['POST'])
@login_required
@role_required(UserRole.TU)
def finance_reconciliation_retry_journal(journal_id):
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.finance_reconciliation'))
    try:
        post_journal(tenant_id=tenant_id, journal_id=journal_id, actor_user_id=current_user.id)
        db.session.commit()
        flash(f'Jurnal #{journal_id} berhasil diposting.', 'success')
    except Exception as exc:
        db.session.rollback()
        flash(f'Gagal posting jurnal #{journal_id}: {exc}', 'warning')
    return redirect(url_for('admin.finance_reconciliation'))


@admin_bp.route('/keuangan/rekonsiliasi/retry-draft-all', methods=['POST'])
@login_required
@role_required(UserRole.TU)
def finance_reconciliation_retry_draft_all():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.finance_reconciliation'))

    drafts = FinanceJournal.query.filter_by(
        tenant_id=tenant_id,
        status=FinanceJournalStatus.DRAFT,
    ).order_by(FinanceJournal.id.asc()).all()
    success = 0
    failed = 0
    for journal in drafts:
        try:
            post_journal(tenant_id=tenant_id, journal_id=journal.id, actor_user_id=current_user.id)
            db.session.commit()
            success += 1
        except Exception:
            db.session.rollback()
            failed += 1
    flash(f'Retry posting draft selesai. Berhasil: {success}, gagal: {failed}.', 'info')
    return redirect(url_for('admin.finance_reconciliation'))


@admin_bp.route('/keuangan/rekonsiliasi/retry-sources', methods=['POST'])
@login_required
@role_required(UserRole.TU)
def finance_reconciliation_retry_sources():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.finance_reconciliation'))

    source = (request.form.get('source') or '').strip().lower()
    success = 0
    failed = 0

    if source == 'invoice':
        for trx in _unposted_invoice_payment_transactions(tenant_id).limit(200).all():
            try:
                post_invoice_payment(tenant_id=tenant_id, transaction_id=trx.id, actor_user_id=current_user.id)
                success += 1
            except Exception:
                db.session.rollback()
                failed += 1
        flash(f'Retry sumber pembayaran selesai. Berhasil: {success}, gagal: {failed}.', 'info')
        return redirect(url_for('admin.finance_reconciliation'))

    if source == 'savings':
        for savings_trx in _unposted_savings_transactions(tenant_id).limit(200).all():
            try:
                post_savings_transaction(
                    tenant_id=tenant_id,
                    savings_transaction_id=savings_trx.id,
                    actor_user_id=current_user.id
                )
                success += 1
            except Exception:
                db.session.rollback()
                failed += 1
        flash(f'Retry sumber tabungan selesai. Berhasil: {success}, gagal: {failed}.', 'info')
        return redirect(url_for('admin.finance_reconciliation'))

    flash('Sumber retry tidak valid.', 'warning')
    return redirect(url_for('admin.finance_reconciliation'))


@admin_bp.route('/keuangan/master-biaya', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.TU)
def manage_fee_types():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        form_type = (request.form.get('form_type') or 'master_fee').strip()

        name = request.form.get('name')
        amount = request.form.get('amount')
        academic_year_id = request.form.get('academic_year_id', type=int)
        amount_rupiah = to_rupiah_int(amount, default=-1)

        if amount_rupiah <= 0:
            flash('Nominal biaya harus lebih dari 0.', 'warning')
            return redirect(url_for('admin.manage_fee_types'))

        try:
            new_fee = FeeType(
                tenant_id=tenant_id,
                name=name,
                amount=amount_rupiah,
                academic_year_id=academic_year_id
            )
            db.session.add(new_fee)
            db.session.commit()
            flash('Jenis Biaya Master berhasil dibuat.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {e}', 'danger')

        return redirect(url_for('admin.manage_fee_types'))

    query = (request.args.get('q') or '').strip()
    fees_query = FeeType.query.filter(FeeType.tenant_id == tenant_id).outerjoin(
        AcademicYear,
        FeeType.academic_year_id == AcademicYear.id,
    )
    if query:
        fees_query = fees_query.filter(
            or_(
                FeeType.name.ilike(f'%{query}%'),
                AcademicYear.name.ilike(f'%{query}%'),
                AcademicYear.semester.ilike(f'%{query}%')
            )
        )

    fees = fees_query.order_by(FeeType.id.desc()).all()
    years = AcademicYear.query.filter_by(is_active=True).all()
    return render_template(
        'admin/finance/fee_types.html',
        fees=fees,
        years=years,
        query=query,
    )


@admin_bp.route('/keuangan/biaya/edit/<int:fee_id>', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.TU)
def edit_fee_type(fee_id):
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.manage_fee_types'))

    fee = FeeType.query.filter_by(id=fee_id, tenant_id=tenant_id).first_or_404()
    years = AcademicYear.query.filter_by(is_active=True).all()

    if request.method == 'POST':
        fee.name = request.form.get('name')
        amount_rupiah = to_rupiah_int(request.form.get('amount'), default=-1)
        if amount_rupiah <= 0:
            flash('Nominal biaya harus lebih dari 0.', 'warning')
            return redirect(url_for('admin.edit_fee_type', fee_id=fee_id))
        fee.amount = amount_rupiah

        # Handle Tahun Ajaran (Bisa None/Null jika berlaku umum)
        year_id = request.form.get('academic_year_id')
        fee.academic_year_id = year_id if year_id else None

        try:
            db.session.commit()
            flash(f'Master Biaya "{fee.name}" berhasil diperbarui.', 'success')
            return redirect(url_for('admin.manage_fee_types'))
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal update: {e}', 'danger')

    return render_template('admin/finance/edit_fee_type.html', fee=fee, years=years)


@admin_bp.route('/keuangan/generate/<int:fee_id>', methods=['POST'])
@login_required
@role_required(UserRole.TU)
def generate_invoices(fee_id):
    """
    Admin berhak menerbitkan tagihan untuk seluruh siswa berdasarkan FeeType.
    Menggunakan logika yang sama seperti modul TU dengan guard agar tidak error
    jika relasi student_candidate belum tersedia.
    """
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.manage_fee_types'))

    fee = FeeType.query.filter_by(id=fee_id, tenant_id=tenant_id).first_or_404()
    students = (
        Student.query.join(User, Student.user_id == User.id)
        .filter(
            Student.is_deleted.is_(False),
            User.tenant_id == tenant_id,
        )
        .all()
    )

    count_success = 0
    due_date_default = local_now() + timedelta(days=10)
    is_monthly_fee = "SPP" in fee.name.upper() or "BULAN" in fee.name.upper()

    try:
        for student in students:
            candidate = getattr(student, "student_candidate", None)

            if candidate:
                if "RQDF" in fee.name.upper() and candidate.program_type.name != 'RQDF_SORE':
                    continue
                if "RQDF" not in fee.name.upper() and candidate.program_type.name == 'RQDF_SORE':
                    continue

            if Invoice.query.filter_by(student_id=student.id, fee_type_id=fee.id, is_deleted=False).first():
                continue

            nominal_final = fee.amount
            if is_monthly_fee and student.custom_spp_fee is not None:
                nominal_final = student.custom_spp_fee
            elif candidate and candidate.scholarship_category.name != 'NON_BEASISWA':
                nominal_final = fee.amount * 0.5

            new_inv = Invoice(
                invoice_number=generate_invoice_number(fee.id, student.id, tenant_id=tenant_id),
                student_id=student.id,
                fee_type_id=fee.id,
                total_amount=to_rupiah_int(nominal_final),
                status=PaymentStatus.UNPAID,
                due_date=due_date_default
            )
            db.session.add(new_inv)
            count_success += 1

        db.session.commit()
        flash(f'Berhasil menerbitkan {count_success} tagihan baru.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')

    return redirect(url_for('admin.manage_fee_types'))




# =========================================================
# 8. MANAJEMEN PPDB
# =========================================================

@admin_bp.route('/ppdb/settings', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def ppdb_settings():
    return ppdb_settings_view(
        settings_endpoint='admin.ppdb_settings',
        form_builder_endpoint='admin.ppdb_form_builder',
    )


@admin_bp.route('/ppdb/form-builder/<int:path_id>', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def ppdb_form_builder(path_id):
    return ppdb_form_builder_view(
        path_id,
        settings_endpoint='admin.ppdb_settings',
        form_builder_endpoint='admin.ppdb_form_builder',
    )


@admin_bp.route('/ppdb/pendaftar')
@login_required
@role_required(UserRole.ADMIN)
def ppdb_list():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    query = (request.args.get('q') or '').strip()
    candidates_query = StudentCandidate.query.filter_by(tenant_id=tenant_id, is_deleted=False)
    if query:
        candidates_query = candidates_query.filter(
            or_(
                StudentCandidate.registration_no.ilike(f'%{query}%'),
                StudentCandidate.full_name.ilike(f'%{query}%'),
                StudentCandidate.parent_phone.ilike(f'%{query}%'),
                StudentCandidate.personal_phone.ilike(f'%{query}%')
            )
        )

    candidates = candidates_query.order_by(StudentCandidate.created_at.desc()).all()
    tenant = Tenant.query.filter_by(id=tenant_id, is_deleted=False).first()
    return render_template(
        'staff/ppdb/list.html',
        candidates=candidates,
        query=query,
        list_endpoint='admin.ppdb_list',
        detail_endpoint='admin.ppdb_detail',
        public_ppdb_tenant_slug=tenant.slug if tenant else None,
    )


@admin_bp.route('/ppdb/detail/<int:candidate_id>')
@login_required
@role_required(UserRole.ADMIN)
def ppdb_detail(candidate_id):
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    candidate = StudentCandidate.query.filter_by(id=candidate_id, tenant_id=tenant_id, is_deleted=False).first_or_404()
    fee_drafts = []
    if candidate.status == RegistrationStatus.PENDING and candidate.program_type != ProgramType.MAJLIS_TALIM:
        fee_drafts = build_candidate_fee_drafts(candidate, tenant_id=tenant_id)
    custom_fields = list_active_ppdb_form_fields(tenant_id, candidate.ppdb_period, candidate.ppdb_path)
    document_requirements = list_active_ppdb_document_requirements(tenant_id, candidate.ppdb_period, candidate.ppdb_path)

    return render_template(
        'staff/ppdb/detail.html',
        candidate=candidate,
        fee_drafts=fee_drafts,
        fee_drafts_total=sum(item.get('nominal', 0) for item in fee_drafts),
        custom_fields=custom_fields,
        extra_answers=_loads_object(candidate.extra_answers_json),
        document_requirements=document_requirements,
        document_status=_loads_object(candidate.document_status_json),
        list_endpoint='admin.ppdb_list',
        accept_endpoint='admin.accept_candidate',
    )


@admin_bp.route('/ppdb/terima/<int:candidate_id>', methods=['POST'])
@login_required
@role_required(UserRole.ADMIN)
def accept_candidate(candidate_id):
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    calon = StudentCandidate.query.filter_by(id=candidate_id, tenant_id=tenant_id, is_deleted=False).first_or_404()

    if calon.status == RegistrationStatus.ACCEPTED:
        flash('Siswa ini sudah diproses sebelumnya.', 'warning')
        return redirect(url_for('admin.ppdb_list'))

    try:
        # Jalur khusus peserta Majelis Ta'lim (tidak membuat akun siswa & tagihan)
        if calon.program_type == ProgramType.MAJLIS_TALIM:
            nomor_majelis = calon.personal_phone or calon.parent_phone
            if not nomor_majelis:
                raise ValueError('Nomor WhatsApp peserta Majelis tidak ditemukan.')

            majlis_user = User.query.filter_by(
                username=nomor_majelis,
                tenant_id=tenant_id,
                is_deleted=False,
            ).first()
            if not majlis_user:
                majlis_user = User(
                    tenant_id=tenant_id,
                    username=nomor_majelis,
                    email=f"majlis.{calon.id}@sekolah.id",
                    password_hash=generate_password_hash(nomor_majelis or "123456"),
                    role=UserRole.MAJLIS_PARTICIPANT,
                    must_change_password=True,
                )
                db.session.add(majlis_user)
                db.session.flush()
            ensure_majlis_participant_acceptance(
                user=majlis_user,
                full_name=calon.full_name,
                phone=nomor_majelis,
                address=calon.address,
                job=calon.personal_job,
            )

            calon.status = RegistrationStatus.ACCEPTED
            db.session.commit()
            flash(f"Peserta Majelis {calon.full_name} berhasil diterima.", 'success')
            return redirect(url_for('admin.ppdb_list'))

        # --- 1. PROSES AKUN ---
        nis_baru = generate_nis()

        # User Wali
        parent_phone = (calon.parent_phone or '').strip()
        if not parent_phone:
            raise ValueError('Nomor Telepon Orang Tua wajib diisi.')

        user_wali = User.query.filter_by(
            username=parent_phone,
            tenant_id=tenant_id,
            is_deleted=False,
        ).first()
        if not user_wali:
            user_wali = User(tenant_id=tenant_id, username=parent_phone, email=f"wali.{nis_baru}@sekolah.id",
                             password_hash=generate_password_hash(parent_phone or "123456"),
                             role=UserRole.WALI_MURID,
                             must_change_password=True)
            db.session.add(user_wali)
            db.session.flush()
        parent_profile = user_wali.parent_profile
        if not parent_profile:
            parent_profile = Parent(
                user_id=user_wali.id,
                full_name=calon.father_name or calon.mother_name or "Wali Murid",
                phone=parent_phone,
                job=calon.father_job,
                address=calon.address
            )
            db.session.add(parent_profile)
            db.session.flush()
        else:
            if not parent_profile.full_name:
                parent_profile.full_name = calon.father_name or calon.mother_name or "Wali Murid"
            if not parent_profile.phone:
                parent_profile.phone = parent_phone
            if not parent_profile.job and calon.father_job:
                parent_profile.job = calon.father_job
            if not parent_profile.address and calon.address:
                parent_profile.address = calon.address

        # User Siswa
        user_siswa = User(tenant_id=tenant_id, username=nis_baru, email=f"{nis_baru}@sekolah.id",
                          password_hash=generate_password_hash(nis_baru), role=UserRole.SISWA,
                          must_change_password=True)
        db.session.add(user_siswa)
        db.session.flush()
        siswa_baru = Student(user_id=user_siswa.id, parent_id=parent_profile.id, nis=nis_baru,
                             full_name=calon.full_name, gender=calon.gender, place_of_birth=calon.place_of_birth,
                             date_of_birth=calon.date_of_birth, address=calon.address)
        db.session.add(siswa_baru)
        db.session.flush()

        # --- 2. SMART INVOICING (VERSI DINAMIS) ---
        tagihan_list = build_candidate_fee_drafts(calon, tenant_id=tenant_id)

        due_date = local_now() + timedelta(days=14)

        ctr = 1
        for item in tagihan_list:
            fee_type = FeeType.query.filter_by(tenant_id=tenant_id, name=item['nama']).first()
            if not fee_type:
                fee_type = FeeType(
                    tenant_id=tenant_id,
                    name=item['nama'],
                    amount=to_rupiah_int(item['nominal']),
                )
                db.session.add(fee_type)
                db.session.flush()

            new_inv = Invoice(
                invoice_number=generate_invoice_number(
                    fee_type.id,
                    siswa_baru.id,
                    sequence=ctr,
                    tenant_id=tenant_id,
                ),
                student_id=siswa_baru.id,
                fee_type_id=fee_type.id,
                total_amount=to_rupiah_int(item['nominal']),
                paid_amount=0,
                status=PaymentStatus.UNPAID,
                due_date=due_date
            )
            db.session.add(new_inv)
            ctr += 1

        calon.status = RegistrationStatus.ACCEPTED
        db.session.commit()
        flash(f'Sukses! Siswa {siswa_baru.full_name} diterima. {len(tagihan_list)} rincian tagihan diterbitkan.',
              'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error: {e}', 'danger')
        print(e)

    return redirect(url_for('admin.ppdb_list'))

# =========================================================
# 8. MANAJEMEN USER
# =========================================================

@admin_bp.route('/student/reset-password/<int:user_id>', methods=['POST'])
@login_required
@role_required(UserRole.ADMIN)
def reset_password(user_id):
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('main.dashboard'))

    user = User.query.filter_by(id=user_id, tenant_id=tenant_id).first_or_404()

    if user.has_role(UserRole.ADMIN):
        flash('Tidak bisa mereset akun Admin lain dari sini.', 'danger')
        return redirect(url_for('main.dashboard'))

    try:
        new_password = "123456"  # Default fallback

        if user.has_role(UserRole.SISWA) and user.student_profile:
            new_password = user.student_profile.nis
        elif user.has_role(UserRole.WALI_MURID) and user.parent_profile:
            new_password = user.parent_profile.phone
        elif user.has_role(UserRole.WALI_ASRAMA) and user.boarding_guardian_profile:
            new_password = user.boarding_guardian_profile.phone or "123456"

        user.password_hash = generate_password_hash(new_password)
        user.must_change_password = True
        db.session.commit()

        flash(f'Password user {user.username} berhasil direset menjadi: {new_password}', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Gagal mereset password: {str(e)}', 'danger')

    return redirect(request.referrer or url_for('main.dashboard'))


# =========================================================
# 9. MANAJEMEN JADWAL PELAJARAN
# =========================================================

@admin_bp.route('/akademik/jadwal', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def manage_schedules():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return render_template(
            'admin/academic/schedules.html',
            classes=[],
            subjects=[],
            teachers=[],
            schedules=[],
            selected_class=None,
        )

    # Ambil parameter filter kelas dari URL (misal: ?class_id=1)
    selected_class_id = request.args.get('class_id', type=int)

    # Pastikan data lama "guru mapel Rumah Qur'an" ditutup.
    cleanup_stats = cleanup_rumah_quran_subject_data(tenant_id=tenant_id)
    if cleanup_stats["closed_assignments"] or cleanup_stats["deleted_schedules"]:
        db.session.commit()

    # Dropdown Data
    classes = (
        scoped_classrooms_query(tenant_id)
        .filter(
            or_(
                ClassRoom.program_type.is_(None),
                ~ClassRoom.program_type.in_([ProgramType.RQDF_SORE, ProgramType.TAKHOSUS_TAHFIDZ]),
            )
        )
        .all()
    )
    subjects = Subject.query.filter_by(is_deleted=False).all()
    teachers = _tenant_teachers_query(tenant_id).all()

    # Jika user mengirim Form Tambah Jadwal
    if request.method == 'POST':
        class_id = request.form.get('class_id', type=int)
        subject_id = request.form.get('subject_id', type=int)
        teacher_id = request.form.get('teacher_id', type=int)
        day = request.form.get('day')
        start_time_str = request.form.get('start_time')
        end_time_str = request.form.get('end_time')

        try:
            target_class = scoped_classrooms_query(tenant_id).filter(ClassRoom.id == class_id).first()
            if target_class is None:
                flash('Kelas tidak valid.', 'warning')
                return redirect(url_for('admin.manage_schedules'))
            if is_rumah_quran_classroom(target_class):
                flash("Kelas Rumah Qur'an tidak menggunakan jadwal mapel.", 'warning')
                return redirect(url_for('admin.manage_schedules', class_id=class_id))
            if _tenant_teachers_query(tenant_id).filter(Teacher.id == teacher_id).first() is None:
                flash('Guru tidak valid untuk tenant aktif.', 'warning')
                return redirect(url_for('admin.manage_schedules', class_id=class_id))

            # 1. Konversi String jam "07:00" menjadi object Time python
            start_time = datetime.strptime(start_time_str, '%H:%M').time()
            end_time = datetime.strptime(end_time_str, '%H:%M').time()

            # 2. Validasi Logika Waktu (Mulai harus sebelum Selesai)
            if start_time >= end_time:
                flash('Jam mulai harus lebih awal dari jam selesai!', 'warning')
                return redirect(url_for('admin.manage_schedules', class_id=class_id))

            # ==========================================
            # 3. CEK BENTROK JADWAL (CLASH DETECTION)
            # ==========================================

            # A. Cek Bentrok KELAS (Kelas ini sudah dipakai belum di jam segitu?)
            clash_class = Schedule.query.filter(
                Schedule.class_id == class_id,
                Schedule.is_deleted.is_(False),
                Schedule.day == day,
                Schedule.start_time < end_time,  # Logic overlap: StartA < EndB
                Schedule.end_time > start_time  # Logic overlap: EndA > StartB
            ).first()

            if clash_class:
                # Ambil nama mapel agar pesan error jelas (menggunakan relationship 'subject')
                mapel_name = clash_class.subject.name if clash_class.subject else "Mapel Lain"
                flash(
                    f'Gagal! Bentrok dengan mapel "{mapel_name}" di kelas ini ({clash_class.start_time.strftime("%H:%M")} - {clash_class.end_time.strftime("%H:%M")}).',
                    'danger')
                return redirect(url_for('admin.manage_schedules', class_id=class_id))

            # B. Cek Bentrok GURU (Guru ini sedang mengajar di kelas lain tidak?)
            clash_teacher = (
                Schedule.query.join(ClassRoom, Schedule.class_id == ClassRoom.id)
                .join(ProgramGroup, ClassRoom.program_group_id == ProgramGroup.id)
                .filter(
                    ProgramGroup.tenant_id == tenant_id,
                    Schedule.teacher_id == teacher_id,
                    Schedule.is_deleted.is_(False),
                    Schedule.day == day,
                    Schedule.start_time < end_time,
                    Schedule.end_time > start_time,
                )
                .first()
            )

            if clash_teacher:
                # Ambil nama kelas tempat guru tsb sedang mengajar
                other_class = clash_teacher.class_room.name if clash_teacher.class_room else "Kelas Lain"
                flash(f'Gagal! Guru tersebut sedang mengajar di "{other_class}" pada jam yang sama.', 'danger')
                return redirect(url_for('admin.manage_schedules', class_id=class_id))

            # ==========================================
            # 4. SIMPAN JIKA LOLOS VALIDASI
            # ==========================================
            new_schedule = Schedule(
                class_id=class_id,
                subject_id=subject_id,
                teacher_id=teacher_id,
                day=day,
                start_time=start_time,
                end_time=end_time
            )
            db.session.add(new_schedule)
            db.session.commit()
            flash('Jadwal berhasil ditambahkan.', 'success')

        except Exception as e:
            db.session.rollback()
            flash(f'Gagal menambah jadwal (System Error): {e}', 'danger')

        # Redirect kembali ke kelas yang sedang dipilih
        return redirect(url_for('admin.manage_schedules', class_id=class_id))

    # Query Jadwal untuk ditampilkan di tabel
    schedules = []
    selected_class = None

    if selected_class_id:
        selected_class = scoped_classrooms_query(tenant_id).filter(ClassRoom.id == selected_class_id).first()
        if selected_class is None:
            flash('Kelas tidak valid.', 'warning')
            return redirect(url_for('admin.manage_schedules'))
        if selected_class and is_rumah_quran_classroom(selected_class):
            flash("Kelas Rumah Qur'an tidak menggunakan jadwal mapel.", 'warning')
            return redirect(url_for('admin.manage_schedules'))
        # Urutkan berdasarkan Hari (Senin-Jumat) dan Jam Mulai
        schedules = Schedule.query.filter_by(class_id=selected_class_id, is_deleted=False) \
            .order_by(Schedule.day, Schedule.start_time).all()

        # Custom sort di python agar harinya urut Senin->Minggu
        days_order = {'Senin': 1, 'Selasa': 2, 'Rabu': 3, 'Kamis': 4, 'Jumat': 5, 'Sabtu': 6, 'Minggu': 7}
        schedules.sort(key=lambda x: (days_order.get(x.day, 8), x.start_time))

    return render_template('admin/academic/schedules.html',
                           classes=classes,
                           subjects=subjects,
                           teachers=teachers,
                           schedules=schedules,
                           selected_class=selected_class)


@admin_bp.route('/akademik/jadwal/edit/<int:id>', methods=['POST'])
@login_required
@role_required(UserRole.ADMIN)
def edit_schedule(id):
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.manage_schedules'))

    schedule = (
        Schedule.query.join(ClassRoom, Schedule.class_id == ClassRoom.id)
        .join(ProgramGroup, ClassRoom.program_group_id == ProgramGroup.id)
        .filter(
            Schedule.id == id,
            Schedule.is_deleted.is_(False),
            ProgramGroup.tenant_id == tenant_id,
        )
        .first_or_404()
    )
    class_id = request.form.get('class_id', type=int) or schedule.class_id  # Fallback

    if schedule.class_room and is_rumah_quran_classroom(schedule.class_room):
        schedule.is_deleted = True
        db.session.commit()
        flash("Jadwal mapel kelas Rumah Qur'an telah dinonaktifkan.", 'warning')
        return redirect(url_for('admin.manage_schedules', class_id=schedule.class_id))

    # Ambil data dari form
    subject_id = request.form.get('subject_id', type=int)
    teacher_id = request.form.get('teacher_id', type=int)
    day = request.form.get('day')
    start_time_str = request.form.get('start_time')
    end_time_str = request.form.get('end_time')

    try:
        target_class = scoped_classrooms_query(tenant_id).filter(ClassRoom.id == class_id).first()
        if target_class is None:
            flash('Kelas tidak valid.', 'warning')
            return redirect(url_for('admin.manage_schedules', class_id=class_id))
        if is_rumah_quran_classroom(target_class):
            flash("Kelas Rumah Qur'an tidak menggunakan jadwal mapel.", 'warning')
            return redirect(url_for('admin.manage_schedules', class_id=class_id))
        if _tenant_teachers_query(tenant_id).filter(Teacher.id == teacher_id).first() is None:
            flash('Guru tidak valid untuk tenant aktif.', 'warning')
            return redirect(url_for('admin.manage_schedules', class_id=class_id))

        start_time = datetime.strptime(start_time_str, '%H:%M').time()
        end_time = datetime.strptime(end_time_str, '%H:%M').time()

        if start_time >= end_time:
            flash('Jam mulai harus lebih awal dari jam selesai!', 'warning')
            return redirect(url_for('admin.manage_schedules', class_id=class_id))

        # Cek Bentrok KELAS
        clash_class = Schedule.query.filter(
            Schedule.id != id,  # PENTING: Jangan cek jadwal diri sendiri
            Schedule.class_id == class_id,
            Schedule.is_deleted.is_(False),
            Schedule.day == day,
            Schedule.start_time < end_time,
            Schedule.end_time > start_time
        ).first()

        if clash_class:
            flash(f'Gagal Update! Bentrok dengan mapel lain di kelas ini.', 'danger')
            return redirect(url_for('admin.manage_schedules', class_id=class_id))

        # Cek Bentrok GURU
        clash_teacher = (
            Schedule.query.join(ClassRoom, Schedule.class_id == ClassRoom.id)
            .join(ProgramGroup, ClassRoom.program_group_id == ProgramGroup.id)
            .filter(
                Schedule.id != id,  # PENTING: Jangan cek jadwal diri sendiri
                ProgramGroup.tenant_id == tenant_id,
                Schedule.teacher_id == teacher_id,
                Schedule.is_deleted.is_(False),
                Schedule.day == day,
                Schedule.start_time < end_time,
                Schedule.end_time > start_time,
            )
            .first()
        )

        if clash_teacher:
            flash(f'Gagal Update! Guru sedang mengajar di kelas lain.', 'danger')
            return redirect(url_for('admin.manage_schedules', class_id=class_id))

        # UPDATE DATA
        schedule.subject_id = subject_id
        schedule.teacher_id = teacher_id
        schedule.day = day
        schedule.start_time = start_time
        schedule.end_time = end_time

        db.session.commit()
        flash('Jadwal berhasil diperbarui.', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Gagal update: {e}', 'danger')

    return redirect(url_for('admin.manage_schedules', class_id=class_id))


@admin_bp.route('/akademik/jadwal/hapus/<int:id>', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def delete_schedule(id):
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.manage_schedules'))

    schedule = (
        Schedule.query.join(ClassRoom, Schedule.class_id == ClassRoom.id)
        .join(ProgramGroup, ClassRoom.program_group_id == ProgramGroup.id)
        .filter(
            Schedule.id == id,
            Schedule.is_deleted.is_(False),
            ProgramGroup.tenant_id == tenant_id,
        )
        .first_or_404()
    )
    class_id = schedule.class_id

    # Optional: Jika ingin strict hanya boleh POST
    if request.method == 'GET':
        # Bisa redirect balik atau tampilkan error
        pass

    db.session.delete(schedule)
    db.session.commit()

    flash('Jadwal dihapus.', 'success')  # Ubah jadi success warna hijau
    return redirect(url_for('admin.manage_schedules', class_id=class_id))


# =========================================================
# 10. MANAJEMEN USER PUSAT (RESET PASSWORD ALL ROLES)
# =========================================================

@admin_bp.route('/users/manage', methods=['GET'])
@login_required
@role_required(UserRole.ADMIN)
def manage_users():
    """Halaman untuk melihat semua user dan reset password"""
    query = (request.args.get('q') or '').strip()
    role_filter = (request.args.get('role') or 'all').strip().lower()

    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return render_template(
            'admin/users/manage.html',
            users=[],
            query=query,
            role_filter=role_filter,
            UserRole=UserRole,
        )

    # Ambil semua user tenant aktif KECUALI Admin (untuk keamanan)
    users_query = User.query.filter(
        User.tenant_id == tenant_id,
        User.role != UserRole.ADMIN,
        User.role != UserRole.SUPER_ADMIN,
        ~User.role_assignments.any(role=UserRole.ADMIN),
        ~User.role_assignments.any(role=UserRole.SUPER_ADMIN),
    )

    role_mapping = {
        'santri': UserRole.SISWA,
        'wali': UserRole.WALI_MURID,
        'wali_asrama': UserRole.WALI_ASRAMA,
        'guru': UserRole.GURU,
        'peserta_majlis': UserRole.MAJLIS_PARTICIPANT,
        'staff': UserRole.TU,
    }
    selected_role = role_mapping.get(role_filter)
    if selected_role:
        users_query = users_query.filter(
            or_(
                User.role == selected_role,
                User.role_assignments.any(role=selected_role)
            )
        )

    users = users_query.order_by(User.role, User.username).all()

    if query:
        keyword = query.lower()
        filtered_users = []
        for u in users:
            owner_name = ''
            if u.student_profile:
                owner_name = u.student_profile.full_name or ''
            elif u.parent_profile:
                owner_name = u.parent_profile.full_name or ''
            elif u.teacher_profile:
                owner_name = u.teacher_profile.full_name or ''
            elif u.majlis_profile:
                owner_name = u.majlis_profile.full_name or ''
            elif u.staff_profile:
                owner_name = u.staff_profile.full_name or ''
            elif u.boarding_guardian_profile:
                owner_name = u.boarding_guardian_profile.full_name or ''

            if (
                keyword in (u.username or '').lower() or
                keyword in (u.role.value or '').lower() or
                any(keyword in rv.lower() for rv in u.all_role_values()) or
                keyword in owner_name.lower()
            ):
                filtered_users.append(u)
        users = filtered_users

    return render_template(
        'admin/users/manage.html',
        users=users,
        query=query,
        role_filter=role_filter,
        UserRole=UserRole,
    )


@admin_bp.route('/users/roles', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def manage_user_roles():
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.manage_users'))

    if request.method == 'POST':
        user_id = request.form.get('user_id', type=int)
        selected_roles_raw = request.form.getlist('roles')
        query = (request.form.get('q') or '').strip()

        user = User.query.filter_by(id=user_id, tenant_id=tenant_id).first_or_404()
        selected_roles = set()
        for item in selected_roles_raw:
            try:
                selected_roles.add(UserRole[item])
            except KeyError:
                pass

        is_valid, message = validate_role_combination(selected_roles)
        if not is_valid:
            flash(message, 'danger')
            return redirect(url_for('admin.manage_user_roles', q=query))
        if UserRole.SUPER_ADMIN in selected_roles:
            flash('Role Super Admin hanya bisa dikelola melalui level platform.', 'danger')
            return redirect(url_for('admin.manage_user_roles', q=query))

        # Pilih role utama berdasarkan prioritas global agar deterministik
        new_primary = None
        for role in ROLE_PRIORITY:
            if role in selected_roles:
                new_primary = role
                break

        if not new_primary:
            flash('Role utama tidak valid.', 'danger')
            return redirect(url_for('admin.manage_user_roles', q=query))

        # Validasi role yang wajib punya profil spesifik
        if UserRole.SISWA in selected_roles and not user.student_profile:
            flash('Role Santri hanya bisa diberikan ke user yang sudah memiliki profil siswa.', 'danger')
            return redirect(url_for('admin.manage_user_roles', q=query))

        if UserRole.WALI_MURID in selected_roles and not user.parent_profile:
            flash('Role Wali Murid hanya bisa diberikan ke user yang sudah memiliki profil wali murid.', 'danger')
            return redirect(url_for('admin.manage_user_roles', q=query))

        if UserRole.MAJLIS_PARTICIPANT in selected_roles and not user.majlis_profile:
            flash("Role Peserta Majlis hanya bisa diberikan ke user yang sudah memiliki profil peserta majlis.", 'danger')
            return redirect(url_for('admin.manage_user_roles', q=query))

        # Auto-provision profil untuk role operasional agar langsung muncul di modul terkait
        display_name = _infer_user_display_name(user)
        phone = _infer_user_phone(user)
        if UserRole.GURU in selected_roles and not user.teacher_profile:
            db.session.add(Teacher(
                user_id=user.id,
                full_name=display_name,
                phone=phone
            ))

        if UserRole.TU in selected_roles and not user.staff_profile:
            db.session.add(Staff(
                user_id=user.id,
                full_name=display_name,
                position='Staff'
            ))

        if UserRole.WALI_ASRAMA in selected_roles and not user.boarding_guardian_profile:
            db.session.add(BoardingGuardian(
                user_id=user.id,
                full_name=display_name,
                phone=phone
            ))

        user.role = new_primary

        # Sinkronkan role assignment tambahan (di luar role utama)
        target_extra_roles = selected_roles - {new_primary}
        existing_assignments = {assignment.role: assignment for assignment in user.role_assignments}

        for role, assignment in list(existing_assignments.items()):
            if role not in target_extra_roles:
                db.session.delete(assignment)

        for role in target_extra_roles:
            if role not in existing_assignments:
                db.session.add(UserRoleAssignment(user_id=user.id, role=role))

        db.session.commit()
        flash(f'Role user {user.username} berhasil diperbarui.', 'success')
        return redirect(url_for('admin.manage_user_roles', q=query))

    query = (request.args.get('q') or '').strip()
    users_query = User.query.filter(
        User.tenant_id == tenant_id,
        User.role != UserRole.ADMIN,
        User.role != UserRole.SUPER_ADMIN,
        ~User.role_assignments.any(role=UserRole.ADMIN),
        ~User.role_assignments.any(role=UserRole.SUPER_ADMIN),
    )
    if query:
        users_query = users_query.filter(
            or_(
                User.username.ilike(f'%{query}%'),
                User.email.ilike(f'%{query}%')
            )
        )

    users = users_query.order_by(User.username.asc()).all()
    return render_template(
        'admin/users/roles.html',
        users=users,
        query=query,
        all_roles=[role for role in UserRole if role not in {UserRole.ADMIN, UserRole.SUPER_ADMIN}],
        role_label=role_label
    )


@admin_bp.route('/users/reset-password-generic', methods=['POST'])
@login_required
@role_required(UserRole.ADMIN)
def generic_reset_password():
    """Route serbaguna untuk reset password via Modal"""
    user_id = request.form.get('user_id')
    new_password = request.form.get('new_password')

    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.manage_users'))

    user = User.query.filter_by(id=user_id, tenant_id=tenant_id).first_or_404()

    # Validasi sederhana
    if not new_password or len(new_password) < 4:
        flash('Password minimal 4 karakter.', 'danger')
        return redirect(url_for('admin.manage_users'))

    try:
        user.set_password(new_password)
        # Opsional: Paksa user ganti password lagi saat login nanti
        user.must_change_password = False
        db.session.commit()

        flash(f'Password untuk {user.username} ({user.role.value}) berhasil diubah.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal mereset password: {str(e)}', 'danger')

    return redirect(url_for('admin.manage_users'))


@admin_bp.route('/users/reset-officer-pin', methods=['POST'])
@login_required
@role_required(UserRole.ADMIN)
def reset_officer_pin():
    user_id = request.form.get('user_id', type=int)
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.manage_users'))

    user = User.query.filter_by(id=user_id, tenant_id=tenant_id).first_or_404()
    if not user.has_role(UserRole.TU, UserRole.WALI_ASRAMA):
        flash('User ini bukan petugas tabungan (TU/Wali Asrama).', 'warning')
        return redirect(url_for('admin.manage_users'))

    try:
        user.withdrawal_pin_hash = None
        user.withdrawal_pin_failed_attempts = 0
        user.withdrawal_pin_locked_until = None
        db.session.commit()
        flash(f'PIN petugas untuk {user.username} berhasil direset.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal reset PIN petugas: {str(e)}', 'danger')

    return redirect(url_for('admin.manage_users'))


@admin_bp.route('/users/reset-student-pin', methods=['POST'])
@login_required
@role_required(UserRole.ADMIN)
def reset_student_pin():
    user_id = request.form.get('user_id', type=int)
    tenant_id = _current_tenant_id()
    if tenant_id is None:
        flash('Tenant default tidak ditemukan.', 'danger')
        return redirect(url_for('admin.manage_users'))

    user = User.query.filter_by(id=user_id, tenant_id=tenant_id).first_or_404()
    if not user.student_profile:
        flash('User ini bukan akun santri.', 'warning')
        return redirect(url_for('admin.manage_users'))

    account = StudentSavingsAccount.query.filter_by(
        tenant_id=tenant_id,
        student_id=user.student_profile.id
    ).first()
    if not account:
        flash('Akun tabungan santri belum ada.', 'warning')
        return redirect(url_for('admin.manage_users'))

    try:
        account.pin_hash = None
        account.pin_failed_attempts = 0
        account.pin_locked_until = None
        db.session.commit()
        flash(f'PIN tabungan santri untuk {user.username} berhasil direset.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal reset PIN santri: {str(e)}', 'danger')

    return redirect(url_for('admin.manage_users'))

